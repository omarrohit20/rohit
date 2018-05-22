import os, logging, sys, json, csv
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Color, PatternFill, Font, Border
from pymongo import MongoClient
from multiprocessing.dummy import Pool as ThreadPool

import quandl, math, time
import pandas as pd
import numpy as np
from talib.abstract import *

import datetime
import time
import gc

connection = MongoClient('localhost', 27017)
db = connection.Nsedata

directory = '../../output/final'
logname = '../../output/final' + '/news' + time.strftime("%d%m%y-%H%M%S")

newsDict = {}
wb = Workbook()
ws = wb.active
ws.append(["scrip", "timestamps", "summary", "Link", "MLIndicator"])
ws_buyNews = wb.create_sheet("BuyNews")
ws_buyNews.append(["timestamps", "summary", "Link", "MLIndicator"])
ws_sellNews = wb.create_sheet("SellNews")
ws_sellNews.append(["timestamps", "summary", "Link", "MLIndicator"])


def saveDailyNews():
    for newslink,newsValue in newsDict.items():
        ws.append([newsValue['scrip'], newsValue['newstime'], newsValue['newssummary'], newslink, newsValue['mlindicator']])

def saveReports():
    # Add a default style with striped rows and banded columns
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
               showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    
    count = 0
    for row in ws.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:E" + str(count))
    tab.tableStyleInfo = style
    ws.add_table(tab)
    
    count = 0
    for row in ws_buyNews.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:C" + str(count))
    tab.tableStyleInfo = style
    ws_buyNews.add_table(tab)
    
    count = 0
    for row in ws_sellNews.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:C" + str(count))
    tab.tableStyleInfo = style
    ws_sellNews.add_table(tab)
    
    wb.save(logname + ".xlsx")
        
def buy_News(scrip):
    scrip_newsList = db.news.find_one({'scrip':scrip})
    ws_buyNews.append([scrip])
    ws_buyNews.append(["#####################"])
    if(scrip_newsList is None):
        print('Missing news for ', scrip)
        return
    
    for scrip_news in scrip_newsList['news']:
        ws_buyNews.append([scrip_news['timestamp'], scrip_news['summary'], scrip_news['link']])
    ws_buyNews.append([" "])
    
def sell_News(scrip):
    scrip_newsList = db.news.find_one({'scrip':scrip})
    ws_sellNews.append([scrip])
    ws_sellNews.append(["#####################"])
    if(scrip_newsList is None):
        print('Missing news for ', scrip)
        return
    for scrip_news in scrip_newsList['news']:
        ws_sellNews.append([scrip_news['timestamp'], scrip_news['summary'], scrip_news['link']])
    ws_sellNews.append([" "])  

def result_data(scrip):
    
    
    regression_data_high = db.regressionhigh.find_one({'scrip':scrip.replace('&','').replace('-','_')})
    if(regression_data_high['kNeighboursValue'] > 0 and regression_data_high['mlpValue'] > 0):
        if(regression_data_high is None):
            print('Missing or very less Data for ', scrip)
        else:
            buy_News(scrip)
            
    regression_data_low = db.regressionlow.find_one({'scrip':scrip.replace('&','').replace('-','_')})
    if(regression_data_low['kNeighboursValue'] < 0 and regression_data_low['mlpValue'] < 0):
        if(regression_data_low is None):
            print('Missing or very less Data for ', scrip)
        else:
            sell_News(scrip)        
         
    start_date = (datetime.datetime.now() - datetime.timedelta(hours=0))
    start_date = datetime.datetime(start_date.year, start_date.month, start_date.day, start_date.hour)
    end_date = (datetime.datetime.now() - datetime.timedelta(hours=19))
    #end_date = (datetime.datetime.now() - datetime.timedelta(hours=67))
    end_date = datetime.datetime(end_date.year, end_date.month, end_date.day, end_date.hour)
    
    scrip_newsList = db.news.find_one({'scrip':scrip})
    if(scrip_newsList is None):
        return
    for news in scrip_newsList['news']:
        newstime = news['timestamp']
        newssummary = news['summary']
        newslink = news['link']
        try:
            news_time = datetime.datetime.strptime(newstime, "%H:%M:%S %d-%m-%Y")
            if start_date > news_time > end_date: 
                if newslink in newsDict:
                    newsDict[newslink]['scrip'] = newsDict[newslink]['scrip'] + ',' + scrip
                    if((regression_data_high['kNeighboursValue'] >= 1) or (regression_data_high['mlpValue'] >= 2 and regression_data_high['kNeighboursValue'] >= 0) 
                        or (regression_data_high['mlpValue'] >= 0 and regression_data_high['kNeighboursValue'] >= 0.5)):
                        newsDict[newslink]['mlindicator'] = newsDict[newslink]['mlindicator'] + ',' + 'Buy:' + scrip
                    if((regression_data_low['kNeighboursValue'] <= -1) or (regression_data_low['mlpValue'] <= -2 and regression_data_low['kNeighboursValue'] <= 0)
                        or (regression_data_low['mlpValue'] <= 0 and regression_data_low['kNeighboursValue'] <= -0.5)):
                        newsDict[newslink]['mlindicator'] = newsDict[newslink]['mlindicator'] + ',' + 'Sell:' + scrip    
                else:
                    newsValue = {}
                    newsValue['newssummary'] = newssummary
                    newsValue['newstime'] = newstime
                    newsValue['scrip'] = scrip
                    newsValue['mlindicator'] = ""
                    if((regression_data_high['kNeighboursValue'] >= 1) or (regression_data_high['mlpValue'] >= 2 and regression_data_high['kNeighboursValue'] >= 0) 
                        or (regression_data_high['mlpValue'] >= 0 and regression_data_high['kNeighboursValue'] >= 0.5)):
                        newsValue['mlindicator'] = 'Buy:' + scrip
                    if((regression_data_low['kNeighboursValue'] <= -1) or (regression_data_low['mlpValue'] <= -2 and regression_data_low['kNeighboursValue'] <= 0)
                        or (regression_data_low['mlpValue'] <= 0 and regression_data_low['kNeighboursValue'] <= -0.5)):
                        newsValue['mlindicator'] = newsValue['mlindicator'] + ',' + 'Sell:' + scrip    
                    newsDict[newslink] = newsValue
                
                
        except:
            pass
                              
def calculateParallel(threads=2):
    pool = ThreadPool(threads)
    scrips = []
    for data in db.scrip.find({'futures':'Yes'}):
        scrips.append(data['scrip'].replace('&','').replace('-','_'))
    scrips.sort()
    pool.map(result_data, scrips)
        
                     
if __name__ == "__main__":
    if not os.path.exists(directory):
        os.makedirs(directory)
    calculateParallel(1)
    connection.close()
    saveDailyNews()
    saveReports()