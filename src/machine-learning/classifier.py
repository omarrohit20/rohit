import os, logging, sys, json
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Color, PatternFill, Font, Border
from pymongo import MongoClient
from multiprocessing.dummy import Pool as ThreadPool

import quandl, math, time
import pandas as pd
pd.options.mode.chained_assignment = None  # default='warn'
import numpy as np
from talib.abstract import *
#from pip.req.req_file import preprocess
from Algorithms.regression_helpers import load_dataset, addFeatures, addFeaturesVolChange, \
    addFeaturesOpenChange, addFeaturesHighChange, addFeaturesLowChange, addFeaturesEMA9Change, addFeaturesEMA21Change, \
    mergeDataframes, count_missing, applyTimeLag, performClassification   
    
from technical import ta_lib_data  

from sklearn.ensemble import RandomForestRegressor
from sklearn.ensemble import BaggingRegressor
from sklearn.ensemble import AdaBoostRegressor
from sklearn.ensemble import GradientBoostingRegressor
from sklearn.neighbors import KNeighborsRegressor
from sklearn.neural_network import MLPRegressor
from sklearn.ensemble import RandomForestClassifier
from sklearn import neighbors
from sklearn.ensemble import AdaBoostClassifier
from sklearn.ensemble import GradientBoostingClassifier
from sklearn.neural_network import MLPClassifier
#from sklearn.svm import SVR
from sklearn.feature_selection import SelectKBest, chi2
from sklearn.svm import SVC, SVR
#from sklearn.qda import QDA
from sklearn.grid_search import GridSearchCV

import time
import gc

connection = MongoClient('localhost', 27017)
db = connection.Nsedata

directory = '../../output/classifier' + '/' + time.strftime("%d%m%y-%H%M%S")
logname = '../../output/classifier' + '/mllog' + time.strftime("%d%m%y-%H%M%S")
logging.basicConfig(filename=logname, filemode='a', stream=sys.stdout, level=logging.INFO)
log = logging.getLogger(__name__)

forecast_out = 1
randomForest = True
mlp = True
bagging = True
adaBoost = False
kNeighbours = True
gradientBoosting = False

wb = Workbook()
ws = wb.active
ws.append(["futures", "train set","BuyIndicators", "SellIndicators","Symbol", "VOL_change", "PCT_change", "Score","RandomForest", "accuracy", "MLP", "accuracy", "Bagging", "accuracy", "AdaBoost", "accuracy", "KNeighbors", "accuracy", "GradientBoosting", "accuracy", "trend", "yHighChange","yLowChange"])
ws_filter = wb.create_sheet("Filter")
ws_filter.append(["futures", "train set","BuyIndicators", "SellIndicators","Symbol", "VOL_change", "PCT_change", "Score","RandomForest", "accuracy", "MLP", "accuracy", "Bagging", "accuracy", "AdaBoost", "accuracy", "KNeighbors", "accuracy", "GradientBoosting", "accuracy", "trend", "yHighChange","yLowChange"])
ws_gtltzero = wb.create_sheet("FilterAllgtlt0")
ws_gtltzero.append(["futures", "train set","BuyIndicators", "SellIndicators","Symbol", "VOL_change", "PCT_change", "Score","RandomForest", "accuracy", "MLP", "accuracy", "Bagging", "accuracy", "AdaBoost", "accuracy", "KNeighbors", "accuracy", "GradientBoosting", "accuracy", "trend", "yHighChange","yLowChange"])

if randomForest:
    ws_RandomForest = wb.create_sheet("RandomForest")
    ws_RandomForest.append(["futures", "train set","BuyIndicators", "SellIndicators","Symbol", "VOL_change", "PCT_change", "Score","RandomForest", "accuracy", "MLP", "accuracy", "Bagging", "accuracy", "AdaBoost", "accuracy", "KNeighbors", "accuracy", "GradientBoosting", "accuracy", "trend", "yHighChange","yLowChange"])

if mlp:    
    ws_SVR = wb.create_sheet("MLP")
    ws_SVR.append(["futures", "train set","BuyIndicators", "SellIndicators","Symbol", "VOL_change", "PCT_change", "Score","RandomForest", "accuracy", "MLP", "accuracy", "Bagging", "accuracy", "AdaBoost", "accuracy", "KNeighbors", "accuracy", "GradientBoosting", "accuracy", "trend", "yHighChange","yLowChange"])

if bagging:    
    ws_Bagging = wb.create_sheet("Bagging")
    ws_Bagging.append(["futures", "train set","BuyIndicators", "SellIndicators","Symbol", "VOL_change", "PCT_change", "Score","RandomForest", "accuracy", "MLP", "accuracy", "Bagging", "accuracy", "AdaBoost", "accuracy", "KNeighbors", "accuracy", "GradientBoosting", "accuracy", "trend", "yHighChange","yLowChange"])

if adaBoost:    
    ws_AdaBoost = wb.create_sheet("AdaBoost")
    ws_AdaBoost.append(["futures", "train set","BuyIndicators", "SellIndicators","Symbol", "VOL_change", "PCT_change", "Score","RandomForest", "accuracy", "MLP", "accuracy", "Bagging", "accuracy", "AdaBoost", "accuracy", "KNeighbors", "accuracy", "GradientBoosting", "accuracy", "trend", "yHighChange","yLowChange"])
    
if kNeighbours:    
    ws_KNeighbors = wb.create_sheet("KNeighbors")
    ws_KNeighbors.append(["futures", "train set","BuyIndicators", "SellIndicators","Symbol", "VOL_change", "PCT_change", "Score","RandomForest", "accuracy", "MLP", "accuracy", "Bagging", "accuracy", "AdaBoost", "accuracy", "KNeighbors", "accuracy", "GradientBoosting", "accuracy", "trend", "yHighChange","yLowChange"])

if gradientBoosting:    
    ws_GradientBoosting = wb.create_sheet("GradientBoosting")
    ws_GradientBoosting.append(["futures", "train set","BuyIndicators", "SellIndicators","Symbol", "VOL_change", "PCT_change", "Score","RandomForest", "accuracy", "MLP", "accuracy", "Bagging", "accuracy", "AdaBoost", "accuracy", "KNeighbors", "accuracy", "GradientBoosting", "accuracy", "trend", "yHighChange","yLowChange"])

def insert_classificationdata(data):     
    json_data = json.loads(json.dumps(data))
    db.classification.insert_one(json_data)  

def getScore(vol_change, pct_change):
    try:
        return float(vol_change)/float(pct_change) 
    except ZeroDivisionError:
        return 0

def saveReports():
    # Add a default style with striped rows and banded columns
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
               showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    
    count = 0
    for row in ws.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:W" + str(count))
    tab.tableStyleInfo = style
    ws.add_table(tab)
    
    count = 0
    for row in ws_filter.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:W" + str(count))
    tab.tableStyleInfo = style
    ws_filter.add_table(tab)
    
    count = 0
    for row in ws_gtltzero.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:W" + str(count))
    tab.tableStyleInfo = style
    ws_gtltzero.add_table(tab)
    
    if randomForest:
        count = 0
        for row in ws_RandomForest.iter_rows(row_offset=1):
            count += 1
        tab = Table(displayName="Table1", ref="A1:W" + str(count))
        tab.tableStyleInfo = style
        ws_RandomForest.add_table(tab)
    
    if mlp:
        count = 0
        for row in ws_SVR.iter_rows(row_offset=1):
            count += 1
        tab = Table(displayName="Table1", ref="A1:W" + str(count))
        tab.tableStyleInfo = style
        ws_SVR.add_table(tab)
    
    if bagging:
        count = 0
        for row in ws_Bagging.iter_rows(row_offset=1):
            count += 1
        tab = Table(displayName="Table1", ref="A1:W" + str(count))
        tab.tableStyleInfo = style
        ws_Bagging.add_table(tab)
    
    if adaBoost:
        count = 0
        for row in ws_AdaBoost.iter_rows(row_offset=1):
            count += 1
        tab = Table(displayName="Table1", ref="A1:W" + str(count))
        tab.tableStyleInfo = style
        ws_AdaBoost.add_table(tab)
    
    if kNeighbours:
        count = 0
        for row in ws_KNeighbors.iter_rows(row_offset=1):
            count += 1
        tab = Table(displayName="Table1", ref="A1:W" + str(count))
        tab.tableStyleInfo = style
        ws_KNeighbors.add_table(tab)
    
    if gradientBoosting:
        count = 0
        for row in ws_GradientBoosting.iter_rows(row_offset=1):
            count += 1
        tab = Table(displayName="Table1", ref="A1:W" + str(count))
        tab.tableStyleInfo = style
        ws_GradientBoosting.add_table(tab)
    
    wb.save(logname + ".xlsx")

def historical_data(data):
    ardate = np.array([x.encode('UTF8') for x in (np.array(data['data'])[:,0][::-1]).tolist()])
    aropen = np.array([float(x.encode('UTF8')) for x in (np.array(data['data'])[:,1][::-1]).tolist()])
    arhigh = np.array([float(x.encode('UTF8')) for x in (np.array(data['data'])[:,2][::-1]).tolist()])
    arlow  = np.array([float(x.encode('UTF8')) for x in (np.array(data['data'])[:,3][::-1]).tolist()])
    arlast = np.array([float(x.encode('UTF8')) for x in (np.array(data['data'])[:,4][::-1]).tolist()])
    arclose= np.array([float(x.encode('UTF8')) for x in (np.array(data['data'])[:,5][::-1]).tolist()])
    arquantity = np.array([float(x.encode('UTF8')) for x in (np.array(data['data'])[:,6][::-1]).tolist()])
    arturnover = np.array([float(x.encode('UTF8')) for x in (np.array(data['data'])[:,7][::-1]).tolist()])
    return ardate, aropen, arhigh, arlow, arlast, arclose, arquantity, arturnover

def get_data_frame(df, regressor="None"):
    if (df is not None):
        df=df.rename(columns = {'total trade quantity':'volume'})
        df=df.rename(columns = {'turnover (lacs)': 'turnover'})
        df['PCT_day_change'] = (((df['close'] - df['open'])/df['open'])*100)
        df['HL_change'] = (((df['high'] - df['low'])/df['low'])*100).astype(int)
        df['volume_pre'] = df['volume'].shift(+1)
        df['close_pre'] = df['close'].shift(+1)
        #df.fillna(-99999, inplace=True)
        df.dropna(inplace=True)
        df['VOL_change'] = (((df['volume'] - df['volume_pre'])/df['volume_pre'])*100)
        df['PCT_change'] = (((df['close'] - df['close_pre'])/df['close_pre'])*100)
        df['EMA9'] = EMA(df,9)
        df['EMA21'] = EMA(df,21)
        
        dfp = df[['VOL_change']]
        maxdelta = 10
        columns = df.columns
        open = columns[1]
        high = columns[2]
        low = columns[3]
        close = columns[4]
        volume = columns[5]
        EMA9 = columns[-2]
        EMA21 = columns[-1]
        for dele in range(1, 11):
            addFeatures(df, dfp, close, dele)
        #if regressor == 'kNeighbours':
        if regressor != 'mlp':    
            for dele in range(1, 2):
                addFeaturesOpenChange(df, dfp, open, dele)    
                addFeaturesLowChange(df, dfp, low, dele) 
                addFeaturesHighChange(df, dfp, high, dele)
                addFeaturesEMA9Change(df, dfp, EMA9, dele)
                addFeaturesEMA21Change(df, dfp, EMA21, dele)
 
        if regressor != 'mlp':    
            dfp['ADX'] = ADX(df).apply(lambda x: 1 if x > 20 else 0) #Average Directional Movement Index http://www.investopedia.com/terms/a/adx.asp
            dfp['ADXR'] = ADXR(df).apply(lambda x: 1 if x > 20 else 0) #Average Directional Movement Index Rating https://www.scottrade.com/knowledge-center/investment-education/research-analysis/technical-analysis/the-indicators/average-directional-movement-index-rating-adxr.html
        dfp['APO'] = APO(df).apply(lambda x: 1 if x > 0 else 0) #Absolute Price Oscillator https://www.fidelity.com/learning-center/trading-investing/technical-analysis/technical-indicator-guide/apo
#         aroon = AROON(df) #Aroon http://www.investopedia.com/terms/a/aroon.asp
#         dfp['AROONUP'], dfp['AROONDOWN'] = aroon['aroonup'], aroon['aroondown']
        dfp['AROONOSC'] = AROONOSC(df).apply(lambda x: 1 if x > 0 else 0)
        dfp['BOP'] = BOP(df).apply(lambda x: 1 if x > 0 else 0) #Balance Of Power https://www.marketvolume.com/technicalanalysis/balanceofpower.asp
        if regressor == 'kNeighbours': 
            dfp['CCI'] = CCI(df) #Commodity Channel Index http://www.investopedia.com/articles/trading/05/041805.asp
#         dfp['CMO'] = CMO(df) #Chande Momentum Oscillator https://www.fidelity.com/learning-center/trading-investing/technical-analysis/technical-indicator-guide/cmo
#         dfp['DX'] = DX(df) #Directional Movement Index http://www.investopedia.com/terms/d/dmi.asp
#         macd = MACD(df)
#         dfp['MACD'], dfp['MACDSIGNAL'], dfp['MACDHIST'] = macd['macd'], macd['macdsignal'], macd['macdhist']
#         #dfp['MACDEXT'] = MACDEXT(df)
#         #dfp['MACDFIX'] = MACDFIX(df)
#         dfp['MFI'] = MFI(df)
#         dfp['MINUS_DI'] = MINUS_DI(df)
#         dfp['MINUS_DM'] = MINUS_DM(df)
#         dfp['MOM'] = MOM(df)
#         dfp['PLUS_DI'] = PLUS_DI(df)
#         dfp['PLUS_DM'] = PLUS_DM(df)
#         dfp['PPO'] = PPO(df)
#         dfp['ROC'] = ROC(df)
#         dfp['ROCP'] = ROCP(df)
#         dfp['ROCR'] = ROCR(df)
#         dfp['ROCR100'] = ROCR100(df)
        if regressor == 'kNeighbours':
            dfp['RSI'] = RSI(df)
        #dfp['STOCH'] = STOCH(df)
        #dfp['STOCHF'] = STOCHF(df)
        #dfp['STOCHRSI'] = STOCHRSI(df)
#         dfp['TRIX'] = TRIX(df)
#         dfp['ULTOSC'] = ULTOSC(df)
#         dfp['WILLR'] = WILLR(df)
#         
#         bbands = BBANDS(df)
#         dfp['BBANDSUPPER'], dfp['BBANDSMIDDLE'], dfp['BBANDSLOWER'] = bbands['upperband'], bbands['middleband'], bbands['lowerband']
#         
# #       dfp['DEMA'] = DEMA(df)
#         dfp['EMA9'] = EMA(df,9)
#         dfp['EMA21'] = EMA(df,21)
#         dfp['EMA25'] = EMA(df,25)
#         dfp['EMA50'] = EMA(df,50)
#         dfp['EMA100'] = EMA(df,100)
#         dfp['EMA200'] = EMA(df,200)
#         dfp['HT_TRENDLINE'] = HT_TRENDLINE(df)
#         dfp['KAMA'] = KAMA(df)
#         dfp['MA'] = MA(df)
#         #dfp['MAMA'] = MAMA(df)
#         #dfp['MAVP'] = MAVP(df)
#         dfp['MIDPOINT'] = MIDPOINT(df)
#         dfp['MIDPRICE'] = MIDPRICE(df)
#         dfp['SAR'] = SAR(df)
#         dfp['SAREXT'] = SAREXT(df)
#         dfp['SMA'] = SMA(df)
#         dfp['SMA9'] = SMA(df, 9)
#         dfp['T3'] = T3(df)
#         dfp['TEMA'] = TEMA(df)
#         dfp['TRIMA'] = TRIMA(df)
#         dfp['WMA'] = WMA(df)
 
 
        dfp['CDL2CROWS'] = CDL2CROWS(df)
        dfp['CDL3BLACKCROWS'] = CDL3BLACKCROWS(df)
        dfp['CDL3INSIDE'] = CDL3INSIDE(df)
        dfp['CDL3LINESTRIKE'] = CDL3LINESTRIKE(df)
        dfp['CDL3OUTSIDE'] = CDL3OUTSIDE(df)
        dfp['CDL3STARSINSOUTH'] = CDL3STARSINSOUTH(df)
        dfp['CDL3WHITESOLDIERS'] = CDL3WHITESOLDIERS(df)
        dfp['CDLABANDONEDBABY'] = CDLABANDONEDBABY(df)
        dfp['CDLADVANCEBLOCK'] = CDLADVANCEBLOCK(df) #Bearish reversal. prior trend Upward. Look for three white candles in an upward price trend. On each candle, price opens within the body of the previous candle. The height of the shadows grow taller on the last two candles.
        dfp['CDLBELTHOLD'] = CDLBELTHOLD(df) # Bearish reversal. prior trend upward. Price opens at the high for the day and closes near the low, forming a tall black candle, often with a small lower shadow.
        dfp['CDLBREAKAWAY'] = CDLBREAKAWAY(df) # Bearish reversal. prior trend upward. Look for 5 candle lines in an upward price trend with the first candle being a tall white one. The second day should be a white candle with a gap between the two bodies, but the shadows can overlap. Day three should have a higher close and the candle can be any color. Day 4 shows a white candle with a higher close. The last day is a tall black candle with a close within the gap between the bodies of the first two candles.
        dfp['CDLCLOSINGMARUBOZU'] = CDLCLOSINGMARUBOZU(df)
        dfp['CDLCONCEALBABYSWALL'] = CDLCONCEALBABYSWALL(df)
        dfp['CDLCOUNTERATTACK'] = CDLCOUNTERATTACK(df)
        dfp['CDLDARKCLOUDCOVER'] = CDLDARKCLOUDCOVER(df)
        dfp['CDLDOJI'] = CDLDOJI(df)
        dfp['CDLDOJISTAR'] = CDLDOJISTAR(df)
        dfp['CDLDRAGONFLYDOJI'] = CDLDRAGONFLYDOJI(df)
        dfp['CDLENGULFING'] = CDLENGULFING(df)
        dfp['CDLEVENINGDOJISTAR'] = CDLEVENINGDOJISTAR(df)
        dfp['CDLEVENINGSTAR'] = CDLEVENINGSTAR(df)
        dfp['CDLGAPSIDESIDEWHITE'] = CDLGAPSIDESIDEWHITE(df)
        dfp['CDLGRAVESTONEDOJI'] = CDLGRAVESTONEDOJI(df)
        dfp['CDLHAMMER'] = CDLHAMMER(df)
        dfp['CDLHANGINGMAN'] = CDLHANGINGMAN(df)
        dfp['CDLHARAMI'] = CDLHARAMI(df)
        dfp['CDLHARAMICROSS'] = CDLHARAMICROSS(df)
        dfp['CDLHIGHWAVE'] = CDLHIGHWAVE(df)
        dfp['CDLHIKKAKE'] = CDLHIKKAKE(df)
        dfp['CDLHIKKAKEMOD'] = CDLHIKKAKEMOD(df)
        dfp['CDLHOMINGPIGEON'] = CDLHOMINGPIGEON(df)
        dfp['CDLIDENTICAL3CROWS'] = CDLIDENTICAL3CROWS(df)
        dfp['CDLINNECK'] = CDLINNECK(df)
        dfp['CDLINVERTEDHAMMER'] = CDLINVERTEDHAMMER(df)
        dfp['CDLKICKING'] = CDLKICKING(df)
        dfp['CDLKICKINGBYLENGTH'] = CDLKICKINGBYLENGTH(df)
        dfp['CDLLADDERBOTTOM'] = CDLLADDERBOTTOM(df)
        dfp['CDLLONGLEGGEDDOJI'] = CDLLONGLEGGEDDOJI(df)
        dfp['CDLLONGLINE'] = CDLLONGLINE(df)
        dfp['CDLMARUBOZU'] = CDLMARUBOZU(df)
        dfp['CDLMATCHINGLOW'] = CDLMATCHINGLOW(df)
        dfp['CDLMATHOLD'] = CDLMATHOLD(df)
        dfp['CDLMORNINGDOJISTAR'] = CDLMORNINGDOJISTAR(df)
        dfp['CDLMORNINGSTAR'] = CDLMORNINGSTAR(df)
        dfp['CDLONNECK'] = CDLONNECK(df)
        dfp['CDLPIERCING'] = CDLPIERCING(df)
        dfp['CDLRICKSHAWMAN'] = CDLRICKSHAWMAN(df)
        dfp['CDLRISEFALL3METHODS'] = CDLRISEFALL3METHODS(df)
        dfp['CDLSEPARATINGLINES'] = CDLSEPARATINGLINES(df)
        dfp['CDLSHOOTINGSTAR'] = CDLSHOOTINGSTAR(df)
        dfp['CDLSHORTLINE'] = CDLSHORTLINE(df)
        dfp['CDLSPINNINGTOP'] = CDLSPINNINGTOP(df)
        dfp['CDLSTALLEDPATTERN'] = CDLSTALLEDPATTERN(df)
        dfp['CDLSTICKSANDWICH'] = CDLSTICKSANDWICH(df)
        dfp['CDLTAKURI'] = CDLTAKURI(df)
        dfp['CDLTASUKIGAP'] = CDLTASUKIGAP(df)
        dfp['CDLTHRUSTING'] = CDLTHRUSTING(df)
        dfp['CDLTRISTAR'] = CDLTRISTAR(df)
        dfp['CDLUNIQUE3RIVER'] = CDLUNIQUE3RIVER(df)
        dfp['CDLUPSIDEGAP2CROWS'] = CDLUPSIDEGAP2CROWS(df)
        dfp['CDLXSIDEGAP3METHODS'] = CDLXSIDEGAP3METHODS(df)

#         dfp['AVGPRICE'] = AVGPRICE(df)
#         dfp['MEDPRICE'] = MEDPRICE(df)
#         dfp['TYPPRICE'] = TYPPRICE(df)
#         dfp['WCLPRICE'] = WCLPRICE(df)
# 
# #         dfp['ATR'] = ATR(df)
# #         dfp['NATR'] = NATR(df)
# #         dfp['TRANGE'] = TRANGE(df)
#         
#        dfp['AD'] = AD(df)
#        dfp['ADOSC'] = ADOSC(df)
#        dfp['OBV'] = OBV(df)
        
        forecast_col = 'PCT_change1'
        dfp.fillna(-99999, inplace=True)
        dfp['label'] = dfp[forecast_col].shift(-forecast_out)
        
#         X = np.array(dfp.drop(['label'], 1))
#         X = preprocessing.scale(X)
#         X = X[:-forecast_out]
#         X_lately = X[-forecast_out:]
#         
#         #dfp.dropna(inplace=True)
#         y = np.array(dfp['label'])
#         
#         X_train, X_test, y_train, y_test = cross_validation.train_test_split(X, y, test_size=0.1)
#         clf = LinearRegression()
#         #clf = svm.SVR(kernel='poly')
#         clf.fit(X_train, y_train)
#         accuracy = clf.score(X_test, y_test)
#         forecast_set = clf.predict(X_lately)
#         print(scrip, accuracy, forecast_set)
#         #print('XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX')     
        dfp = dfp.ix[50:]
        dfp.drop(dfp[dfp.label > 7].index, inplace=True)
        dfp.drop(dfp[dfp.label < -7].index, inplace=True)
    return dfp

def create_csv(regressionResult):
    ws.append(regressionResult)
    trainSize = int(regressionResult[1])
    buyIndia = str(regressionResult[2])
    sellIndia = str(regressionResult[3])
    scrip = str(regressionResult[4])
    forecast_day_VOL_change = int(regressionResult[5])
    forecast_day_PCT_change = float(regressionResult[6])
    score = float(regressionResult[7])
    randomForestValue = float(regressionResult[8])
    randomForestAccuracy = float(regressionResult[9])
    mlpValue = float(regressionResult[10])
    mlpAccuracy = float(regressionResult[11])
    baggingValue = float(regressionResult[12])
    baggingAccuracy = float(regressionResult[13])
    adaBoostValue = float(regressionResult[14])
    adaBoostAccuracy = float(regressionResult[15])
    kNeighboursValue = float(regressionResult[16])
    kNeighboursAccuracy = float(regressionResult[17])
    gradientBoostingValue = float(regressionResult[18])
    gradientBoostingAccuracy = float(regressionResult[19])
    trend = str(regressionResult[20])
    yearHighChange = float(regressionResult[21])
    yearLowChange = float(regressionResult[22])
    
    
    if mlp and kNeighbours:
        #ws_filter = wb.create_sheet("Filter")
        if((trainSize> 1000) and (kNeighboursValue > 0) and (mlpValue > 0) and len(sellIndia) < 2):
            ws_filter.append(regressionResult)     
        elif((trainSize> 1000) and (kNeighboursValue < 0) and (mlpValue < 0) and len(buyIndia) < 2):
            ws_filter.append(regressionResult)
               
    if mlp and kNeighbours:    
        #ws_gtltzero = wb.create_sheet("FilterAllgtlt0")
        if((trainSize> 1000) and (randomForestValue == 0) and (kNeighboursValue == 0) and (mlpValue == 0) and (baggingValue == 0)):
            #do Nothing
            x=None
        elif((trainSize> 1000) and (randomForestValue >= 0) and (kNeighboursValue >= 0) and (mlpValue >= 0) and (baggingValue >= 0)):
            ws_gtltzero.append(regressionResult)           
        elif((trainSize> 1000) and (randomForestValue <= 0) and (kNeighboursValue <= 0) and (mlpValue <= 0) and (baggingValue <= 0)):
            ws_gtltzero.append(regressionResult)   
    
    if randomForest:    
        #ws_RandomForest = wb.create_sheet("RandomForest")
        if((trainSize> 1000) and (randomForestValue > 0) and (kNeighboursValue >= 0) and (mlpValue >= 0) and (baggingValue >= 0) and (randomForestAccuracy > .5)):
            ws_RandomForest.append(regressionResult)  
        elif((trainSize> 1000) and (randomForestValue > 0) and (kNeighboursValue > 0) and (mlpValue >= 0) and (baggingValue >= 0)):
            ws_RandomForest.append(regressionResult)    
        elif((trainSize> 1000) and (randomForestValue < 0) and (kNeighboursValue <= 0) and (mlpValue <= 0) and (baggingValue <= 0) and (randomForestAccuracy > .5)):
            ws_RandomForest.append(regressionResult) 
        elif((trainSize> 1000) and (randomForestValue < 0) and (kNeighboursValue < 0) and (mlpValue <= 0) and (baggingValue <= 0)):
            ws_RandomForest.append(regressionResult)          
    
    if mlp:    
        #ws_SVR = wb.create_sheet("MLP")
        if((trainSize> 1000) and (randomForestValue >= 0) and (kNeighboursValue >= 0) and (mlpValue > 0) and (baggingValue >= 0) and (mlpAccuracy > .35)):
            ws_SVR.append(regressionResult) 
            
        elif((trainSize> 1000) and (randomForestValue <= 0) and (kNeighboursValue <= 0) and (mlpValue < 0) and (baggingValue <= 0) and (mlpAccuracy > .35)):
            ws_SVR.append(regressionResult)        
    
    if bagging:       
        #ws_Bagging = wb.create_sheet("Bagging")
        if((trainSize> 1000) and (randomForestValue >= 0) and (kNeighboursValue >= 0) and (mlpValue >= 0) and (baggingValue > 0) and (baggingAccuracy > .5)):
            ws_Bagging.append(regressionResult)  
            
        elif((trainSize> 1000) and (randomForestValue <= 0) and (kNeighboursValue <= 0) and (mlpValue <= 0) and (baggingValue < 0) and (baggingAccuracy > .5)):
            ws_Bagging.append(regressionResult)       
        
    if adaBoost:    
        #ws_AdaBoost = wb.create_sheet("AdaBoost")
        if((trainSize> 1000) and (adaBoostValue > 0) and (adaBoostAccuracy > .5)):
            ws_AdaBoost.append(regressionResult)
        if((trainSize> 1000) and (adaBoostValue < 0) and (adaBoostAccuracy > .5)):
            ws_AdaBoost.append(regressionResult)    
    
    if kNeighbours:    
        #ws_KNeighbors = wb.create_sheet("KNeighbors")
        if((trainSize> 1000) and (randomForestValue <= 0) and (kNeighboursValue < 0) and (mlpValue <= 0) and (baggingValue <= 0) and (kNeighboursAccuracy > .5)):
            ws_KNeighbors.append(regressionResult) 
        elif((trainSize> 1000) and (randomForestValue >= 0) and (kNeighboursValue > 0) and (mlpValue >= 0) and (baggingValue >= 0) and (kNeighboursAccuracy > .5)):
            ws_KNeighbors.append(regressionResult)        
        
    if gradientBoosting:    
        #ws_GradientBoosting = wb.create_sheet("GradientBoosting")
        if((trainSize> 1000) and (gradientBoostingValue > 0) and score > 0):
            ws_GradientBoosting.append(regressionResult)
        if((trainSize> 1000) and (gradientBoostingValue < 0) and score < 0):
            ws_GradientBoosting.append(regressionResult) 
            
    #Insert in db
    data = {}
    data['trainSize'] = trainSize
    data['buyIndia'] = buyIndia
    data['sellIndia'] = sellIndia
    data['scrip'] = scrip
    data['forecast_day_VOL_change'] = forecast_day_VOL_change
    data['forecast_day_PCT_change'] = forecast_day_PCT_change
    data['score'] = score
    data['randomForestValue'] = randomForestValue
    data['randomForestAccuracy'] = randomForestAccuracy
    data['mlpValue'] = mlpValue
    data['mlpAccuracy'] = mlpAccuracy
    data['baggingValue'] = baggingValue
    data['baggingAccuracy'] = baggingAccuracy
    data['adaBoostValue'] = adaBoostValue
    data['adaBoostAccuracy'] = adaBoostAccuracy
    data['kNeighboursValue'] = kNeighboursValue
    data['kNeighboursAccuracy'] = kNeighboursAccuracy
    data['gradientBoostingValue'] = gradientBoostingValue 
    data['gradientBoostingAccuracy'] = gradientBoostingAccuracy 
    data['trend'] = trend 
    data['yearHighChange'] = yearHighChange 
    data['yearLowChange'] = yearLowChange 
    insert_classificationdata(data)   

def regression_ta_data(scrip):
    data = db.history.find_one({'dataset_code':scrip})
    if(data is None or (np.array(data['data'])).size < 200):
        print('Missing or very less Data for ', scrip)
        return
        
    hsdate, hsopen, hshigh, hslow, hslast, hsclose, hsquantity, hsturnover = historical_data(data)   
    df = pd.DataFrame({
        'date': hsdate,
        'open': hsopen,
        'high': hshigh,
        'low': hslow,
        'close': hsclose,
        'volume': hsquantity,
        'turnover':hsturnover
    })
    df = df[['date','open','high','low','close','volume','turnover']]
    print(scrip)
    dfp = get_data_frame(df)
    #dfp.to_csv(directory + '/' + scrip + '.csv', encoding='utf-8')
    forecast_day_PCT_change = dfp.iloc[-forecast_out:, 1].values[0]
    forecast_day_VOL_change = dfp.iloc[-forecast_out:, 0].values[0]
    score = getScore(forecast_day_VOL_change, forecast_day_PCT_change) 
    buy, sell, trend, yearHighChange, yearLowChange = ta_lib_data(scrip) 
    trainSize = int((df.shape)[0])
    
    regressionResult = [ ]
    regressionResult.append(data['futures'])
    regressionResult.append(str(trainSize))
    regressionResult.append(str(buy))
    regressionResult.append(str(sell))
    regressionResult.append(str(scrip))
    regressionResult.append(forecast_day_VOL_change)
    regressionResult.append(forecast_day_PCT_change)
    regressionResult.append(score)
      
    if randomForest:
        regressionResult.extend(performClassification(dfp, 0.98, scrip, directory, forecast_out, RandomForestClassifier(n_estimators=10, n_jobs=1), True))
    else: 
        regressionResult.extend([0,0])    
            
    if mlp:
        regressionResult.extend(performClassification(get_data_frame(df, 'mlp'), 0.98, scrip, directory, forecast_out, MLPClassifier(activation='tanh', solver='adam', max_iter=500, hidden_layer_sizes=(84, 40, 10))))
    else:
        regressionResult.extend([0,0])
        
    if bagging:
        regressionResult.extend(performClassification(dfp, 0.98, scrip, directory, forecast_out, SVC(C=1.0, cache_size=200, class_weight=None, coef0=0.0,
                                                                                                     decision_function_shape='ovr', degree=3, gamma='auto', kernel='sigmoid'), True))
    else:
        regressionResult.extend([0,0])
        
    if adaBoost:
        regressionResult.extend(performClassification(dfp, 0.98, scrip, directory, forecast_out, AdaBoostClassifier()))
    else:
        regressionResult.extend([0,0])
        
    if kNeighbours:
        #dfp_kneighbour = get_data_frame(df, 'kNeighbours')
        regressionResult.extend(performClassification(get_data_frame(df, 'kNeighbours'), 0.98, scrip, directory, forecast_out, neighbors.KNeighborsClassifier(n_jobs=1), True))
    else:
        regressionResult.extend([0,0])
        
    if gradientBoosting:
        regressionResult.extend(performClassification(dfp, 0.98, scrip, directory, forecast_out, GradientBoostingClassifier()))
    else:
        regressionResult.extend([0,0])
    
    regressionResult.append(trend)
    regressionResult.append(yearHighChange)
    regressionResult.append(yearLowChange)
    create_csv(regressionResult)   
                                                          
def calculateParallel(threads=2):
    pool = ThreadPool(threads)
    
    scrips = []
    for data in db.scrip.find():
        scrips.append(data['scrip'].replace('&','').replace('-','_'))
    scrips.sort()
    
    pool.map(regression_ta_data, scrips)
    pool.map(regression_ta_data, scrips)
                     
if __name__ == "__main__":
    if not os.path.exists(directory):
        os.makedirs(directory)
    calculateParallel(1)
    connection.close()
    saveReports()