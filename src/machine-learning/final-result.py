import os, logging, sys, json, csv
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Color, PatternFill, Font, Border
from pymongo import MongoClient
from multiprocessing.dummy import Pool as ThreadPool

import quandl, math, time
import pandas as pd
import numpy as np
from talib.abstract import *

import datetime
import time
import gc

connection = MongoClient('localhost', 27017)
db = connection.Nsedata

directory = '../../output/final'
logname = '../../output/final' + '/final-result' + time.strftime("%d%m%y-%H%M%S")

newsDict = {}
wb = Workbook()
ws_buyAll = wb.create_sheet("BuyAll")
ws_buyAll.append(["futures", "train set","BuyIndicators", "SellIndicators","Symbol", "VOL_change", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_DAY", "Score","RandomForest", "accuracy", "MLP", "accuracy", "Bagging", "accuracy", "AdaBoost", "accuracy", "KNeighbors", "accuracy", "GradientBoosting", "accuracy", "trend", "yHighChange","yLowChange"])
ws_sellAll = wb.create_sheet("SellAll")
ws_sellAll.append(["futures", "train set","BuyIndicators", "SellIndicators","Symbol", "VOL_change", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_DAY", "Score","RandomForest", "accuracy", "MLP", "accuracy", "Bagging", "accuracy", "AdaBoost", "accuracy", "KNeighbors", "accuracy", "GradientBoosting", "accuracy", "trend", "yHighChange","yLowChange"])
ws_buyFinal = wb.create_sheet("BuyFinal")
ws_buyFinal.append(["futures", "train set","BuyIndicators", "SellIndicators","Symbol", "VOL_change", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_DAY", "Score","RandomForest", "accuracy", "MLP", "accuracy", "Bagging", "accuracy", "AdaBoost", "accuracy", "KNeighbors", "accuracy", "GradientBoosting", "accuracy", "trend", "yHighChange","yLowChange"])
ws_sellFinal = wb.create_sheet("SellFinal")
ws_sellFinal.append(["futures", "train set","BuyIndicators", "SellIndicators","Symbol", "VOL_change", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_DAY", "Score","RandomForest", "accuracy", "MLP", "accuracy", "Bagging", "accuracy", "AdaBoost", "accuracy", "KNeighbors", "accuracy", "GradientBoosting", "accuracy", "trend", "yHighChange","yLowChange"])
ws_buyFinal1 = wb.create_sheet("BuyFinal1")
ws_buyFinal1.append(["futures", "train set","BuyIndicators", "SellIndicators","Symbol", "VOL_change", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_DAY", "Score","RandomForest", "accuracy", "MLP", "accuracy", "Bagging", "accuracy", "AdaBoost", "accuracy", "KNeighbors", "accuracy", "GradientBoosting", "accuracy", "trend", "yHighChange","yLowChange"])
ws_sellFinal1 = wb.create_sheet("SellFinal1")
ws_sellFinal1.append(["futures", "train set","BuyIndicators", "SellIndicators","Symbol", "VOL_change", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_DAY", "Score","RandomForest", "accuracy", "MLP", "accuracy", "Bagging", "accuracy", "AdaBoost", "accuracy", "KNeighbors", "accuracy", "GradientBoosting", "accuracy", "trend", "yHighChange","yLowChange"])
ws_buy = wb.create_sheet("Buy")
ws_buy.append(["futures", "train set","BuyIndicators", "SellIndicators","Symbol", "VOL_change", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_DAY", "Score","RandomForest", "accuracy", "MLP", "accuracy", "Bagging", "accuracy", "AdaBoost", "accuracy", "KNeighbors", "accuracy", "GradientBoosting", "accuracy", "trend", "yHighChange","yLowChange"])
ws_sell = wb.create_sheet("Sell")
ws_sell.append(["futures", "train set","BuyIndicators", "SellIndicators","Symbol", "VOL_change", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_DAY", "Score","RandomForest", "accuracy", "MLP", "accuracy", "Bagging", "accuracy", "AdaBoost", "accuracy", "KNeighbors", "accuracy", "GradientBoosting", "accuracy", "trend", "yHighChange","yLowChange"])
ws_buyPattern = wb.create_sheet("BuyPattern")
ws_buyPattern.append(["futures", "train set","BuyIndicators", "SellIndicators","Symbol", "VOL_change", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_DAY", "Score","RandomForest", "accuracy", "MLP", "accuracy", "Bagging", "accuracy", "AdaBoost", "accuracy", "KNeighbors", "accuracy", "GradientBoosting", "accuracy", "trend", "yHighChange","yLowChange"])
ws_sellPattern = wb.create_sheet("SellPattern")
ws_sellPattern.append(["futures", "train set","BuyIndicators", "SellIndicators","Symbol", "VOL_change", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_DAY", "Score","RandomForest", "accuracy", "MLP", "accuracy", "Bagging", "accuracy", "AdaBoost", "accuracy", "KNeighbors", "accuracy", "GradientBoosting", "accuracy", "trend", "yHighChange","yLowChange"])
ws_buyPattern1 = wb.create_sheet("BuyPattern1")
ws_buyPattern1.append(["futures", "train set","BuyIndicators", "SellIndicators","Symbol", "VOL_change", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_DAY", "Score","RandomForest", "accuracy", "MLP", "accuracy", "Bagging", "accuracy", "AdaBoost", "accuracy", "KNeighbors", "accuracy", "GradientBoosting", "accuracy", "trend", "yHighChange","yLowChange"])
ws_sellPattern1 = wb.create_sheet("SellPattern1")
ws_sellPattern1.append(["futures", "train set","BuyIndicators", "SellIndicators","Symbol", "VOL_change", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_DAY", "Score","RandomForest", "accuracy", "MLP", "accuracy", "Bagging", "accuracy", "AdaBoost", "accuracy", "KNeighbors", "accuracy", "GradientBoosting", "accuracy", "trend", "yHighChange","yLowChange"])
ws_buyOthers = wb.create_sheet("BuyOthers")
ws_buyOthers.append(["futures", "train set","BuyIndicators", "SellIndicators","Symbol", "VOL_change", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_DAY", "Score","RandomForest", "accuracy", "MLP", "accuracy", "Bagging", "accuracy", "AdaBoost", "accuracy", "KNeighbors", "accuracy", "GradientBoosting", "accuracy", "trend", "yHighChange","yLowChange"])
ws_sellOthers = wb.create_sheet("SellOthers")
ws_sellOthers.append(["futures", "train set","BuyIndicators", "SellIndicators","Symbol", "VOL_change", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_DAY", "Score","RandomForest", "accuracy", "MLP", "accuracy", "Bagging", "accuracy", "AdaBoost", "accuracy", "KNeighbors", "accuracy", "GradientBoosting", "accuracy", "trend", "yHighChange","yLowChange"])

def saveReports(run_type=None):
    # Add a default style with striped rows and banded columns
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
               showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    
    count = 0
    for row in ws_buyAll.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:AD" + str(count))
    tab.tableStyleInfo = style
    ws_buyAll.add_table(tab)
    
    count = 0
    for row in ws_sellAll.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:AD" + str(count))
    tab.tableStyleInfo = style
    ws_sellAll.add_table(tab)
    
    count = 0
    for row in ws_buyFinal.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:AD" + str(count))
    tab.tableStyleInfo = style
    ws_buyFinal.add_table(tab)
    
    count = 0
    for row in ws_sellFinal.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:AD" + str(count))
    tab.tableStyleInfo = style
    ws_sellFinal.add_table(tab)
    
    count = 0
    for row in ws_buyFinal1.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:AD" + str(count))
    tab.tableStyleInfo = style
    ws_buyFinal1.add_table(tab)
    
    count = 0
    for row in ws_sellFinal1.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:AD" + str(count))
    tab.tableStyleInfo = style
    ws_sellFinal1.add_table(tab)
    
    count = 0
    for row in ws_buy.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:AD" + str(count))
    tab.tableStyleInfo = style
    ws_buy.add_table(tab)
    
    count = 0
    for row in ws_sell.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:AD" + str(count))
    tab.tableStyleInfo = style
    ws_sell.add_table(tab)
    
    count = 0
    for row in ws_buyPattern.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:AD" + str(count))
    tab.tableStyleInfo = style
    ws_buyPattern.add_table(tab)
    
    count = 0
    for row in ws_sellPattern.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:AD" + str(count))
    tab.tableStyleInfo = style
    ws_sellPattern.add_table(tab)
      
    count = 0
    for row in ws_buyPattern1.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:AD" + str(count))
    tab.tableStyleInfo = style
    ws_buyPattern1.add_table(tab)
    
    count = 0
    for row in ws_sellPattern1.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:AD" + str(count))
    tab.tableStyleInfo = style
    ws_sellPattern1.add_table(tab)
    
    count = 0
    for row in ws_buyOthers.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:AD" + str(count))
    tab.tableStyleInfo = style
    ws_buyOthers.add_table(tab)
    
    count = 0
    for row in ws_sellOthers.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:AD" + str(count))
    tab.tableStyleInfo = style
    ws_sellOthers.add_table(tab)
      
    if(run_type == 'broker'):
        wb.save(logname + "broker_buy.xlsx")
    elif(run_type == 'result'):
        wb.save(logname + "result.xlsx")
    elif(run_type == 'result_declared'):
        wb.save(logname + "result_declared.xlsx")       
    else:
        wb.save(logname + ".xlsx")

def buy_News(scrip):
    scrip_newsList = db.news.find_one({'scrip':scrip})
    ws_buyNews.append([scrip])
    ws_buyNews.append(["#####################"])
    if(scrip_newsList is None):
        print('Missing news for ', scrip)
        return
    
    for scrip_news in scrip_newsList['news']:
        ws_buyNews.append([scrip_news['timestamp'], scrip_news['summary'], scrip_news['link']])
    ws_buyNews.append([" "])
    
def sell_News(scrip):
    scrip_newsList = db.news.find_one({'scrip':scrip})
    ws_sellNews.append([scrip])
    ws_sellNews.append(["#####################"])
    if(scrip_newsList is None):
        print('Missing news for ', scrip)
        return
    for scrip_news in scrip_newsList['news']:
        ws_sellNews.append([scrip_news['timestamp'], scrip_news['summary'], scrip_news['link']])
    ws_sellNews.append([" "])  

def result_data(scrip):
    regression_data = db.regressionhigh.find_one({'scrip':scrip.replace('&','').replace('-','_')})
    classification_data = db.classificationhigh.find_one({'scrip':scrip.replace('&','').replace('-','_')})
    if(classification_data is not None and regression_data is not None):
        regressionResult = [ ]
        regressionResult.append(regression_data['futures'])
        regressionResult.append(regression_data['trainSize'])
        regressionResult.append(regression_data['buyIndia'])
        regressionResult.append(regression_data['sellIndia'])
        regressionResult.append(regression_data['scrip'])
        regressionResult.append(regression_data['forecast_day_VOL_change'])
        regressionResult.append(regression_data['forecast_day_PCT_change'])
        regressionResult.append(regression_data['forecast_day_PCT2_change'])
        regressionResult.append(regression_data['forecast_day_PCT3_change'])
        regressionResult.append(regression_data['forecast_day_PCT4_change'])
        regressionResult.append(regression_data['forecast_day_PCT5_change'])
        regressionResult.append(regression_data['forecast_day_PCT7_change'])
        regressionResult.append(regression_data['forecast_day_PCT10_change'])
        regressionResult.append(regression_data['PCT_day_change'])
        regressionResult.append(regression_data['score'])
        regressionResult.append(regression_data['randomForestValue'])
        regressionResult.append(regression_data['randomForestAccuracy'])
        regressionResult.append(regression_data['mlpValue'])
        regressionResult.append(regression_data['mlpAccuracy'])
        regressionResult.append(regression_data['baggingValue'])
        regressionResult.append(regression_data['baggingAccuracy'])
        regressionResult.append(regression_data['adaBoostValue'])
        regressionResult.append(regression_data['adaBoostAccuracy'])
        regressionResult.append(regression_data['kNeighboursValue'])
        regressionResult.append(regression_data['kNeighboursAccuracy'])
        regressionResult.append(regression_data['gradientBoostingValue'])
        regressionResult.append(regression_data['gradientBoostingAccuracy'])
        regressionResult.append(regression_data['trend'])
        regressionResult.append(regression_data['yearHighChange'])
        regressionResult.append(regression_data['yearLowChange'])
        classificationResult = [ ]
        classificationResult.append(classification_data['futures'])
        classificationResult.append(classification_data['trainSize'])
        classificationResult.append(classification_data['buyIndia'])
        classificationResult.append(classification_data['sellIndia'])
        classificationResult.append(classification_data['scrip'])
        classificationResult.append(classification_data['forecast_day_VOL_change'])
        classificationResult.append(classification_data['forecast_day_PCT_change'])
        classificationResult.append(classification_data['forecast_day_PCT2_change'])
        classificationResult.append(classification_data['forecast_day_PCT3_change'])
        classificationResult.append(classification_data['forecast_day_PCT4_change'])
        classificationResult.append(classification_data['forecast_day_PCT5_change'])
        classificationResult.append(classification_data['forecast_day_PCT7_change'])
        classificationResult.append(classification_data['forecast_day_PCT10_change'])
        classificationResult.append(classification_data['PCT_day_change'])
        classificationResult.append(classification_data['score'])
        classificationResult.append(classification_data['randomForestValue'])
        classificationResult.append(classification_data['randomForestAccuracy'])
        classificationResult.append(classification_data['mlpValue'])
        classificationResult.append(classification_data['mlpAccuracy'])
        classificationResult.append(classification_data['baggingValue'])
        classificationResult.append(classification_data['baggingAccuracy'])
        classificationResult.append(classification_data['adaBoostValue'])
        classificationResult.append(classification_data['adaBoostAccuracy'])
        classificationResult.append(classification_data['kNeighboursValue'])
        classificationResult.append(classification_data['kNeighboursAccuracy'])
        classificationResult.append(classification_data['gradientBoostingValue'])
        classificationResult.append(classification_data['gradientBoostingAccuracy'])
        classificationResult.append(classification_data['trend'])
        classificationResult.append(classification_data['yearHighChange'])
        classificationResult.append(classification_data['yearLowChange'])
        score = ''
        if(regression_data['score'] == '10' or regression_data['score'] == '1-1'):
            score = 'up'
        if(((regression_data['mlpValue'] >= 1 and regression_data['kNeighboursValue'] >= 0.5) or (regression_data['mlpValue'] >= 0.5 and regression_data['kNeighboursValue'] >= 1)) 
               and ((classification_data['mlpValue'] >= 1 and classification_data['kNeighboursValue'] >= 0) or (classification_data['mlpValue'] >= 1 and classification_data['kNeighboursValue'] >= 1))):
            ws_buyAll.append(regressionResult)
            ws_buyAll.append(classificationResult)
            if(5 > regression_data['PCT_day_change'] > 0 and str(regression_data['sellIndia']) == ''):
                if(regression_data['forecast_day_PCT5_change'] <=0 and regression_data['forecast_day_PCT7_change'] <=0 and regression_data['forecast_day_PCT10_change'] <=0 and 5 > regression_data['forecast_day_PCT_change'] >= 0):
                    ws_buyFinal.append(regressionResult) 
                    ws_buyFinal.append(classificationResult) 
                elif(regression_data['forecast_day_PCT5_change'] <=1 and regression_data['forecast_day_PCT7_change'] <=1):
                    ws_buyFinal1.append(regressionResult)
                    ws_buyFinal1.append(classificationResult)
                else:
                    ws_buy.append(regressionResult)
                    ws_buy.append(classificationResult)
         
        if(4 > regression_data['PCT_day_change']
           and ((regression_data['mlpValue'] >= 1 and regression_data['kNeighboursValue'] >= 0.5) or (regression_data['mlpValue'] >= 0.5 and regression_data['kNeighboursValue'] >= 1))
           and ((classification_data['mlpValue'] >= 1 and classification_data['kNeighboursValue'] >= 0) or (classification_data['mlpValue'] >= 1 and classification_data['kNeighboursValue'] >= 1))
           ):
            if(('MARUBOZU' in str(regression_data['buyIndia']) and regression_data['forecast_day_PCT5_change'] <= 0)
               or 'HAMMER' in str(regression_data['buyIndia'])
               #or 'ENGULFING' in str(regression_data['buyIndia'])
               #or 'PIERCING' in str(regression_data['buyIndia'])
               or 'MORNINGSTAR' in str(regression_data['buyIndia'])
               #or ':DOJISTAR' in str(regression_data['buyIndia'])
               #or 'MORNINGDOJISTAR' in str(regression_data['buyIndia'])
               or 'ABANDONEDBABY' in str(regression_data['buyIndia'])
               or 'COUNTERATTACK' in str(regression_data['buyIndia'])
               or 'KICKING' in str(regression_data['buyIndia'])
               or 'BREAKAWAY' in str(regression_data['buyIndia'])
               #or 'TRISTAR' in str(regression_data['buyIndia'])
               #or '3WHITESOLDIERS' in str(regression_data['buyIndia'])
               #or '3INSIDE' in str(regression_data['buyIndia'])
               or ('CCI:BOP' in str(regression_data['buyIndia']) and 'BELTHOLD' in str(regression_data['buyIndia']))
               or ('AROON:BOP' in str(regression_data['buyIndia']) and 'BELTHOLD' in str(regression_data['buyIndia']) and 'ENGULFING' in str(regression_data['buyIndia']))
               or ('BELTHOLD' == str(regression_data['buyIndia']) and regression_data['forecast_day_PCT5_change'] <= 0 and score == 'up')
               or ('3OUTSIDE' in str(regression_data['buyIndia']) and regression_data['forecast_day_PCT5_change'] <= 0 and score == 'up')
               or ('HARAMI' in str(regression_data['buyIndia']) and regression_data['forecast_day_PCT5_change'] <= 0 and score == 'up')
               or (regression_data['yearHighChange'] <= -35 and 'HARAMI' in str(regression_data['buyIndia']) and 'SHORTLINE' in str(regression_data['buyIndia']))
               or (regression_data['yearHighChange'] <= -35 and 'BELTHOLD' in str(regression_data['buyIndia']) and 'LONGLINE' in str(regression_data['buyIndia']))
               or (regression_data['yearHighChange'] <= -35 and ',CCI:BOP' in str(regression_data['buyIndia']) and 'LONGLINE' in str(regression_data['buyIndia']))
               ):
                ws_buyPattern.append(regressionResult)
                ws_buyPattern.append(classificationResult)       
            elif(str(regression_data['buyIndia']) != ''):
                ws_buyPattern1.append(regressionResult)
                ws_buyPattern1.append(classificationResult)
            elif(str(regression_data['buyIndia']) == '' and regression_data['yearLowChange'] >= 10): 
                ws_buyOthers.append(regressionResult)
                ws_buyOthers.append(classificationResult)                    
    
        
    regression_data = db.regressionlow.find_one({'scrip':scrip.replace('&','').replace('-','_')})
    classification_data = db.classificationlow.find_one({'scrip':scrip.replace('&','').replace('-','_')})
    if(classification_data is not None or regression_data is not None):
        regressionResult = [ ]
        regressionResult.append(regression_data['futures'])
        regressionResult.append(regression_data['trainSize'])
        regressionResult.append(regression_data['buyIndia'])
        regressionResult.append(regression_data['sellIndia'])
        regressionResult.append(regression_data['scrip'])
        regressionResult.append(regression_data['forecast_day_VOL_change'])
        regressionResult.append(regression_data['forecast_day_PCT_change'])
        regressionResult.append(regression_data['forecast_day_PCT2_change'])
        regressionResult.append(regression_data['forecast_day_PCT3_change'])
        regressionResult.append(regression_data['forecast_day_PCT4_change'])
        regressionResult.append(regression_data['forecast_day_PCT5_change'])
        regressionResult.append(regression_data['forecast_day_PCT7_change'])
        regressionResult.append(regression_data['forecast_day_PCT10_change'])
        regressionResult.append(regression_data['PCT_day_change'])
        regressionResult.append(regression_data['score'])
        regressionResult.append(regression_data['randomForestValue'])
        regressionResult.append(regression_data['randomForestAccuracy'])
        regressionResult.append(regression_data['mlpValue'])
        regressionResult.append(regression_data['mlpAccuracy'])
        regressionResult.append(regression_data['baggingValue'])
        regressionResult.append(regression_data['baggingAccuracy'])
        regressionResult.append(regression_data['adaBoostValue'])
        regressionResult.append(regression_data['adaBoostAccuracy'])
        regressionResult.append(regression_data['kNeighboursValue'])
        regressionResult.append(regression_data['kNeighboursAccuracy'])
        regressionResult.append(regression_data['gradientBoostingValue'])
        regressionResult.append(regression_data['gradientBoostingAccuracy'])
        regressionResult.append(regression_data['trend'])
        regressionResult.append(regression_data['yearHighChange'])
        regressionResult.append(regression_data['yearLowChange'])
        classificationResult = [ ]
        classificationResult.append(classification_data['futures'])
        classificationResult.append(classification_data['trainSize'])
        classificationResult.append(classification_data['buyIndia'])
        classificationResult.append(classification_data['sellIndia'])
        classificationResult.append(classification_data['scrip'])
        classificationResult.append(classification_data['forecast_day_VOL_change'])
        classificationResult.append(classification_data['forecast_day_PCT_change'])
        classificationResult.append(classification_data['forecast_day_PCT2_change'])
        classificationResult.append(classification_data['forecast_day_PCT3_change'])
        classificationResult.append(classification_data['forecast_day_PCT4_change'])
        classificationResult.append(classification_data['forecast_day_PCT5_change'])
        classificationResult.append(classification_data['forecast_day_PCT7_change'])
        classificationResult.append(classification_data['forecast_day_PCT10_change'])
        classificationResult.append(classification_data['PCT_day_change'])
        classificationResult.append(classification_data['score'])
        classificationResult.append(classification_data['randomForestValue'])
        classificationResult.append(classification_data['randomForestAccuracy'])
        classificationResult.append(classification_data['mlpValue'])
        classificationResult.append(classification_data['mlpAccuracy'])
        classificationResult.append(classification_data['baggingValue'])
        classificationResult.append(classification_data['baggingAccuracy'])
        classificationResult.append(classification_data['adaBoostValue'])
        classificationResult.append(classification_data['adaBoostAccuracy'])
        classificationResult.append(classification_data['kNeighboursValue'])
        classificationResult.append(classification_data['kNeighboursAccuracy'])
        classificationResult.append(classification_data['gradientBoostingValue'])
        classificationResult.append(classification_data['gradientBoostingAccuracy'])
        classificationResult.append(classification_data['trend'])
        classificationResult.append(classification_data['yearHighChange'])
        classificationResult.append(classification_data['yearLowChange'])
        score = ''
        if(regression_data['score'] == '01' or regression_data['score'] == '0-1'):
            score = 'down'
        if(((regression_data['mlpValue'] <= -1 and regression_data['kNeighboursValue'] <= -0.5) or (regression_data['mlpValue'] <= -0.5 and regression_data['kNeighboursValue'] <= -1))
               and ((classification_data['mlpValue'] <= -1 and classification_data['kNeighboursValue'] <= 0) or (classification_data['mlpValue'] <= -1 and classification_data['kNeighboursValue'] <= -1))):
            ws_sellAll.append(regressionResult)
            ws_sellAll.append(classificationResult)
            if(-5 < regression_data['PCT_day_change'] < 0 and str(regression_data['buyIndia']) == ''):
                if(regression_data['forecast_day_PCT5_change'] >= 0 and regression_data['forecast_day_PCT7_change'] >= 0 and regression_data['forecast_day_PCT10_change'] >= 0 and -5 < regression_data['forecast_day_PCT_change'] <= 0):
                    ws_sellFinal.append(regressionResult) 
                    ws_sellFinal.append(classificationResult)
                elif(regression_data['forecast_day_PCT5_change'] >= 1 and regression_data['forecast_day_PCT7_change'] >= 1):
                    ws_sellFinal1.append(regressionResult)
                    ws_sellFinal1.append(classificationResult) 
                else:
                    ws_sell.append(regressionResult)
                    ws_sell.append(classificationResult)  
                               
        if(-4 < regression_data['PCT_day_change'] 
               and ((regression_data['mlpValue'] <= -1 and regression_data['kNeighboursValue'] <= -0.5) or (regression_data['mlpValue'] <= -0.5 and regression_data['kNeighboursValue'] <= -1))
               and ((classification_data['mlpValue'] <= -1 and classification_data['kNeighboursValue'] <= 0) or (classification_data['mlpValue'] <= -1 and classification_data['kNeighboursValue'] <= -1))
               ):
            if(('MARUBOZU' in str(regression_data['sellIndia'])
               or 'HANGINGMAN' in str(regression_data['sellIndia'])
               #or 'ENGULFING' in str(regression_data['sellIndia'])
               or 'EVENINGSTAR' in str(regression_data['sellIndia'])
               #or ':DOJISTAR' in str(regression_data['sellIndia'])
               #or 'EVENINGDOJISTAR' in str(regression_data['sellIndia'])
               or 'ABANDONEDBABY' in str(regression_data['sellIndia'])
               or 'COUNTERATTACK' in str(regression_data['sellIndia'])
               or 'KICKING' in str(regression_data['sellIndia'])
               or 'BREAKAWAY' in str(regression_data['sellIndia'])
               #or 'TRISTAR' in str(regression_data['sellIndia'])
               or 'SHOOTINGSTAR' in str(regression_data['sellIndia'])
               or 'DARKCLOUDCOVER' in str(regression_data['sellIndia'])
               #or '3INSIDE' in str(regression_data['sellIndia'])
               #or '3OUTSIDE' in str(regression_data['sellIndia'])
               #or '2CROWS' in str(regression_data['sellIndia'])
               #or '3BLACKCROWS' in str(regression_data['sellIndia'])
               or ('HARAMI' in str(regression_data['sellIndia']) and regression_data['forecast_day_PCT5_change'] >= 0 and score == 'down')
               ) and (regression_data['forecast_day_PCT5_change'] >= 0)):
                ws_sellPattern.append(regressionResult)
                ws_sellPattern.append(classificationResult) 
            elif(str(regression_data['sellIndia']) != ''):
                ws_sellPattern1.append(regressionResult)
                ws_sellPattern1.append(classificationResult)
            elif(str(regression_data['sellIndia']) == '' and regression_data['yearHighChange'] <= -10):
                ws_sellOthers.append(regressionResult)
                ws_sellOthers.append(classificationResult)    
                                   
def calculateParallel(threads=2, run_type=None, futures=None):
    pool = ThreadPool(threads)
    if(run_type == 'broker'):
        count=0
        scrips = []
        with open('../data-import/nselist/ind_broker_buy.csv') as csvfile:
            readCSV = csv.reader(csvfile, delimiter=',')
            for row in readCSV:
                if (count != 0):
                    scrips.append(row[0].replace('&','').replace('-','_'))
                count = count + 1
                
            scrips.sort()
            pool.map(result_data, scrips)
    elif(run_type == 'result'):
        count=0
        scrips = []
        with open('../data-import/nselist/ind_result.csv') as csvfile:
            readCSV = csv.reader(csvfile, delimiter=',')
            for row in readCSV:
                if (count != 0):
                    scrips.append(row[0].replace('&','').replace('-','_'))
                count = count + 1
                
            scrips.sort()
            pool.map(result_data, scrips)  
    elif(run_type == 'result_declared'):
        count=0
        scrips = []
        with open('../data-import/nselist/ind_result_declared.csv') as csvfile:
            readCSV = csv.reader(csvfile, delimiter=',')
            for row in readCSV:
                if (count != 0):
                    scrips.append(row[0].replace('&','').replace('-','_'))
                count = count + 1
                
            scrips.sort()
            pool.map(result_data, scrips)               
    else:
        scrips = []
        for data in db.scrip.find({'futures':futures}):
            scrips.append(data['scrip'].replace('&','').replace('-','_'))
        scrips.sort()
        pool.map(result_data, scrips)
                     
if __name__ == "__main__":
    if not os.path.exists(directory):
        os.makedirs(directory)
    calculateParallel(1, sys.argv[1], sys.argv[2])
    connection.close()
    saveReports(sys.argv[1])