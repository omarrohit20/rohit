import os, logging, sys, json
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Color, PatternFill, Font, Border
from pymongo import MongoClient
from multiprocessing.dummy import Pool as ThreadPool

import quandl, math, time
import pandas as pd
import numpy as np
from talib.abstract import *
from pip.req.req_file import preprocess

import time
import gc

connection = MongoClient('localhost', 27017)
db = connection.Nsedata

directory = '../../output/final'
logname = '../../output/final' + '/results' + time.strftime("%d%m%y-%H%M%S")

wb = Workbook()
ws = wb.active
ws.append(["futures", "train set","BuyIndicators", "SellIndicators","Symbol", "VOL_change", "PCT_change", "Score","RandomForest", "accuracy", "MLP", "accuracy", "Bagging", "accuracy", "AdaBoost", "accuracy", "KNeighbors", "accuracy", "GradientBoosting", "accuracy", "trend", "yHighChange","yLowChange"])
ws_buy = wb.create_sheet("Buy")
ws_buy.append(["futures", "train set","BuyIndicators", "SellIndicators","Symbol", "VOL_change", "PCT_change", "Score","RandomForest", "accuracy", "MLP", "accuracy", "Bagging", "accuracy", "AdaBoost", "accuracy", "KNeighbors", "accuracy", "GradientBoosting", "accuracy", "trend", "yHighChange","yLowChange"])
ws_sell = wb.create_sheet("Sell")
ws_sell.append(["futures", "train set","BuyIndicators", "SellIndicators","Symbol", "VOL_change", "PCT_change", "Score","RandomForest", "accuracy", "MLP", "accuracy", "Bagging", "accuracy", "AdaBoost", "accuracy", "KNeighbors", "accuracy", "GradientBoosting", "accuracy", "trend", "yHighChange","yLowChange"])
ws_buyAll = wb.create_sheet("BuyAll")
ws_buyAll.append(["futures", "train set","BuyIndicators", "SellIndicators","Symbol", "VOL_change", "PCT_change", "Score","RandomForest", "accuracy", "MLP", "accuracy", "Bagging", "accuracy", "AdaBoost", "accuracy", "KNeighbors", "accuracy", "GradientBoosting", "accuracy", "trend", "yHighChange","yLowChange"])
ws_sellAll = wb.create_sheet("SellAll")
ws_sellAll.append(["futures", "train set","BuyIndicators", "SellIndicators","Symbol", "VOL_change", "PCT_change", "Score","RandomForest", "accuracy", "MLP", "accuracy", "Bagging", "accuracy", "AdaBoost", "accuracy", "KNeighbors", "accuracy", "GradientBoosting", "accuracy", "trend", "yHighChange","yLowChange"])

def saveReports():
    # Add a default style with striped rows and banded columns
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
               showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    
    count = 0
    for row in ws.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:W" + str(count))
    tab.tableStyleInfo = style
    ws.add_table(tab)
    
    count = 0
    for row in ws_buy.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:W" + str(count))
    tab.tableStyleInfo = style
    ws_buy.add_table(tab)
    
    count = 0
    for row in ws_sell.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:W" + str(count))
    tab.tableStyleInfo = style
    ws_sell.add_table(tab)
    
    count = 0
    for row in ws_buyAll.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:W" + str(count))
    tab.tableStyleInfo = style
    ws_buyAll.add_table(tab)
    
    count = 0
    for row in ws_sellAll.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:W" + str(count))
    tab.tableStyleInfo = style
    ws_sellAll.add_table(tab)
      
    wb.save(logname + ".xlsx")

def result_data(scrip):
    classification_data = db.classification.find_one({'scrip':str(scrip.encode('UTF8')).replace('&','').replace('-','_')})
    regression_data = db.regression.find_one({'scrip':str(scrip.encode('UTF8')).replace('&','').replace('-','_')})
    
    if(classification_data is None or regression_data is None):
        print('Missing or very less Data for ', scrip.encode('UTF8'))
        return
    
    #Buy Indicators
    regressionResult = [ ]
    regressionResult.append("YES")
    regressionResult.append(regression_data['trainSize'])
    regressionResult.append(regression_data['buyIndia'])
    regressionResult.append(regression_data['sellIndia'])
    regressionResult.append(regression_data['scrip'])
    regressionResult.append(regression_data['forecast_day_VOL_change'])
    regressionResult.append(regression_data['forecast_day_PCT_change'])
    regressionResult.append(regression_data['score'])
    regressionResult.append(regression_data['randomForestValue'])
    regressionResult.append(regression_data['randomForestAccuracy'])
    regressionResult.append(regression_data['mlpValue'])
    regressionResult.append(regression_data['mlpAccuracy'])
    regressionResult.append(regression_data['baggingValue'])
    regressionResult.append(regression_data['baggingAccuracy'])
    regressionResult.append(regression_data['adaBoostValue'])
    regressionResult.append(regression_data['adaBoostAccuracy'])
    regressionResult.append(regression_data['kNeighboursValue'])
    regressionResult.append(regression_data['kNeighboursAccuracy'])
    regressionResult.append(regression_data['gradientBoostingValue'])
    regressionResult.append(regression_data['gradientBoostingAccuracy'])
    regressionResult.append(regression_data['trend'])
    regressionResult.append(regression_data['yearHighChange'])
    regressionResult.append(regression_data['yearLowChange'])
    if(classification_data['kNeighboursValue'] >= 0 and regression_data['kNeighboursValue'] > .5 and len(regression_data['sellIndia'].encode('utf8')) < 1):
        ws_buy.append(regressionResult)
    if(regression_data['kNeighboursValue'] > .5):
        ws_buyAll.append(regressionResult)    
        
    #Sell Indicators
    regressionResult = [ ]
    regressionResult.append("YES")
    regressionResult.append(classification_data['trainSize'])
    regressionResult.append(classification_data['buyIndia'])
    regressionResult.append(classification_data['sellIndia'])
    regressionResult.append(classification_data['scrip'])
    regressionResult.append(classification_data['forecast_day_VOL_change'])
    regressionResult.append(classification_data['forecast_day_PCT_change'])
    regressionResult.append(classification_data['score'])
    regressionResult.append(classification_data['randomForestValue'])
    regressionResult.append(classification_data['randomForestAccuracy'])
    regressionResult.append(classification_data['mlpValue'])
    regressionResult.append(classification_data['mlpAccuracy'])
    regressionResult.append(classification_data['baggingValue'])
    regressionResult.append(classification_data['baggingAccuracy'])
    regressionResult.append(classification_data['adaBoostValue'])
    regressionResult.append(classification_data['adaBoostAccuracy'])
    regressionResult.append(classification_data['kNeighboursValue'])
    regressionResult.append(classification_data['kNeighboursAccuracy'])
    regressionResult.append(classification_data['gradientBoostingValue'])
    regressionResult.append(classification_data['gradientBoostingAccuracy'])
    regressionResult.append(classification_data['trend'])
    regressionResult.append(classification_data['yearHighChange'])
    regressionResult.append(classification_data['yearLowChange'])
    if(classification_data['kNeighboursValue'] < 0 and regression_data['kNeighboursValue'] <= 0 and len(regression_data['buyIndia'].encode('utf8')) < 1):            
        ws_sell.append(regressionResult)  
    if(classification_data['kNeighboursValue'] < 0):            
        ws_sellAll.append(regressionResult)     
                                  

def calculateParallel(threads=2):
    pool = ThreadPool(threads)
    
    scrips = []
    for data in db.scrip.find():
        scrips.append(str((data['scrip']).encode('UTF8')).replace('&','').replace('-','_'))
    scrips.sort()
    
    pool.map(result_data, scrips)
                     
if __name__ == "__main__":
    if not os.path.exists(directory):
        os.makedirs(directory)
    calculateParallel(1)
    connection.close()
    saveReports()