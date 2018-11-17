import os, logging, sys, json, csv
sys.path.insert(0, '../')

from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Color, PatternFill, Font, Border
from pymongo import MongoClient
from multiprocessing.dummy import Pool as ThreadPool

import quandl, math, time
import pandas as pd
import numpy as np
from talib.abstract import *

import datetime
import time
import gc

from util.util import getScore, all_day_pct_change_negative, all_day_pct_change_positive, no_doji_or_spinning_buy_india, no_doji_or_spinning_sell_india, scrip_patterns_to_dict
from util.util import is_algo_buy, is_algo_sell
from util.util import get_regressionResult
from util.util import buy_pattern_from_history, buy_all_rule, buy_year_high, buy_year_low, buy_up_trend, buy_down_trend, buy_final, buy_high_indicators, buy_pattern
from util.util import sell_pattern_from_history, sell_all_rule, sell_year_high, sell_year_low, sell_up_trend, sell_down_trend, sell_final, sell_high_indicators, sell_pattern
from util.util import buy_pattern_without_mlalgo, sell_pattern_without_mlalgo, buy_oi, sell_oi, all_withoutml
from util.util import buy_oi_candidate, sell_oi_candidate
from util.util import buy_all_filter, buy_all_common, sell_all_filter, sell_all_common
from util.util import buy_all_rule_classifier, sell_all_rule_classifier
from util.util import is_algo_buy_classifier, is_algo_sell_classifier
from util.util import sell_oi_negative, sell_day_high, buy_oi_negative, buy_day_low



connection = MongoClient('localhost', 27017)
db = connection.Nsedata

buyPatternsDict=scrip_patterns_to_dict('../../data-import/nselist/patterns-buy.csv')
sellPatternsDict=scrip_patterns_to_dict('../../data-import/nselist/patterns-sell.csv')

directory = '../../output/final'
logname = '../../output/final' + '/all-result' + time.strftime("%d%m%y-%H%M%S")

newsDict = {}
wb = Workbook()
ws_buyAll = wb.create_sheet("BuyAllBoth")
ws_buyAll.append(["BuyIndicators", "SellIndicators", "Symbol", "seriesTrend", "SMA25", "SMA50", "SMA100", "SMA200", "ResultDate", "ResultDeclared", "ResultSentiment", "ResultComment", "VOL_change", "OI_change", "Contract_change", "OI_change_next", "Contract_change_next", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "MLP", "KNeighbors", "MLP_Other", "KNeighbors_Other", "yHigh2Change", "yLow2Change", "yHighChange", "yLowChange", "m6HighChange", "m6LowChange", "m3HighChange", "m3LowChange", "trend", "Score", "HighTail", "LowTail", "PCT_Day_Change", "PCT_Change", "Symbol", "Filter", "Filter1", "Filter2", "Filter3"])
ws_buyAllReg = wb.create_sheet("BuyAllReg")
ws_buyAllReg.append(["BuyIndicators", "SellIndicators", "Symbol", "seriesTrend", "SMA25", "SMA50", "SMA100", "SMA200", "ResultDate", "ResultDeclared", "ResultSentiment", "ResultComment", "VOL_change", "OI_change", "Contract_change", "OI_change_next", "Contract_change_next", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "MLP", "KNeighbors", "MLP_Other", "KNeighbors_Other", "yHigh2Change", "yLow2Change", "yHighChange", "yLowChange", "m6HighChange", "m6LowChange", "m3HighChange", "m3LowChange", "trend", "Score", "HighTail", "LowTail", "PCT_Day_Change", "PCT_Change", "Symbol", "Filter", "Filter1", "Filter2", "Filter3"])
ws_buyOI = wb.create_sheet("BuyOIBoth")
ws_buyOI.append(["BuyIndicators", "SellIndicators", "Symbol", "seriesTrend", "SMA25", "SMA50", "SMA100", "SMA200", "ResultDate", "ResultDeclared", "ResultSentiment", "ResultComment", "VOL_change", "OI_change", "Contract_change", "OI_change_next", "Contract_change_next", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "MLP", "KNeighbors", "MLP_Other", "KNeighbors_Other", "yHigh2Change", "yLow2Change", "yHighChange", "yLowChange", "m6HighChange", "m6LowChange", "m3HighChange", "m3LowChange", "trend", "Score", "HighTail", "LowTail", "PCT_Day_Change", "PCT_Change", "Symbol", "Filter", "Filter1", "Filter2", "Filter3"])
ws_buyOIReg = wb.create_sheet("BuyOIReg")
ws_buyOIReg.append(["BuyIndicators", "SellIndicators", "Symbol", "seriesTrend", "SMA25", "SMA50", "SMA100", "SMA200", "ResultDate", "ResultDeclared", "ResultSentiment", "ResultComment", "VOL_change", "OI_change", "Contract_change", "OI_change_next", "Contract_change_next", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "MLP", "KNeighbors", "MLP_Other", "KNeighbors_Other", "yHigh2Change", "yLow2Change", "yHighChange", "yLowChange", "m6HighChange", "m6LowChange", "m3HighChange", "m3LowChange", "trend", "Score", "HighTail", "LowTail", "PCT_Day_Change", "PCT_Change", "Symbol", "Filter", "Filter1", "Filter2", "Filter3"])

ws_sellAll = wb.create_sheet("SellAllBoth")
ws_sellAll.append(["BuyIndicators", "SellIndicators", "Symbol", "seriesTrend", "SMA25", "SMA50", "SMA100", "SMA200", "ResultDate", "ResultDeclared", "ResultSentiment", "ResultComment", "VOL_change", "OI_change", "Contract_change", "OI_change_next", "Contract_change_next", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "MLP", "KNeighbors", "MLP_Other", "KNeighbors_Other", "yHigh2Change", "yLow2Change", "yHighChange", "yLowChange", "m6HighChange", "m6LowChange", "m3HighChange", "m3LowChange", "trend", "Score", "HighTail", "LowTail", "PCT_Day_Change", "PCT_Change", "Symbol", "Filter", "Filter1", "Filter2", "Filter3"])
ws_sellAllReg = wb.create_sheet("SellAllReg")
ws_sellAllReg.append(["BuyIndicators", "SellIndicators", "Symbol", "seriesTrend", "SMA25", "SMA50", "SMA100", "SMA200", "ResultDate", "ResultDeclared", "ResultSentiment", "ResultComment", "VOL_change", "OI_change", "Contract_change", "OI_change_next", "Contract_change_next", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "MLP", "KNeighbors", "MLP_Other", "KNeighbors_Other", "yHigh2Change", "yLow2Change", "yHighChange", "yLowChange", "m6HighChange", "m6LowChange", "m3HighChange", "m3LowChange", "trend", "Score", "HighTail", "LowTail", "PCT_Day_Change", "PCT_Change", "Symbol", "Filter", "Filter1", "Filter2", "Filter3"])
ws_sellOI = wb.create_sheet("SellOIBoth")
ws_sellOI.append(["BuyIndicators", "SellIndicators", "Symbol", "seriesTrend", "SMA25", "SMA50", "SMA100", "SMA200", "ResultDate", "ResultDeclared", "ResultSentiment", "ResultComment", "VOL_change", "OI_change", "Contract_change", "OI_change_next", "Contract_change_next", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "MLP", "KNeighbors", "MLP_Other", "KNeighbors_Other", "yHigh2Change", "yLow2Change", "yHighChange", "yLowChange", "m6HighChange", "m6LowChange", "m3HighChange", "m3LowChange", "trend", "Score", "HighTail", "LowTail", "PCT_Day_Change", "PCT_Change", "Symbol", "Filter", "Filter1", "Filter2", "Filter3"])
ws_sellOIReg = wb.create_sheet("SellOIReg")
ws_sellOIReg.append(["BuyIndicators", "SellIndicators", "Symbol", "seriesTrend", "SMA25", "SMA50", "SMA100", "SMA200", "ResultDate", "ResultDeclared", "ResultSentiment", "ResultComment", "VOL_change", "OI_change", "Contract_change", "OI_change_next", "Contract_change_next", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "MLP", "KNeighbors", "MLP_Other", "KNeighbors_Other", "yHigh2Change", "yLow2Change", "yHighChange", "yLowChange", "m6HighChange", "m6LowChange", "m3HighChange", "m3LowChange", "trend", "Score", "HighTail", "LowTail", "PCT_Day_Change", "PCT_Change", "Symbol", "Filter", "Filter1", "Filter2", "Filter3"])

ws_buyAllCla = wb.create_sheet("BuyAllCla")
ws_buyAllCla.append(["BuyIndicators", "SellIndicators", "Symbol", "seriesTrend", "SMA25", "SMA50", "SMA100", "SMA200", "ResultDate", "ResultDeclared", "ResultSentiment", "ResultComment", "VOL_change", "OI_change", "Contract_change", "OI_change_next", "Contract_change_next", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "MLP", "KNeighbors", "MLP_Other", "KNeighbors_Other", "yHigh2Change", "yLow2Change", "yHighChange", "yLowChange", "m6HighChange", "m6LowChange", "m3HighChange", "m3LowChange", "trend", "Score", "HighTail", "LowTail", "PCT_Day_Change", "PCT_Change", "Symbol", "Filter", "Filter1", "Filter2", "Filter3"])
ws_buyOICla = wb.create_sheet("BuyOICla")
ws_buyOICla.append(["BuyIndicators", "SellIndicators", "Symbol", "seriesTrend", "SMA25", "SMA50", "SMA100", "SMA200", "ResultDate", "ResultDeclared", "ResultSentiment", "ResultComment", "VOL_change", "OI_change", "Contract_change", "OI_change_next", "Contract_change_next", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "MLP", "KNeighbors", "MLP_Other", "KNeighbors_Other", "yHigh2Change", "yLow2Change", "yHighChange", "yLowChange", "m6HighChange", "m6LowChange", "m3HighChange", "m3LowChange", "trend", "Score", "HighTail", "LowTail", "PCT_Day_Change", "PCT_Change", "Symbol", "Filter", "Filter1", "Filter2", "Filter3"])
ws_sellAllCla = wb.create_sheet("SellAllCla")
ws_sellAllCla.append(["BuyIndicators", "SellIndicators", "Symbol", "seriesTrend", "SMA25", "SMA50", "SMA100", "SMA200", "ResultDate", "ResultDeclared", "ResultSentiment", "ResultComment", "VOL_change", "OI_change", "Contract_change", "OI_change_next", "Contract_change_next", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "MLP", "KNeighbors", "MLP_Other", "KNeighbors_Other", "yHigh2Change", "yLow2Change", "yHighChange", "yLowChange", "m6HighChange", "m6LowChange", "m3HighChange", "m3LowChange", "trend", "Score", "HighTail", "LowTail", "PCT_Day_Change", "PCT_Change", "Symbol", "Filter", "Filter1", "Filter2", "Filter3"])
ws_sellOICla = wb.create_sheet("SellOICla")
ws_sellOICla.append(["BuyIndicators", "SellIndicators", "Symbol", "seriesTrend", "SMA25", "SMA50", "SMA100", "SMA200", "ResultDate", "ResultDeclared", "ResultSentiment", "ResultComment", "VOL_change", "OI_change", "Contract_change", "OI_change_next", "Contract_change_next", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "MLP", "KNeighbors", "MLP_Other", "KNeighbors_Other", "yHigh2Change", "yLow2Change", "yHighChange", "yLowChange", "m6HighChange", "m6LowChange", "m3HighChange", "m3LowChange", "trend", "Score", "HighTail", "LowTail", "PCT_Day_Change", "PCT_Change", "Symbol", "Filter", "Filter1", "Filter2", "Filter3"])
ws_highAll = wb.create_sheet("HighAll")
ws_highAll.append(["BuyIndicators", "SellIndicators", "Symbol", "seriesTrend", "SMA25", "SMA50", "SMA100", "SMA200", "ResultDate", "ResultDeclared", "ResultSentiment", "ResultComment", "VOL_change", "OI_change", "Contract_change", "OI_change_next", "Contract_change_next", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "MLP", "KNeighbors", "MLP_Other", "KNeighbors_Other", "yHigh2Change", "yLow2Change", "yHighChange", "yLowChange", "m6HighChange", "m6LowChange", "m3HighChange", "m3LowChange", "trend", "Score", "HighTail", "LowTail", "PCT_Day_Change", "PCT_Change", "Symbol", "Filter", "Filter1", "Filter2", "Filter3"])
ws_lowAll = wb.create_sheet("LowAll")
ws_lowAll.append(["BuyIndicators", "SellIndicators", "Symbol", "seriesTrend", "SMA25", "SMA50", "SMA100", "SMA200", "ResultDate", "ResultDeclared", "ResultSentiment", "ResultComment", "VOL_change", "OI_change", "Contract_change", "OI_change_next", "Contract_change_next", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "MLP", "KNeighbors", "MLP_Other", "KNeighbors_Other", "yHigh2Change", "yLow2Change", "yHighChange", "yLowChange", "m6HighChange", "m6LowChange", "m3HighChange", "m3LowChange", "trend", "Score", "HighTail", "LowTail", "PCT_Day_Change", "PCT_Change", "Symbol", "Filter", "Filter1", "Filter2", "Filter3"])

def saveReports(run_type=None):
    ws_buyAll.append([""])
    ws_buyOI.append([""])
    ws_buyAllReg.append([""])
    ws_buyOIReg.append([""])
    ws_buyAllCla.append([""])
    ws_buyOICla.append([""])
    ws_highAll.append([""])
        
    ws_sellAll.append([""])
    ws_sellOI.append([""])
    ws_sellAllReg.append([""])
    ws_sellOIReg.append([""])
    ws_sellAllCla.append([""])
    ws_sellOICla.append([""])
    ws_lowAll.append([""])
    
    # Add a default style with striped rows and banded columns
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
               showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    
    count = 0
    for row in ws_buyAll.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:AU" + str(count))
    tab.tableStyleInfo = style
    ws_buyAll.add_table(tab)
    
    count = 0
    for row in ws_buyOI.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:AU" + str(count))
    tab.tableStyleInfo = style
    ws_buyOI.add_table(tab)
    
    count = 0
    for row in ws_buyAllReg.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:AU" + str(count))
    tab.tableStyleInfo = style
    ws_buyAllReg.add_table(tab)
    
    count = 0
    for row in ws_buyOIReg.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:AU" + str(count))
    tab.tableStyleInfo = style
    ws_buyOIReg.add_table(tab)
    
    count = 0
    for row in ws_buyAllCla.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:AU" + str(count))
    tab.tableStyleInfo = style
    ws_buyAllCla.add_table(tab)
    
    count = 0
    for row in ws_buyOICla.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:AU" + str(count))
    tab.tableStyleInfo = style
    ws_buyOICla.add_table(tab)
    
    count = 0
    for row in ws_highAll.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:AU" + str(count))
    tab.tableStyleInfo = style
    ws_highAll.add_table(tab)
    
    
    count = 0
    for row in ws_sellAll.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:AU" + str(count))
    tab.tableStyleInfo = style
    ws_sellAll.add_table(tab)
    
    count = 0
    for row in ws_sellOI.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:AU" + str(count))
    tab.tableStyleInfo = style
    ws_sellOI.add_table(tab)
    
    count = 0
    for row in ws_sellAllReg.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:AU" + str(count))
    tab.tableStyleInfo = style
    ws_sellAllReg.add_table(tab)
    
    count = 0
    for row in ws_sellOIReg.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:AU" + str(count))
    tab.tableStyleInfo = style
    ws_sellOIReg.add_table(tab)
    
    count = 0
    for row in ws_sellAllCla.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:AU" + str(count))
    tab.tableStyleInfo = style
    ws_sellAllCla.add_table(tab)
    
    count = 0
    for row in ws_sellOICla.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:AU" + str(count))
    tab.tableStyleInfo = style
    ws_sellOICla.add_table(tab)
    
    count = 0
    for row in ws_lowAll.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:AU" + str(count))
    tab.tableStyleInfo = style
    ws_lowAll.add_table(tab)
    
    
    wb.save(logname + ".xlsx")
      
def result_data(scrip):
    regression_high = db.regressionhigh.find_one({'scrip':scrip.replace('&','').replace('-','_')})
    regression_low = db.regressionlow.find_one({'scrip':scrip.replace('&','').replace('-','_')})
    classification_high = db.classificationhigh.find_one({'scrip':scrip.replace('&','').replace('-','_')})
    classification_low = db.classificationlow.find_one({'scrip':scrip.replace('&','').replace('-','_')})
    
    regression_data = regression_high
    classification_data = classification_high
    regressionResult = get_regressionResult(regression_data, scrip, db, regression_low['mlpValue'], regression_low['kNeighboursValue'])
    classificationResult = get_regressionResult(classification_data, scrip, db, classification_low['mlpValue'], classification_low['kNeighboursValue'])
    
    re_oi = buy_oi_candidate(regression_data, regressionResult, None)
    cl_oi = buy_oi_candidate(classification_data, classificationResult, None)
    if(cl_oi and re_oi):
        all_withoutml(regression_data, regressionResult, ws_buyOI)
        all_withoutml(classification_data, classificationResult, ws_buyOI)    
    
    buyIndiaAvgReg, result = buy_pattern_from_history(regression_data, regressionResult, None)
    buyIndiaAvgCla, result = buy_pattern_from_history(classification_data, classificationResult, None)
    if (buy_all_rule(regression_data, regressionResult, buyIndiaAvgReg, None)
        and ( buy_all_rule(classification_data, classificationResult, buyIndiaAvgCla, None)
              or buy_all_rule_classifier(classification_data, classificationResult, buyIndiaAvgCla, None)
            )
        ):
        buy_all_filter(regression_data, regressionResult, None)
        buy_all_common(regression_data, regressionResult, None)
        if (buy_all_rule(regression_data, regressionResult, buyIndiaAvgReg, ws_buyAll)
            or buy_all_rule_classifier(regression_data, regressionResult, buyIndiaAvgReg, ws_buyAll)):
            print('')
        buy_all_filter(classification_data, classificationResult, None)
        buy_all_common(classification_data, classificationResult, None)
        if (buy_all_rule(classification_data, classificationResult, buyIndiaAvgCla, ws_buyAll)
            or buy_all_rule_classifier(classification_data, classificationResult, buyIndiaAvgCla, ws_buyAll)):
            print('')
    
    
    regression_data = regression_low
    classification_data = classification_low
    regressionResult = get_regressionResult(regression_data, scrip, db, regression_high['mlpValue'], regression_high['kNeighboursValue'])
    classificationResult = get_regressionResult(classification_data, scrip, db, classification_high['mlpValue'], classification_high['kNeighboursValue'])
    
    re_oi = sell_oi_candidate(regression_data, regressionResult, None)
    cl_oi = sell_oi_candidate(classification_data, classificationResult, None)
    if(cl_oi and re_oi):
        all_withoutml(regression_data, regressionResult, ws_sellOI)
        all_withoutml(classification_data, classificationResult, ws_sellOI) 
     
    sellIndiaAvgReg, result = sell_pattern_from_history(regression_data, regressionResult, None)
    sellIndiaAvgCla, result = sell_pattern_from_history(classification_data, classificationResult, None)
    if (sell_all_rule(regression_data, regressionResult, sellIndiaAvgReg, None)
        and ( sell_all_rule(classification_data, classificationResult, sellIndiaAvgCla, None)
              or sell_all_rule_classifier(classification_data, classificationResult, sellIndiaAvgCla, None)
            )
        ):
        sell_all_filter(regression_data, regressionResult, None)
        sell_all_common(regression_data, regressionResult, None)
        if (sell_all_rule(regression_data, regressionResult, sellIndiaAvgReg, ws_sellAll)
            or sell_all_rule_classifier(regression_data, regressionResult, sellIndiaAvgReg, ws_sellAll)):
            print('')
        sell_all_filter(classification_data, classificationResult, None)
        sell_all_common(classification_data, classificationResult, None)
        if (sell_all_rule(classification_data, classificationResult, sellIndiaAvgCla, ws_sellAll)
            or sell_all_rule_classifier(classification_data, classificationResult, sellIndiaAvgCla, ws_sellAll)):
            print('')                          

def result_data_reg(scrip):
    regression_high = db.regressionhigh.find_one({'scrip':scrip.replace('&','').replace('-','_')})
    regression_low = db.regressionlow.find_one({'scrip':scrip.replace('&','').replace('-','_')})
    regression_data = regression_high
    if(regression_data is not None):
        regressionResult = get_regressionResult(regression_data, scrip, db, regression_low['mlpValue'], regression_low['kNeighboursValue'])
        buy_pattern_without_mlalgo(regression_data, regressionResult, None, None)
        oi = buy_oi_candidate(regression_data, regressionResult, None)
        if oi:
            all_withoutml(regression_data, regressionResult, ws_buyOIReg)
        buyIndiaAvg, result = buy_pattern_from_history(regression_data, regressionResult, None)
        if buy_all_rule(regression_data, regressionResult, buyIndiaAvg, None):
            buy_year_high(regression_data, regressionResult, None)
            buy_year_low(regression_data, regressionResult, None, None)
            buy_final(regression_data, regressionResult, None, None)
            buy_up_trend(regression_data, regressionResult, None)
            buy_down_trend(regression_data, regressionResult, None)
            buy_oi(regression_data, regressionResult, None)
            buy_high_indicators(regression_data, regressionResult, None)
            buy_pattern(regression_data, regressionResult, None, None)
            buy_all_rule(regression_data, regressionResult, buyIndiaAvg, ws_buyAllReg)
        all_withoutml(regression_data, regressionResult, ws_highAll)
                
    regression_data = regression_low
    if(regression_data is not None):
        regressionResult = get_regressionResult(regression_data, scrip, db, regression_high['mlpValue'], regression_high['kNeighboursValue'])
        sell_pattern_without_mlalgo(regression_data, regressionResult, None, None)
        oi = sell_oi_candidate(regression_data, regressionResult, None)
        if oi:
            all_withoutml(regression_data, regressionResult, ws_sellOIReg)
        sellIndiaAvg, result = sell_pattern_from_history(regression_data, regressionResult, None)
        if sell_all_rule(regression_data, regressionResult, sellIndiaAvg, None):
            sell_year_high(regression_data, regressionResult, None, None)
            sell_year_low(regression_data, regressionResult, None)
            sell_final(regression_data, regressionResult, None, None)
            sell_up_trend(regression_data, regressionResult, None)
            #sell_down_trend(regression_data, regressionResult, None)
            sell_oi(regression_data, regressionResult, None)
            sell_high_indicators(regression_data, regressionResult, None)
            sell_pattern(regression_data, regressionResult, None, None)
            sell_all_rule(regression_data, regressionResult, sellIndiaAvg, ws_sellAllReg)                                
        all_withoutml(regression_data, regressionResult, ws_lowAll)
        
def result_data_cla(scrip):
    classification_high = db.classificationhigh.find_one({'scrip':scrip.replace('&','').replace('-','_')})
    classification_low = db.classificationlow.find_one({'scrip':scrip.replace('&','').replace('-','_')})
    
    regression_data = classification_high
    if(regression_data is not None):
        regressionResult = get_regressionResult(regression_data, scrip, db, classification_low['mlpValue'], classification_low['kNeighboursValue'])
        buy_pattern_without_mlalgo(regression_data, regressionResult, None, None)
        oi = buy_oi_candidate(regression_data, regressionResult, None)
        if oi:
            all_withoutml(regression_data, regressionResult, ws_buyOICla)
        buyIndiaAvg, result = buy_pattern_from_history(regression_data, regressionResult, None)
        if buy_all_rule(regression_data, regressionResult, buyIndiaAvg, None):
            buy_year_high(regression_data, regressionResult, None)
            buy_year_low(regression_data, regressionResult, None, None)
            buy_final(regression_data, regressionResult, None, None)
            buy_up_trend(regression_data, regressionResult, None)
            buy_down_trend(regression_data, regressionResult, None)
            #buy_oi(regression_data, regressionResult, None)
            buy_high_indicators(regression_data, regressionResult, None)
            buy_all_rule(regression_data, regressionResult, buyIndiaAvg, ws_buyAllCla)
                        
    regression_data = classification_low
    if(regression_data is not None):
        regressionResult = get_regressionResult(regression_data, scrip, db, classification_high['mlpValue'], classification_high['kNeighboursValue'])
        sell_pattern_without_mlalgo(regression_data, regressionResult, None, None)
        oi = sell_oi_candidate(regression_data, regressionResult, None)
        if oi:
            all_withoutml(regression_data, regressionResult, ws_sellOICla)
        sellIndiaAvg, result = sell_pattern_from_history(regression_data, regressionResult, None)
        if sell_all_rule(regression_data, regressionResult, sellIndiaAvg, None):
            sell_year_high(regression_data, regressionResult, None, None)
            sell_year_low(regression_data, regressionResult, None)
            sell_final(regression_data, regressionResult, None, None)
            sell_up_trend(regression_data, regressionResult, None)
            #sell_down_trend(regression_data, regressionResult, None)
            #sell_oi(regression_data, regressionResult, None)
            sell_high_indicators(regression_data, regressionResult, None)
            sell_all_rule(regression_data, regressionResult, sellIndiaAvg, ws_sellAllCla)                                 
                                             
def calculateParallel(threads=2, futures=None):
    pool = ThreadPool(threads)
    scrips = []
    for data in db.scrip.find({'futures':futures}):
        scrips.append(data['scrip'].replace('&','').replace('-','_'))
    scrips.sort()
    pool.map(result_data, scrips)
    pool.map(result_data_reg, scrips)
    pool.map(result_data_cla, scrips)
                      
if __name__ == "__main__":
    if not os.path.exists(directory):
        os.makedirs(directory)
    calculateParallel(1, sys.argv[1])
    connection.close()
    saveReports(sys.argv[1])