import os, logging, sys, json, csv
sys.path.insert(0, '../')

from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Color, PatternFill, Font, Border
from pymongo import MongoClient
from multiprocessing.dummy import Pool as ThreadPool
from bson import json_util

import math, time
import pandas as pd
import numpy as np
from talib.abstract import *

import datetime
import time
import gc
import copy
import json

from util.util import pct_change_filter, getScore, all_day_pct_change_negative, all_day_pct_change_positive, no_doji_or_spinning_buy_india, no_doji_or_spinning_sell_india, scrip_patterns_to_dict
from util.util import is_algo_buy, is_algo_sell, is_filter_all_accuracy, is_filter_hist_accuracy, is_any_reg_algo_gt1, is_any_reg_algo_lt_minus1, is_any_reg_algo_gt1_not_other, is_any_reg_algo_lt_minus1_not_other
from util.util import get_regressionResult
from util.util import buy_all_rule, buy_year_high, buy_year_low, buy_up_trend, buy_down_trend, buy_final, trend_positive
from util.util import sell_all_rule, sell_year_high, sell_year_low, sell_up_trend, sell_down_trend, sell_final, trend_negative
from util.util import buy_oi, sell_oi, all_withoutml, withoutml
from util.util import buy_other_indicator, buy_all_common, buy_all_common_High_Low, sell_other_indicator, sell_all_common, sell_all_common_High_Low
from util.util import buy_all_rule_classifier, sell_all_rule_classifier
from util.util import is_algo_buy_classifier, is_algo_sell_classifier
from util.util import sell_day_high, buy_day_low
from util.util import filter_avg_gt_2_count_any, filter_avg_lt_minus_2_count_any
from util.util import buy_filter_345_accuracy, sell_filter_345_accuracy
from util.util import buy_filter_accuracy, sell_filter_accuracy
from util.util import buy_filter_pct_change_accuracy, sell_filter_pct_change_accuracy
from util.util import buy_filter_all_accuracy, sell_filter_all_accuracy
from util.util_buy import buy_high_volatility     
from util.util_sell import sell_high_volatility
from util.util import insert_year2LowReversal, insert_year5LowBreakoutY2H, insert_year5LowBreakoutYH, insert_year2HighNearBreakout, insert_year5LowBreakoutMonthHigh, insert_year5HighNearBreakout
from util.util import historical_data



connection = MongoClient('localhost', 27017)
db = connection.Nsedata
dbchartlink = connection.chartlink

directory = '../../output/final'
logname = '../../output/final' + '/all-result' + time.strftime("%d%m%y-%H%M%S")

newsDict = {}
wb = Workbook()

ws_allFilterAcc = wb.create_sheet("AllFilterAcc")
ws_allFilterAcc.append(["BuyIndicators", "SellIndicators", "Symbol", "seriesTrend", "SMA4_2daysBack", "SMA9_2daysBack", "SMA4", "SMA9", "SMA25", "SMA50", "SMA100", "SMA200", "ResultDate", "ResultDeclared", "ResultSentiment", "ResultComment", "VOL_change", "OI_change", "Contract_change", "OI_change_next", "Contract_change_next", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "MLP_reg", "KNeighbors_reg", "MLP_cla", "KNeighbors_cla","MLP_reg_Other", "KNeighbors_reg_Other", "MLP_cla_Other", "KNeighbors_cla_Other", "forecast_mlpValue_reg", "forecast_kNeighboursValue_reg", "forecast_mlpValue_cla", "forecast_kNeighboursValue_cla", "yHigh5Change", "yLow5Change", "yHigh2Change", "yLow2Change", "yHighChange", "yLowChange", "m6HighChange", "m6LowChange", "m3HighChange", "m3LowChange", "mHighChange", "mLowChange", "w2HighChange", "w2LowChange", "wHighChange", "wLowChange", "trend", "Score", "HighTail", "LowTail", "Close", "PCT_Day_Change", "PCT_Change", "Symbol", "Industry", "Filter1", "Filter2", "Filter3", "Filter4", "Filter5", "Filter", "FilterBuy", "FilterSell", "Filter345Acc", "Filter345Count", "Filter345Pct", "Filter1Acc", "Filter1Count", "Filter1Pct", "FilterPctDayChangeAcc", "FilterPctDayChangeCount", "FilterPctDayChangePct", "FilterAllAcc", "FilterAllCount", "FilterAllPct", "FilterTechAcc", "FilterTechCount", "FilterTechPct", "FilterTechAllAcc", "FilterTechAllCount", "FilterTechAllPct", "FilterTechAllPctChangeAcc", "FilterTechAllPctChangeCount", "FilterTechAllPctChangePct","filterst_avg5","filterst_pct5","filterst_avg10","filterst_pct10","filterst_count","filter3st_avg5","filter3st_pct5","filter3st_avg10","filter3st_pct10","filter3st_count","filter4st_avg5","filter4st_pct5","filter4st_avg10","filter4st_pct10", "filter4st_count"])
ws_highBuyStrongFilterAcc = wb.create_sheet("HighBuyAllFilterAcc")
ws_highBuyStrongFilterAcc.append(["BuyIndicators", "SellIndicators", "Symbol", "seriesTrend", "SMA4_2daysBack", "SMA9_2daysBack", "SMA4", "SMA9", "SMA25", "SMA50", "SMA100", "SMA200", "ResultDate", "ResultDeclared", "ResultSentiment", "ResultComment", "VOL_change", "OI_change", "Contract_change", "OI_change_next", "Contract_change_next", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "MLP_reg", "KNeighbors_reg", "MLP_cla", "KNeighbors_cla","MLP_reg_Other", "KNeighbors_reg_Other", "MLP_cla_Other", "KNeighbors_cla_Other", "forecast_mlpValue_reg", "forecast_kNeighboursValue_reg", "forecast_mlpValue_cla", "forecast_kNeighboursValue_cla", "yHigh5Change", "yLow5Change", "yHigh2Change", "yLow2Change", "yHighChange", "yLowChange", "m6HighChange", "m6LowChange", "m3HighChange", "m3LowChange", "mHighChange", "mLowChange", "w2HighChange", "w2LowChange", "wHighChange", "wLowChange", "trend", "Score", "HighTail", "LowTail", "Close", "PCT_Day_Change", "PCT_Change", "Symbol", "Industry", "Filter1", "Filter2", "Filter3", "Filter4", "Filter5", "Filter", "FilterBuy", "FilterSell", "Filter345Acc", "Filter345Count", "Filter345Pct", "Filter1Acc", "Filter1Count", "Filter1Pct", "FilterPctDayChangeAcc", "FilterPctDayChangeCount", "FilterPctDayChangePct", "FilterAllAcc", "FilterAllCount", "FilterAllPct", "FilterTechAcc", "FilterTechCount", "FilterTechPct", "FilterTechAllAcc", "FilterTechAllCount", "FilterTechAllPct", "FilterTechAllPctChangeAcc", "FilterTechAllPctChangeCount", "FilterTechAllPctChangePct","filterst_avg5","filterst_pct5","filterst_avg10","filterst_pct10","filterst_count","filter3st_avg5","filter3st_pct5","filter3st_avg10","filter3st_pct10","filter3st_count","filter4st_avg5","filter4st_pct5","filter4st_avg10","filter4st_pct10", "filter4st_count"])
ws_lowSellStrongFilterAcc = wb.create_sheet("LowSellAllFilterAcc")
ws_lowSellStrongFilterAcc.append(["BuyIndicators", "SellIndicators", "Symbol", "seriesTrend", "SMA4_2daysBack", "SMA9_2daysBack", "SMA4", "SMA9", "SMA25", "SMA50", "SMA100", "SMA200", "ResultDate", "ResultDeclared", "ResultSentiment", "ResultComment", "VOL_change", "OI_change", "Contract_change", "OI_change_next", "Contract_change_next", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "MLP_reg", "KNeighbors_reg", "MLP_cla", "KNeighbors_cla","MLP_reg_Other", "KNeighbors_reg_Other", "MLP_cla_Other", "KNeighbors_cla_Other", "forecast_mlpValue_reg", "forecast_kNeighboursValue_reg", "forecast_mlpValue_cla", "forecast_kNeighboursValue_cla", "yHigh5Change", "yLow5Change", "yHigh2Change", "yLow2Change", "yHighChange", "yLowChange", "m6HighChange", "m6LowChange", "m3HighChange", "m3LowChange", "mHighChange", "mLowChange", "w2HighChange", "w2LowChange", "wHighChange", "wLowChange", "trend", "Score", "HighTail", "LowTail", "Close", "PCT_Day_Change", "PCT_Change", "Symbol", "Industry", "Filter1", "Filter2", "Filter3", "Filter4", "Filter5", "Filter", "FilterBuy", "FilterSell", "Filter345Acc", "Filter345Count", "Filter345Pct", "Filter1Acc", "Filter1Count", "Filter1Pct", "FilterPctDayChangeAcc", "FilterPctDayChangeCount", "FilterPctDayChangePct", "FilterAllAcc", "FilterAllCount", "FilterAllPct", "FilterTechAcc", "FilterTechCount", "FilterTechPct", "FilterTechAllAcc", "FilterTechAllCount", "FilterTechAllPct", "FilterTechAllPctChangeAcc", "FilterTechAllPctChangeCount", "FilterTechAllPctChangePct","filterst_avg5","filterst_pct5","filterst_avg10","filterst_pct10","filterst_count","filter3st_avg5","filter3st_pct5","filter3st_avg10","filter3st_pct10","filter3st_count","filter4st_avg5","filter4st_pct5","filter4st_avg10","filter4st_pct10", "filter4st_count"])
ws_highBuy = wb.create_sheet("HighBuy")
ws_highBuy.append(['scrip', 'PCT_change', 'PCT_day_change', 'ml', 'filter', 'filter2', 'filter3'])
ws_lowSell = wb.create_sheet("LowSell")
ws_lowSell.append(['scrip', 'PCT_change', 'PCT_day_change', 'ml', 'filter', 'filter2', 'filter3'])

ws_highAnalysis = wb.create_sheet("HighAnalysis")
ws_highAnalysis.append(["BuyIndicators", "SellIndicators", "Symbol", "seriesTrend", "SMA4_2daysBack", "SMA9_2daysBack", "SMA4", "SMA9", "SMA25", "SMA50", "SMA100", "SMA200", "ResultDate", "ResultDeclared", "ResultSentiment", "ResultComment", "VOL_change", "OI_change", "Contract_change", "OI_change_next", "Contract_change_next", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "MLP_reg", "KNeighbors_reg", "MLP_cla", "KNeighbors_cla","MLP_reg_Other", "KNeighbors_reg_Other", "MLP_cla_Other", "KNeighbors_cla_Other", "forecast_mlpValue_reg", "forecast_kNeighboursValue_reg", "forecast_mlpValue_cla", "forecast_kNeighboursValue_cla", "yHigh5Change", "yLow5Change", "yHigh2Change", "yLow2Change", "yHighChange", "yLowChange", "m6HighChange", "m6LowChange", "m3HighChange", "m3LowChange", "mHighChange", "mLowChange", "w2HighChange", "w2LowChange", "wHighChange", "wLowChange", "trend", "Score", "HighTail", "LowTail", "Close", "PCT_Day_Change", "PCT_Change", "Symbol", "Industry", "Filter1", "Filter2", "Filter3", "Filter4", "Filter5", "Filter", "FilterBuy", "FilterSell", "Filter345Acc", "Filter345Count", "Filter345Pct", "Filter1Acc", "Filter1Count", "Filter1Pct", "FilterPctDayChangeAcc", "FilterPctDayChangeCount", "FilterPctDayChangePct", "FilterAllAcc", "FilterAllCount", "FilterAllPct", "FilterTechAcc", "FilterTechCount", "FilterTechPct", "FilterTechAllAcc", "FilterTechAllCount", "FilterTechAllPct", "FilterTechAllPctChangeAcc", "FilterTechAllPctChangeCount", "FilterTechAllPctChangePct","filterst_avg5","filterst_pct5","filterst_avg10","filterst_pct10","filterst_count","filter3st_avg5","filter3st_pct5","filter3st_avg10","filter3st_pct10","filter3st_count","filter4st_avg5","filter4st_pct5","filter4st_avg10","filter4st_pct10", "filter4st_count"])
ws_lowAnalysis = wb.create_sheet("LowAnalysis")
ws_lowAnalysis.append(["BuyIndicators", "SellIndicators", "Symbol", "seriesTrend", "SMA4_2daysBack", "SMA9_2daysBack", "SMA4", "SMA9", "SMA25", "SMA50", "SMA100", "SMA200", "ResultDate", "ResultDeclared", "ResultSentiment", "ResultComment", "VOL_change", "OI_change", "Contract_change", "OI_change_next", "Contract_change_next", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "MLP_reg", "KNeighbors_reg", "MLP_cla", "KNeighbors_cla","MLP_reg_Other", "KNeighbors_reg_Other", "MLP_cla_Other", "KNeighbors_cla_Other", "forecast_mlpValue_reg", "forecast_kNeighboursValue_reg", "forecast_mlpValue_cla", "forecast_kNeighboursValue_cla", "yHigh5Change", "yLow5Change", "yHigh2Change", "yLow2Change", "yHighChange", "yLowChange", "m6HighChange", "m6LowChange", "m3HighChange", "m3LowChange", "mHighChange", "mLowChange", "w2HighChange", "w2LowChange", "wHighChange", "wLowChange", "trend", "Score", "HighTail", "LowTail", "Close", "PCT_Day_Change", "PCT_Change", "Symbol", "Industry", "Filter1", "Filter2", "Filter3", "Filter4", "Filter5", "Filter", "FilterBuy", "FilterSell", "Filter345Acc", "Filter345Count", "Filter345Pct", "Filter1Acc", "Filter1Count", "Filter1Pct", "FilterPctDayChangeAcc", "FilterPctDayChangeCount", "FilterPctDayChangePct", "FilterAllAcc", "FilterAllCount", "FilterAllPct", "FilterTechAcc", "FilterTechCount", "FilterTechPct", "FilterTechAllAcc", "FilterTechAllCount", "FilterTechAllPct", "FilterTechAllPctChangeAcc", "FilterTechAllPctChangeCount", "FilterTechAllPctChangePct","filterst_avg5","filterst_pct5","filterst_avg10","filterst_pct10","filterst_count","filter3st_avg5","filter3st_pct5","filter3st_avg10","filter3st_pct10","filter3st_count","filter4st_avg5","filter4st_pct5","filter4st_avg10","filter4st_pct10", "filter4st_count"])

ws_highBuyStrongBoth = wb.create_sheet("HighBuyStrongBoth")
ws_highBuyStrongBoth.append(["BuyIndicators", "SellIndicators", "Symbol", "seriesTrend", "SMA4_2daysBack", "SMA9_2daysBack", "SMA4", "SMA9", "SMA25", "SMA50", "SMA100", "SMA200", "ResultDate", "ResultDeclared", "ResultSentiment", "ResultComment", "VOL_change", "OI_change", "Contract_change", "OI_change_next", "Contract_change_next", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "MLP_reg", "KNeighbors_reg", "MLP_cla", "KNeighbors_cla","MLP_reg_Other", "KNeighbors_reg_Other", "MLP_cla_Other", "KNeighbors_cla_Other", "forecast_mlpValue_reg", "forecast_kNeighboursValue_reg", "forecast_mlpValue_cla", "forecast_kNeighboursValue_cla", "yHigh5Change", "yLow5Change", "yHigh2Change", "yLow2Change", "yHighChange", "yLowChange", "m6HighChange", "m6LowChange", "m3HighChange", "m3LowChange", "mHighChange", "mLowChange", "w2HighChange", "w2LowChange", "wHighChange", "wLowChange", "trend", "Score", "HighTail", "LowTail", "Close", "PCT_Day_Change", "PCT_Change", "Symbol", "Industry", "Filter1", "Filter2", "Filter3", "Filter4", "Filter5", "Filter", "FilterBuy", "FilterSell", "Filter345Acc", "Filter345Count", "Filter345Pct", "Filter1Acc", "Filter1Count", "Filter1Pct", "FilterPctDayChangeAcc", "FilterPctDayChangeCount", "FilterPctDayChangePct", "FilterAllAcc", "FilterAllCount", "FilterAllPct", "FilterTechAcc", "FilterTechCount", "FilterTechPct", "FilterTechAllAcc", "FilterTechAllCount", "FilterTechAllPct", "FilterTechAllPctChangeAcc", "FilterTechAllPctChangeCount", "FilterTechAllPctChangePct","filterst_avg5","filterst_pct5","filterst_avg10","filterst_pct10","filterst_count","filter3st_avg5","filter3st_pct5","filter3st_avg10","filter3st_pct10","filter3st_count","filter4st_avg5","filter4st_pct5","filter4st_avg10","filter4st_pct10", "filter4st_count"])
ws_lowSellStrongBoth = wb.create_sheet("LowSellStrongBoth")
ws_lowSellStrongBoth.append(["BuyIndicators", "SellIndicators", "Symbol", "seriesTrend", "SMA4_2daysBack", "SMA9_2daysBack", "SMA4", "SMA9", "SMA25", "SMA50", "SMA100", "SMA200", "ResultDate", "ResultDeclared", "ResultSentiment", "ResultComment", "VOL_change", "OI_change", "Contract_change", "OI_change_next", "Contract_change_next", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "MLP_reg", "KNeighbors_reg", "MLP_cla", "KNeighbors_cla","MLP_reg_Other", "KNeighbors_reg_Other", "MLP_cla_Other", "KNeighbors_cla_Other", "forecast_mlpValue_reg", "forecast_kNeighboursValue_reg", "forecast_mlpValue_cla", "forecast_kNeighboursValue_cla", "yHigh5Change", "yLow5Change", "yHigh2Change", "yLow2Change", "yHighChange", "yLowChange", "m6HighChange", "m6LowChange", "m3HighChange", "m3LowChange", "mHighChange", "mLowChange", "w2HighChange", "w2LowChange", "wHighChange", "wLowChange", "trend", "Score", "HighTail", "LowTail", "Close", "PCT_Day_Change", "PCT_Change", "Symbol", "Industry", "Filter1", "Filter2", "Filter3", "Filter4", "Filter5", "Filter", "FilterBuy", "FilterSell", "Filter345Acc", "Filter345Count", "Filter345Pct", "Filter1Acc", "Filter1Count", "Filter1Pct", "FilterPctDayChangeAcc", "FilterPctDayChangeCount", "FilterPctDayChangePct", "FilterAllAcc", "FilterAllCount", "FilterAllPct", "FilterTechAcc", "FilterTechCount", "FilterTechPct", "FilterTechAllAcc", "FilterTechAllCount", "FilterTechAllPct", "FilterTechAllPctChangeAcc", "FilterTechAllPctChangeCount", "FilterTechAllPctChangePct","filterst_avg5","filterst_pct5","filterst_avg10","filterst_pct10","filterst_count","filter3st_avg5","filter3st_pct5","filter3st_avg10","filter3st_pct10","filter3st_count","filter4st_avg5","filter4st_pct5","filter4st_avg10","filter4st_pct10", "filter4st_count"])
ws_highBuyStrong = wb.create_sheet("HighBuyStrong")
ws_highBuyStrong.append(["BuyIndicators", "SellIndicators", "Symbol", "seriesTrend", "SMA4_2daysBack", "SMA9_2daysBack", "SMA4", "SMA9", "SMA25", "SMA50", "SMA100", "SMA200", "ResultDate", "ResultDeclared", "ResultSentiment", "ResultComment", "VOL_change", "OI_change", "Contract_change", "OI_change_next", "Contract_change_next", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "MLP_reg", "KNeighbors_reg", "MLP_cla", "KNeighbors_cla","MLP_reg_Other", "KNeighbors_reg_Other", "MLP_cla_Other", "KNeighbors_cla_Other", "forecast_mlpValue_reg", "forecast_kNeighboursValue_reg", "forecast_mlpValue_cla", "forecast_kNeighboursValue_cla", "yHigh5Change", "yLow5Change", "yHigh2Change", "yLow2Change", "yHighChange", "yLowChange", "m6HighChange", "m6LowChange", "m3HighChange", "m3LowChange", "mHighChange", "mLowChange", "w2HighChange", "w2LowChange", "wHighChange", "wLowChange", "trend", "Score", "HighTail", "LowTail", "Close", "PCT_Day_Change", "PCT_Change", "Symbol", "Industry", "Filter1", "Filter2", "Filter3", "Filter4", "Filter5", "Filter", "FilterBuy", "FilterSell", "Filter345Acc", "Filter345Count", "Filter345Pct", "Filter1Acc", "Filter1Count", "Filter1Pct", "FilterPctDayChangeAcc", "FilterPctDayChangeCount", "FilterPctDayChangePct", "FilterAllAcc", "FilterAllCount", "FilterAllPct", "FilterTechAcc", "FilterTechCount", "FilterTechPct", "FilterTechAllAcc", "FilterTechAllCount", "FilterTechAllPct", "FilterTechAllPctChangeAcc", "FilterTechAllPctChangeCount", "FilterTechAllPctChangePct","filterst_avg5","filterst_pct5","filterst_avg10","filterst_pct10","filterst_count","filter3st_avg5","filter3st_pct5","filter3st_avg10","filter3st_pct10","filter3st_count","filter4st_avg5","filter4st_pct5","filter4st_avg10","filter4st_pct10", "filter4st_count"])
ws_lowSellStrong = wb.create_sheet("LowSellStrong")
ws_lowSellStrong.append(["BuyIndicators", "SellIndicators", "Symbol", "seriesTrend", "SMA4_2daysBack", "SMA9_2daysBack", "SMA4", "SMA9", "SMA25", "SMA50", "SMA100", "SMA200", "ResultDate", "ResultDeclared", "ResultSentiment", "ResultComment", "VOL_change", "OI_change", "Contract_change", "OI_change_next", "Contract_change_next", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "MLP_reg", "KNeighbors_reg", "MLP_cla", "KNeighbors_cla","MLP_reg_Other", "KNeighbors_reg_Other", "MLP_cla_Other", "KNeighbors_cla_Other", "forecast_mlpValue_reg", "forecast_kNeighboursValue_reg", "forecast_mlpValue_cla", "forecast_kNeighboursValue_cla", "yHigh5Change", "yLow5Change", "yHigh2Change", "yLow2Change", "yHighChange", "yLowChange", "m6HighChange", "m6LowChange", "m3HighChange", "m3LowChange", "mHighChange", "mLowChange", "w2HighChange", "w2LowChange", "wHighChange", "wLowChange", "trend", "Score", "HighTail", "LowTail", "Close", "PCT_Day_Change", "PCT_Change", "Symbol", "Industry", "Filter1", "Filter2", "Filter3", "Filter4", "Filter5", "Filter", "FilterBuy", "FilterSell", "Filter345Acc", "Filter345Count", "Filter345Pct", "Filter1Acc", "Filter1Count", "Filter1Pct", "FilterPctDayChangeAcc", "FilterPctDayChangeCount", "FilterPctDayChangePct", "FilterAllAcc", "FilterAllCount", "FilterAllPct", "FilterTechAcc", "FilterTechCount", "FilterTechPct", "FilterTechAllAcc", "FilterTechAllCount", "FilterTechAllPct", "FilterTechAllPctChangeAcc", "FilterTechAllPctChangeCount", "FilterTechAllPctChangePct","filterst_avg5","filterst_pct5","filterst_avg10","filterst_pct10","filterst_count","filter3st_avg5","filter3st_pct5","filter3st_avg10","filter3st_pct10","filter3st_count","filter4st_avg5","filter4st_pct5","filter4st_avg10","filter4st_pct10", "filter4st_count"])

ws_lowBuyStrong = wb.create_sheet("LowBuyStrong")
ws_lowBuyStrong.append(["BuyIndicators", "SellIndicators", "Symbol", "seriesTrend", "SMA4_2daysBack", "SMA9_2daysBack", "SMA4", "SMA9", "SMA25", "SMA50", "SMA100", "SMA200", "ResultDate", "ResultDeclared", "ResultSentiment", "ResultComment", "VOL_change", "OI_change", "Contract_change", "OI_change_next", "Contract_change_next", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "MLP_reg", "KNeighbors_reg", "MLP_cla", "KNeighbors_cla","MLP_reg_Other", "KNeighbors_reg_Other", "MLP_cla_Other", "KNeighbors_cla_Other", "forecast_mlpValue_reg", "forecast_kNeighboursValue_reg", "forecast_mlpValue_cla", "forecast_kNeighboursValue_cla", "yHigh5Change", "yLow5Change", "yHigh2Change", "yLow2Change", "yHighChange", "yLowChange", "m6HighChange", "m6LowChange", "m3HighChange", "m3LowChange", "mHighChange", "mLowChange", "w2HighChange", "w2LowChange", "wHighChange", "wLowChange", "trend", "Score", "HighTail", "LowTail", "Close", "PCT_Day_Change", "PCT_Change", "Symbol", "Industry", "Filter1", "Filter2", "Filter3", "Filter4", "Filter5", "Filter", "FilterBuy", "FilterSell", "Filter345Acc", "Filter345Count", "Filter345Pct", "Filter1Acc", "Filter1Count", "Filter1Pct", "FilterPctDayChangeAcc", "FilterPctDayChangeCount", "FilterPctDayChangePct", "FilterAllAcc", "FilterAllCount", "FilterAllPct", "FilterTechAcc", "FilterTechCount", "FilterTechPct", "FilterTechAllAcc", "FilterTechAllCount", "FilterTechAllPct", "FilterTechAllPctChangeAcc", "FilterTechAllPctChangeCount", "FilterTechAllPctChangePct","filterst_avg5","filterst_pct5","filterst_avg10","filterst_pct10","filterst_count","filter3st_avg5","filter3st_pct5","filter3st_avg10","filter3st_pct10","filter3st_count","filter4st_avg5","filter4st_pct5","filter4st_avg10","filter4st_pct10", "filter4st_count"])
ws_highSellStrong = wb.create_sheet("HighSellStrong")
ws_highSellStrong.append(["BuyIndicators", "SellIndicators", "Symbol", "seriesTrend", "SMA4_2daysBack", "SMA9_2daysBack", "SMA4", "SMA9", "SMA25", "SMA50", "SMA100", "SMA200", "ResultDate", "ResultDeclared", "ResultSentiment", "ResultComment", "VOL_change", "OI_change", "Contract_change", "OI_change_next", "Contract_change_next", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "MLP_reg", "KNeighbors_reg", "MLP_cla", "KNeighbors_cla","MLP_reg_Other", "KNeighbors_reg_Other", "MLP_cla_Other", "KNeighbors_cla_Other", "forecast_mlpValue_reg", "forecast_kNeighboursValue_reg", "forecast_mlpValue_cla", "forecast_kNeighboursValue_cla", "yHigh5Change", "yLow5Change", "yHigh2Change", "yLow2Change", "yHighChange", "yLowChange", "m6HighChange", "m6LowChange", "m3HighChange", "m3LowChange", "mHighChange", "mLowChange", "w2HighChange", "w2LowChange", "wHighChange", "wLowChange", "trend", "Score", "HighTail", "LowTail", "Close", "PCT_Day_Change", "PCT_Change", "Symbol", "Industry", "Filter1", "Filter2", "Filter3", "Filter4", "Filter5", "Filter", "FilterBuy", "FilterSell", "Filter345Acc", "Filter345Count", "Filter345Pct", "Filter1Acc", "Filter1Count", "Filter1Pct", "FilterPctDayChangeAcc", "FilterPctDayChangeCount", "FilterPctDayChangePct", "FilterAllAcc", "FilterAllCount", "FilterAllPct", "FilterTechAcc", "FilterTechCount", "FilterTechPct", "FilterTechAllAcc", "FilterTechAllCount", "FilterTechAllPct", "FilterTechAllPctChangeAcc", "FilterTechAllPctChangeCount", "FilterTechAllPctChangePct","filterst_avg5","filterst_pct5","filterst_avg10","filterst_pct10","filterst_count","filter3st_avg5","filter3st_pct5","filter3st_avg10","filter3st_pct10","filter3st_count","filter4st_avg5","filter4st_pct5","filter4st_avg10","filter4st_pct10", "filter4st_count"])
ws_highBuyReg = wb.create_sheet("HighBuyReg")
ws_highBuyReg.append(["BuyIndicators", "SellIndicators", "Symbol", "seriesTrend", "SMA4_2daysBack", "SMA9_2daysBack", "SMA4", "SMA9", "SMA25", "SMA50", "SMA100", "SMA200", "ResultDate", "ResultDeclared", "ResultSentiment", "ResultComment", "VOL_change", "OI_change", "Contract_change", "OI_change_next", "Contract_change_next", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "MLP_reg", "KNeighbors_reg", "MLP_cla", "KNeighbors_cla","MLP_reg_Other", "KNeighbors_reg_Other", "MLP_cla_Other", "KNeighbors_cla_Other", "forecast_mlpValue_reg", "forecast_kNeighboursValue_reg", "forecast_mlpValue_cla", "forecast_kNeighboursValue_cla", "yHigh5Change", "yLow5Change", "yHigh2Change", "yLow2Change", "yHighChange", "yLowChange", "m6HighChange", "m6LowChange", "m3HighChange", "m3LowChange", "mHighChange", "mLowChange", "w2HighChange", "w2LowChange", "wHighChange", "wLowChange", "trend", "Score", "HighTail", "LowTail", "Close", "PCT_Day_Change", "PCT_Change", "Symbol", "Industry", "Filter1", "Filter2", "Filter3", "Filter4", "Filter5", "Filter", "FilterBuy", "FilterSell", "Filter345Acc", "Filter345Count", "Filter345Pct", "Filter1Acc", "Filter1Count", "Filter1Pct", "FilterPctDayChangeAcc", "FilterPctDayChangeCount", "FilterPctDayChangePct", "FilterAllAcc", "FilterAllCount", "FilterAllPct", "FilterTechAcc", "FilterTechCount", "FilterTechPct", "FilterTechAllAcc", "FilterTechAllCount", "FilterTechAllPct", "FilterTechAllPctChangeAcc", "FilterTechAllPctChangeCount", "FilterTechAllPctChangePct","filterst_avg5","filterst_pct5","filterst_avg10","filterst_pct10","filterst_count","filter3st_avg5","filter3st_pct5","filter3st_avg10","filter3st_pct10","filter3st_count","filter4st_avg5","filter4st_pct5","filter4st_avg10","filter4st_pct10", "filter4st_count"])
ws_lowSellReg = wb.create_sheet("LowSellReg")
ws_lowSellReg.append(["BuyIndicators", "SellIndicators", "Symbol", "seriesTrend", "SMA4_2daysBack", "SMA9_2daysBack", "SMA4", "SMA9", "SMA25", "SMA50", "SMA100", "SMA200", "ResultDate", "ResultDeclared", "ResultSentiment", "ResultComment", "VOL_change", "OI_change", "Contract_change", "OI_change_next", "Contract_change_next", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "MLP_reg", "KNeighbors_reg", "MLP_cla", "KNeighbors_cla","MLP_reg_Other", "KNeighbors_reg_Other", "MLP_cla_Other", "KNeighbors_cla_Other", "forecast_mlpValue_reg", "forecast_kNeighboursValue_reg", "forecast_mlpValue_cla", "forecast_kNeighboursValue_cla", "yHigh5Change", "yLow5Change", "yHigh2Change", "yLow2Change", "yHighChange", "yLowChange", "m6HighChange", "m6LowChange", "m3HighChange", "m3LowChange", "mHighChange", "mLowChange", "w2HighChange", "w2LowChange", "wHighChange", "wLowChange", "trend", "Score", "HighTail", "LowTail", "Close", "PCT_Day_Change", "PCT_Change", "Symbol", "Industry", "Filter1", "Filter2", "Filter3", "Filter4", "Filter5", "Filter", "FilterBuy", "FilterSell", "Filter345Acc", "Filter345Count", "Filter345Pct", "Filter1Acc", "Filter1Count", "Filter1Pct", "FilterPctDayChangeAcc", "FilterPctDayChangeCount", "FilterPctDayChangePct", "FilterAllAcc", "FilterAllCount", "FilterAllPct", "FilterTechAcc", "FilterTechCount", "FilterTechPct", "FilterTechAllAcc", "FilterTechAllCount", "FilterTechAllPct", "FilterTechAllPctChangeAcc", "FilterTechAllPctChangeCount", "FilterTechAllPctChangePct","filterst_avg5","filterst_pct5","filterst_avg10","filterst_pct10","filterst_count","filter3st_avg5","filter3st_pct5","filter3st_avg10","filter3st_pct10","filter3st_count","filter4st_avg5","filter4st_pct5","filter4st_avg10","filter4st_pct10", "filter4st_count"])

ws_high = wb.create_sheet("HighAll")
ws_high.append(["BuyIndicators", "SellIndicators", "Symbol", "seriesTrend", "SMA4_2daysBack", "SMA9_2daysBack", "SMA4", "SMA9", "SMA25", "SMA50", "SMA100", "SMA200", "ResultDate", "ResultDeclared", "ResultSentiment", "ResultComment", "VOL_change", "OI_change", "Contract_change", "OI_change_next", "Contract_change_next", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "MLP_reg", "KNeighbors_reg", "MLP_cla", "KNeighbors_cla","MLP_reg_Other", "KNeighbors_reg_Other", "MLP_cla_Other", "KNeighbors_cla_Other", "forecast_mlpValue_reg", "forecast_kNeighboursValue_reg", "forecast_mlpValue_cla", "forecast_kNeighboursValue_cla", "yHigh5Change", "yLow5Change", "yHigh2Change", "yLow2Change", "yHighChange", "yLowChange", "m6HighChange", "m6LowChange", "m3HighChange", "m3LowChange", "mHighChange", "mLowChange", "w2HighChange", "w2LowChange", "wHighChange", "wLowChange", "trend", "Score", "HighTail", "LowTail", "Close", "PCT_Day_Change", "PCT_Change", "Symbol", "Industry", "Filter1", "Filter2", "Filter3", "Filter4", "Filter5", "Filter", "FilterBuy", "FilterSell", "Filter345Acc", "Filter345Count", "Filter345Pct", "Filter1Acc", "Filter1Count", "Filter1Pct", "FilterPctDayChangeAcc", "FilterPctDayChangeCount", "FilterPctDayChangePct", "FilterAllAcc", "FilterAllCount", "FilterAllPct", "FilterTechAcc", "FilterTechCount", "FilterTechPct", "FilterTechAllAcc", "FilterTechAllCount", "FilterTechAllPct", "FilterTechAllPctChangeAcc", "FilterTechAllPctChangeCount", "FilterTechAllPctChangePct","filterst_avg5","filterst_pct5","filterst_avg10","filterst_pct10","filterst_count","filter3st_avg5","filter3st_pct5","filter3st_avg10","filter3st_pct10","filter3st_count","filter4st_avg5","filter4st_pct5","filter4st_avg10","filter4st_pct10", "filter4st_count"])
ws_low = wb.create_sheet("LowAll")
ws_low.append(["BuyIndicators", "SellIndicators", "Symbol", "seriesTrend", "SMA4_2daysBack", "SMA9_2daysBack", "SMA4", "SMA9", "SMA25", "SMA50", "SMA100", "SMA200", "ResultDate", "ResultDeclared", "ResultSentiment", "ResultComment", "VOL_change", "OI_change", "Contract_change", "OI_change_next", "Contract_change_next", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "MLP_reg", "KNeighbors_reg", "MLP_cla", "KNeighbors_cla","MLP_reg_Other", "KNeighbors_reg_Other", "MLP_cla_Other", "KNeighbors_cla_Other", "forecast_mlpValue_reg", "forecast_kNeighboursValue_reg", "forecast_mlpValue_cla", "forecast_kNeighboursValue_cla", "yHigh5Change", "yLow5Change", "yHigh2Change", "yLow2Change", "yHighChange", "yLowChange", "m6HighChange", "m6LowChange", "m3HighChange", "m3LowChange", "mHighChange", "mLowChange", "w2HighChange", "w2LowChange", "wHighChange", "wLowChange", "trend", "Score", "HighTail", "LowTail", "Close", "PCT_Day_Change", "PCT_Change", "Symbol", "Industry", "Filter1", "Filter2", "Filter3", "Filter4", "Filter5", "Filter", "FilterBuy", "FilterSell", "Filter345Acc", "Filter345Count", "Filter345Pct", "Filter1Acc", "Filter1Count", "Filter1Pct", "FilterPctDayChangeAcc", "FilterPctDayChangeCount", "FilterPctDayChangePct", "FilterAllAcc", "FilterAllCount", "FilterAllPct", "FilterTechAcc", "FilterTechCount", "FilterTechPct", "FilterTechAllAcc", "FilterTechAllCount", "FilterTechAllPct", "FilterTechAllPctChangeAcc", "FilterTechAllPctChangeCount", "FilterTechAllPctChangePct","filterst_avg5","filterst_pct5","filterst_avg10","filterst_pct10","filterst_count","filter3st_avg5","filter3st_pct5","filter3st_avg10","filter3st_pct10","filter3st_count","filter4st_avg5","filter4st_pct5","filter4st_avg10","filter4st_pct10", "filter4st_count"])

def is_buy_filter(regression_data):
    if('%%:' in regression_data['filter']
        or "$$" in regression_data['filter']
        or 'ConsolidationBreakout' in regression_data['filter']
        or '%%HLTF:mayBuyTail-tailGT2-allDayLT0' in regression_data['filter']
        or '%%HLTF:mayBuyTail-tailGT2-7,10thDayLT0' in regression_data['filter']
        or '$$MayBuy-CheckChart(downTrend-mayReverseLast4DaysDown)' in regression_data['filter']
        or 'buyYearHigh-0' in regression_data['filter']
        or 'BuyYearLow' in regression_data['filter']
        or '%%:buyDownTrend-month3Low' in regression_data['filter']
        or 'buyFinal' in regression_data['filter']
        or 'buyMorningStar-HighLowerTail' in regression_data['filter']
        or 'sellEveningStar-0' in regression_data['filter']
        or 'checkCupUp' in regression_data['filter']
        or 'checkBuyConsolidationBreakUp' in regression_data['filter']
        or 'buyYear2LowBreakingUp' in regression_data['filter']
        or 'VOL:buy' in regression_data['filter']
        or 'VOL:Buy' in regression_data['filter']
        or regression_data['filter'].startswith('check')
        or regression_data['filter'].startswith('[MLBuy]:check')
        or regression_data['filter'].startswith('[MLSell]:check')
        or regression_data['filter'].startswith('buy')
        or regression_data['filter'].startswith('[MLBuy]:buy')
        or regression_data['filter'].startswith('[MLSell]:buy')
        or regression_data['filter'].startswith('Buy')
        or regression_data['filter'].startswith('[MLBuy]:Buy')
        or regression_data['filter'].startswith('[MLSell]:Buy')
        ):
        return True
    else:
        return False

def is_sell_filter(regression_data):
    if('%%:' in regression_data['filter']
        or "$$" in regression_data['filter']
        or 'ConsolidationBreakdown' in regression_data['filter']
        or '%%HLTF:maySellTail-tailGT2-allDayGT0' in regression_data['filter']
        or '%%HLTF:maySellTail-tailGT2-7,10thDayGT0' in regression_data['filter']
        or '$$MaySell-CheckChart(downTrend-mayReverseLast4DaysUp)' in regression_data['filter']
        or 'sellYearLow' in regression_data['filter']
        or 'sellYearHigh' in regression_data['filter']
        or '%%:sellUpTrend-month3High' in regression_data['filter']
        or 'sellFinal' in regression_data['filter']
        or 'sellEveningStar-0' in regression_data['filter']
        or 'checkCupDown' in regression_data['filter']
        or 'checkSellConsolidationBreakDown' in regression_data['filter']
        or 'VOL:sell' in regression_data['filter']
        or 'VOL:Sell' in regression_data['filter']
        or regression_data['filter'].startswith('check')
        or regression_data['filter'].startswith('[MLBuy]:check')
        or regression_data['filter'].startswith('[MLSell]:check')
        or regression_data['filter'].startswith('sell')
        or regression_data['filter'].startswith('[MLBuy]:sell')
        or regression_data['filter'].startswith('[MLSell]:sell')
        or regression_data['filter'].startswith('Sell')
        or regression_data['filter'].startswith('[MLBuy]:Sell')
        or regression_data['filter'].startswith('[MLSell]:Sell')
        ):
        return True
    else:
        return False
    

def saveReports():
    ws_allFilterAcc.append([""])
    ws_highBuyStrongFilterAcc.append([""])
    ws_lowSellStrongFilterAcc.append([""])
    
    ws_highAnalysis.append([""])
    ws_lowAnalysis.append([""])
    
    ws_lowBuyStrong.append([""])
    ws_highSellStrong.append([""])
    ws_highBuyStrong.append([""])
    ws_lowSellStrong.append([""])
    
    ws_highBuyStrongBoth.append([""])
    ws_lowSellStrongBoth.append([""])
    ws_highBuyReg.append([""])
    ws_lowSellReg.append([""])
    
    ws_high.append([""])
    ws_low.append([""])
    
    
    # Add a default style with striped rows and banded columns
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
               showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    
    count = 0
    for row in ws_allFilterAcc.iter_rows(min_row=2):
        count += 1
    tab = Table(displayName="allFilterAcc", ref="A1:DE" + str(count))
    tab.tableStyleInfo = style
    ws_allFilterAcc.add_table(tab)
    
    count = 0
    for row in ws_highBuyReg.iter_rows(min_row=2):
        count += 1
    tab = Table(displayName="highBuyReg", ref="A1:DE" + str(count))
    tab.tableStyleInfo = style
    ws_highBuyReg.add_table(tab)
    
    count = 0
    for row in ws_highBuyStrong.iter_rows(min_row=2):
        count += 1
    tab = Table(displayName="highBuyStrong", ref="A1:DE" + str(count))
    tab.tableStyleInfo = style
    ws_highBuyStrong.add_table(tab)
    
    count = 0
    for row in ws_highBuyStrongFilterAcc.iter_rows(min_row=2):
        count += 1
    tab = Table(displayName="highBuyStrongFilterAcc", ref="A1:DE" + str(count))
    tab.tableStyleInfo = style
    ws_highBuyStrongFilterAcc.add_table(tab)
    
    count = 0
    for row in ws_highAnalysis.iter_rows(min_row=2):
        count += 1
    tab = Table(displayName="highAnalysis", ref="A1:DE" + str(count))
    tab.tableStyleInfo = style
    ws_highAnalysis.add_table(tab)
    
    count = 0
    for row in ws_high.iter_rows(min_row=2):
        count += 1
    tab = Table(displayName="high", ref="A1:DE" + str(count))
    tab.tableStyleInfo = style
    ws_high.add_table(tab)
    
    
    count = 0
    for row in ws_lowSellReg.iter_rows(min_row=2):
        count += 1
    tab = Table(displayName="lowSellReg", ref="A1:DE" + str(count))
    tab.tableStyleInfo = style
    ws_lowSellReg.add_table(tab)
    
    count = 0
    for row in ws_lowSellStrong.iter_rows(min_row=2):
        count += 1
    tab = Table(displayName="lowSellStrong", ref="A1:DE" + str(count))
    tab.tableStyleInfo = style
    ws_lowSellStrong.add_table(tab)
      
    count = 0
    for row in ws_lowSellStrongFilterAcc.iter_rows(min_row=2):
        count += 1
    tab = Table(displayName="lowSellStrongFilterAcc", ref="A1:DE" + str(count))
    tab.tableStyleInfo = style
    ws_lowSellStrongFilterAcc.add_table(tab)
    
    count = 0
    for row in ws_lowAnalysis.iter_rows(min_row=2):
        count += 1
    tab = Table(displayName="lowAnalysis", ref="A1:DE" + str(count))
    tab.tableStyleInfo = style
    ws_lowAnalysis.add_table(tab)
        
    count = 0
    for row in ws_low.iter_rows(min_row=2):
        count += 1
    tab = Table(displayName="low", ref="A1:DE" + str(count))
    tab.tableStyleInfo = style
    ws_low.add_table(tab)
    
    count = 0
    for row in ws_highBuyStrongBoth.iter_rows(min_row=2):
        count += 1
    tab = Table(displayName="highBuyStrongBoth", ref="A1:DE" + str(count))
    tab.tableStyleInfo = style
    ws_highBuyStrongBoth.add_table(tab)
    
    count = 0
    for row in ws_lowSellStrongBoth.iter_rows(min_row=2):
        count += 1
    tab = Table(displayName="lowSellStrongBoth", ref="A1:DE" + str(count))
    tab.tableStyleInfo = style
    ws_lowSellStrongBoth.add_table(tab)
    
    count = 0
    for row in ws_lowBuyStrong.iter_rows(min_row=2):
        count += 1
    tab = Table(displayName="lowBuyStrong", ref="A1:DE" + str(count))
    tab.tableStyleInfo = style
    ws_lowBuyStrong.add_table(tab)
    
    count = 0
    for row in ws_highSellStrong.iter_rows(min_row=2):
        count += 1
    tab = Table(displayName="highSellStrong", ref="A1:DE" + str(count))
    tab.tableStyleInfo = style
    ws_highSellStrong.add_table(tab)
    
    count = 0
    for row in ws_highBuy.iter_rows(min_row=2):
        count += 1
    tab = Table(displayName="highBuy", ref="A1:G" + str(count))
    tab.tableStyleInfo = style
    ws_highBuy.add_table(tab)
    
    count = 0
    for row in ws_lowSell.iter_rows(min_row=2):
        count += 1
    tab = Table(displayName="lowSell", ref="A1:G" + str(count))
    tab.tableStyleInfo = style
    ws_lowSell.add_table(tab)
    
    wb.save(logname + ".xlsx")
      
def result_data(scrip):
    regression_high = db.regressionhigh.find_one({'scrip':scrip})
    regression_low = db.regressionlow.find_one({'scrip':scrip})
    if(regression_high is None or regression_low is None):
        return
    
    
    regression_data = regression_high
    regressionResult = get_regressionResult(regression_data, scrip, db, 
                                            regression_low['mlpValue_reg'], 
                                            regression_low['kNeighboursValue_reg'],
                                            regression_low['mlpValue_cla'], 
                                            regression_low['kNeighboursValue_cla'],
                                            )
    if (buy_all_rule(regression_data, regressionResult, None, None)
        and buy_all_rule_classifier(regression_data, regressionResult, None, None)
        ):
        buy_other_indicator(regression_data, regressionResult, True, None)
        buy_other_indicator(regression_data, regressionResult, False, None)
        buy_filter_all_accuracy(regression_data, regressionResult)
        buy_all_common(regression_data, regressionResult, True, None)
    if (sell_all_rule(regression_data, regressionResult, None, None)
        and sell_all_rule_classifier(regression_data, regressionResult, None, None)
        ):
        sell_other_indicator(regression_data, regressionResult, True, None)
        sell_other_indicator(regression_data, regressionResult, False, None)
        sell_filter_all_accuracy(regression_data, regressionResult)
        sell_all_common(regression_data, regressionResult, True, None)
    if (buy_all_rule(regression_data, regressionResult, None, None)
        and buy_all_rule_classifier(regression_data, regressionResult, None, None)
        ):
        all_withoutml(regression_data, regressionResult, ws_highBuyStrongBoth)
    
    regression_data = regression_low
    regressionResult = get_regressionResult(regression_data, scrip, db, 
                                            regression_high['mlpValue_reg'], 
                                            regression_high['kNeighboursValue_reg'],
                                            regression_high['mlpValue_cla'], 
                                            regression_high['kNeighboursValue_cla'],
                                            False
                                            )
    if (sell_all_rule(regression_data, regressionResult, None, None)
        and sell_all_rule_classifier(regression_data, regressionResult, None, None)
        ):
        sell_other_indicator(regression_data, regressionResult, True, None)
        sell_other_indicator(regression_data, regressionResult, False, None)
        sell_filter_all_accuracy(regression_data, regressionResult)
        sell_all_common(regression_data, regressionResult, True, None)
    if (buy_all_rule(regression_data, regressionResult, None, None)
        and buy_all_rule_classifier(regression_data, regressionResult, None, None)
        ):
        buy_other_indicator(regression_data, regressionResult, True, None)
        buy_other_indicator(regression_data, regressionResult, False, None)
        buy_filter_all_accuracy(regression_data, regressionResult)
        buy_all_common(regression_data, regressionResult, True, None)
    if (sell_all_rule(regression_data, regressionResult, None, None)
        and sell_all_rule_classifier(regression_data, regressionResult, None, None)
        ):
        all_withoutml(regression_data, regressionResult, ws_lowSellStrongBoth)


def intraday_tech_data(regression_data):
    intradaytech = ''
    scrip = regression_data['scrip']

    #if (regression_data['week2HighChange'] < 0 and regression_data['weekHighChange'] > 0):
    #    intradaytech = intradaytech + '|' + "#BYYWEEK2HIGH<LT0WeekHighGT0,"
    if(regression_data['PCT_day_change'] > 1
            or (regression_data['PCT_day_change_pre1'] > 3 and regression_data['PCT_day_change'] > -0.5)
            or (regression_data['PCT_day_change_pre2'] > 3 and regression_data['PCT_day_change'] > -0.5)
            ) :
            if (regression_data['week2HighChange'] > 0 and regression_data['weekHighChange'] > 0):
                intradaytech = intradaytech + '|' + "#BYYWEEK2HIGH>GT0,"

    #if (regression_data['week2LowChange'] > 0 and regression_data['weekLowChange'] < 0):
    #    intradaytech = intradaytech + '|' + "#SLLWEEK2LOW>GT0WeekLowLT0,"
    if (regression_data['PCT_day_change'] < -1
            or (regression_data['PCT_day_change_pre1'] < -3 and regression_data['PCT_day_change'] < -0.5)
            or (regression_data['PCT_day_change_pre2'] < -3 and regression_data['PCT_day_change'] < 0)
            ):
            if (regression_data['week2LowChange'] < 0 and regression_data['weekLowChange'] < 0):
                intradaytech = intradaytech + '|' + "#SLLWEEK2LOW<LT0,"

    if (abs(regression_data['PCT_day_change']) < 0.85
    and abs(regression_data['PCT_day_change_pre1']) < 1
    and (regression_data['close'] - regression_data['high'])/regression_data['close'] < 0.5
    and (regression_data['close'] - regression_data['high_pre1'])/regression_data['close'] < 0.3
    and (regression_data['close'] - regression_data['low'])/regression_data['close'] > -0.5
    and (regression_data['close'] - regression_data['low_pre1'])/regression_data['close'] > -0.5
    and (regression_data['close'] - regression_data['high_pre2'])/regression_data['close'] < 0
    and (regression_data['close'] - regression_data['high_pre3'])/regression_data['close'] < 0
    ):
        intradaytech = intradaytech + '|' + "Z&&&Consolidation-2Day,"
        return intradaytech
    elif (abs(regression_data['PCT_day_change']) < 1.3
    and regression_data['PCT_day_change'] < -0.5
    and regression_data['PCT_day_change_pre1'] < -1.3
    and regression_data['year5HighChange'] < 0
    and regression_data['month3LowChange'] > 30
    and (regression_data['close'] - regression_data['high'])/regression_data['close'] < 0.3
    and (regression_data['close'] - regression_data['high_pre1']) / regression_data['close'] < 0.3
    and (regression_data['close'] - regression_data['high_pre2']) / regression_data['close'] < 0
    and (regression_data['close'] - regression_data['high_pre3']) / regression_data['close'] < 0.5
    and (regression_data['close'] - regression_data['low']) / regression_data['close'] > -0.3
    and (regression_data['close'] - regression_data['low_pre1']) / regression_data['close'] > -0.3
    and ((regression_data['close'] - regression_data['low_pre2']) / regression_data['close'] > -0.75
        or (regression_data['close'] - regression_data['low_pre3']) / regression_data['close'] > -0.75
        )
    ):
        intradaytech = intradaytech + '|' + "Z&&ConsolidationHigh-3Day,"
        return intradaytech
    elif (abs(regression_data['PCT_day_change']) < 1.3
    and regression_data['PCT_day_change'] > 0.5
    and regression_data['PCT_day_change_pre1'] > 1.3
    and regression_data['year5LowChange'] > 0
    and regression_data['month3HighChange'] < -30
    and (regression_data['close'] - regression_data['high'])/regression_data['close'] < 0.3
    and (regression_data['close'] - regression_data['high_pre1']) / regression_data['close'] < 0.3
    and ((regression_data['close'] - regression_data['high_pre2']) / regression_data['close'] < 0.75
          or (regression_data['close'] - regression_data['high_pre3']) / regression_data['close'] < 0.75)
    and (regression_data['close'] - regression_data['low']) / regression_data['close'] > -0.3
    and (regression_data['close'] - regression_data['low_pre1']) / regression_data['close'] > -0.3
    and (regression_data['close'] - regression_data['low_pre2']) / regression_data['close'] > 0
    and (regression_data['close'] - regression_data['low_pre3']) / regression_data['close'] > -0.5
    ):
        intradaytech = intradaytech + '|' + "Z&&ConsolidationLow-3Day,"
        return intradaytech
    elif (abs(regression_data['PCT_day_change']) < 0.85
    and abs(regression_data['PCT_day_change_pre1']) < 1
    and abs(regression_data['PCT_day_change_pre2']) < 1
    and (abs(regression_data['month3HighChange']) > 15 or abs(regression_data['month3LowChange']) > 15)
    and (regression_data['close'] - regression_data['high'])/regression_data['close'] < 0.5
    and (regression_data['close'] - regression_data['high_pre1'])/regression_data['close'] < 0.3
    and (regression_data['close'] - regression_data['low'])/regression_data['close'] > -0.5
    and (regression_data['close'] - regression_data['low_pre1'])/regression_data['close'] > -0.5
    ):
        intradaytech = intradaytech + '|' + "ZConsolidation-last-day,"
        return intradaytech



    data = db.history15m.find_one({'dataset_code': scrip})
    if (data is None or (np.array(data['data'])).size < 25):
        if (regression_data['forecast_day_PCT2_change'] > 4 and abs(regression_data['PCT_day_change']) < 1.5):
            intradaytech = "ZPre1_uptrend"
        elif (regression_data['forecast_day_PCT3_change'] > 4 and abs(regression_data['PCT_day_change']) < 1.5):
            intradaytech = "ZPre2_uptrend"
        elif (regression_data['forecast_day_PCT4_change'] > 4 and abs(regression_data['PCT_day_change']) < 1.5):
            intradaytech = "ZPre3_uptrend"
        elif (regression_data['forecast_day_PCT5_change'] > 4 and abs(regression_data['PCT_day_change']) < 1.5):
            intradaytech = "ZPre4_uptrend"
        if (regression_data['forecast_day_PCT2_change'] < -4 and abs(regression_data['PCT_day_change']) < 1.5):
            intradaytech = "ZPre1_downtrend"
        elif (regression_data['forecast_day_PCT3_change'] < -4 and abs(regression_data['PCT_day_change']) < 1.5):
            intradaytech = "ZPre2_downtrend"
        elif (regression_data['forecast_day_PCT4_change'] < -4 and abs(regression_data['PCT_day_change']) < 1.5):
            intradaytech = "ZPre3_downtrend"
        elif (regression_data['forecast_day_PCT5_change'] < -4 and abs(regression_data['PCT_day_change']) < 1.5):
            intradaytech = "ZPre4_downtrend"
        return intradaytech

    hsdate, hsopen, hshigh, hslow, hsclose, hsquantity = historical_data(data)
    df = pd.DataFrame({
        'date': hsdate,
        'open': hsopen,
        'high': hshigh,
        'low': hslow,
        'close': hsclose,
        'volume': hsquantity
    })
    df = df[['date', 'open', 'high', 'low', 'close', 'volume']]
    df = df.rename(columns={'total trade quantity': 'volume'})
    forecast_out = 0
    open = df.tail(25).loc[-forecast_out:, 'open'].values[0]
    open_cndl12 = df.tail(12).loc[-forecast_out:, 'open'].values[0]
    open_cndl25 = df.tail(25).loc[-forecast_out:, 'open'].values[0]
    high = df.tail(1).loc[-forecast_out:, 'high'].values[0]
    high_cndl12 = df.tail(12).loc[-forecast_out:, 'high'].values[0]
    high_cndl25 = df.tail(25).loc[-forecast_out:, 'high'].values[0]
    low = df.tail(1).loc[-forecast_out:, 'low'].values[0]
    low_cndl12 = df.tail(12).loc[-forecast_out:, 'low'].values[0]
    low_cndl25 = df.tail(25).loc[-forecast_out:, 'low'].values[0]
    close = df.tail(1).loc[-forecast_out:, 'close'].values[0]
    close_cndl12 = df.tail(12).loc[-forecast_out:, 'close'].values[0]
    close_cndl25 = df.tail(25).loc[-forecast_out:, 'close'].values[0]
    daychange = (close - open_cndl25) * 100 / open_cndl25
    postlunchchange_high = (high - high_cndl12) * 100 / high_cndl12
    morningchange_high = (high_cndl12 - high_cndl25) * 100 / high_cndl25
    postlunchchange_low = (low - low_cndl12) * 100 / low_cndl12
    morningchange_low = (low_cndl12 - low_cndl25) * 100 / low_cndl25


    if ( daychange > 1 and (daychange < 4 or (regression_data['PCT_day_change_pre1'] < 5 and daychange < 7)) and (postlunchchange_high > daychange / 3) and (morningchange_high > daychange / 4)):
        if (1 < regression_data['PCT_day_change_pre1']):
            intradaytech = "UpStairs(SellAt10-Last2DayUP)"
        else:
            intradaytech = "UpStairs"
    elif (daychange > 2
        and abs(postlunchchange_high) < abs(daychange) / 3
        and close > high_cndl25
        ):
        if (1.5 < regression_data['PCT_day_change_pre1']):
            intradaytech = "UpPostLunchConsolidation(Last2DayUP)"
        else:
            intradaytech = "UpPostLunchConsolidation"
    elif (daychange > 2
        and abs(morningchange_high) < abs(daychange) / 3
        and close > high_cndl12
        ):
        if (1.5 < regression_data['PCT_day_change_pre1']):
            intradaytech = "UpMorningConsolidation(Last2DayUP)"
        else:
            intradaytech = "UpMorningConsolidation"


    if (daychange < -1 and (daychange > -4 or (regression_data['PCT_day_change_pre1'] > -5 and daychange > -7)) and (postlunchchange_low < daychange / 3) and (morningchange_low < daychange / 4)):
        if (regression_data['PCT_day_change_pre1'] < -1):
            intradaytech = "DownStairs(BuyAt10-Last2DayDown)"
        else:
            intradaytech = "DownStairs"
    elif (daychange < -2
        and abs(postlunchchange_low) < abs(daychange) / 3
        and close < low_cndl25
        ):
        if (regression_data['PCT_day_change_pre1'] < -1.5):
            intradaytech = "DownPostLunchConsolidation(Last2DayDown)"
        else:
            intradaytech = "DownPostLunchConsolidation"
    elif (daychange < -2
        and abs(morningchange_low) < abs(daychange) / 3
        and close < low_cndl12
        ):
        if (regression_data['PCT_day_change_pre1'] < -1.5):
            intradaytech = "DownMorningConsolidation(Last2DayDown)"
        else:
            intradaytech = "DownMorningConsolidation"

    if ('Up' in intradaytech and regression_data['year5HighChange'] < -20 and regression_data['year2HighChange'] > -5
        and regression_data['highTail'] < 1
        ):
        intradaytech = intradaytech + ':AtYear2High:Year5HighChangeLT-30'
    elif ('Up' in intradaytech and regression_data['year5HighChange'] < -35 and regression_data['year2HighChange'] > -50
        and regression_data['highTail'] < 1
        ):
        intradaytech = intradaytech + ':Year5HighChangeLT-30'
    if ('Down' in intradaytech and regression_data['year5LowChange'] > 20 and regression_data['year2LowChange'] < 5
        and regression_data['lowTail'] < 1
        ):
        intradaytech = intradaytech + ':AtYear2Low:Year5LowChangeGT30'
    elif ('Down' in intradaytech and regression_data['year5LowChange'] > 35 and regression_data['year2LowChange'] < 50
        and regression_data['lowTail'] < 1
        ):
        intradaytech = intradaytech + ':Year5LowChangeGT30'
    
    if(intradaytech != ''):
        return intradaytech

    open_pre1 = df.tail(50).loc[-forecast_out:, 'open'].values[0]
    open_cndl12_pre1 = df.tail(37).loc[-forecast_out:, 'open'].values[0]
    open_cndl25_pre1 = df.tail(50).loc[-forecast_out:, 'open'].values[0]
    high_pre1 = df.tail(26).loc[-forecast_out:, 'high'].values[0]
    high_cndl12_pre1 = df.tail(37).loc[-forecast_out:, 'high'].values[0]
    high_cndl25_pre1 = df.tail(50).loc[-forecast_out:, 'high'].values[0]
    low_pre1 = df.tail(26).loc[-forecast_out:, 'low'].values[0]
    low_cndl12_pre1 = df.tail(37).loc[-forecast_out:, 'low'].values[0]
    low_cndl25_pre1 = df.tail(50).loc[-forecast_out:, 'low'].values[0]
    close_pre1 = df.tail(26).loc[-forecast_out:, 'close'].values[0]
    close_cndl12_pre1 = df.tail(37).loc[-forecast_out:, 'close'].values[0]
    close_cndl25_pre1 = df.tail(50).loc[-forecast_out:, 'close'].values[0]
    daychange_pre1 = (close_pre1 - open_cndl25_pre1) * 100 / open_cndl25_pre1
    postlunchchange_high_pre1 = (high_pre1 - high_cndl12_pre1) * 100 / high_cndl12_pre1
    morningchange_high_pre1 = (high_cndl12_pre1 - high_cndl25_pre1) * 100 / high_cndl25_pre1
    postlunchchange_low_pre1 = (low_pre1 - low_cndl12_pre1) * 100 / low_cndl12_pre1
    morningchange_low_pre1 = (low_cndl12_pre1 - low_cndl25_pre1) * 100 / low_cndl25_pre1

    if (intradaytech == '' and -1 < daychange < 1 and abs(regression_data['PCT_day_change']) < 1 and abs(daychange_pre1) > 3 * abs(daychange)):
        if ( 1.5 < daychange_pre1 < 5 and (postlunchchange_high_pre1 > daychange / 3) and (morningchange_high_pre1 > daychange / 4)):
            intradaytech = "ZPre1_UpStairs"
        elif (1.5 < daychange_pre1 < 5 and abs(postlunchchange_high_pre1) < abs(daychange) / 4):
            intradaytech = "ZPre1_UpPostLunchConsolidation"

        if (-5 < daychange_pre1 < -1.5 and (postlunchchange_low_pre1 < daychange / 3) and (morningchange_low_pre1 > daychange / 4)):
            intradaytech = "ZPre1_DownStairs"
        elif (-5 < daychange_pre1 < -1.5 and abs(postlunchchange_low_pre1) < abs(daychange) / 4):
            intradaytech = "ZPre1_DownPostLunchConsolidation"
    if (intradaytech == '' and -1 < daychange < 1 and abs(regression_data['PCT_day_change']) < 1 and abs(daychange_pre1) > 3 * abs(daychange)):
        if ( 1 < daychange_pre1 < 5 and postlunchchange_high_pre1 > 1 and morningchange_high_pre1 > 1):
            intradaytech = "ZPre1_UpStairs1"
        if (-5 < daychange_pre1 < -1 and postlunchchange_low_pre1 < -1 and morningchange_low_pre1 < -1):
            intradaytech = "ZPre1_DownStairs1"
    if (intradaytech == '' and -1 < daychange < 1 and abs(regression_data['PCT_day_change']) < 1 and abs(daychange_pre1) > 3 * abs(daychange)):
        if ( regression_data['PCT_day_change_pre1'] > 2.5 and postlunchchange_high_pre1 > 1.5):
            intradaytech = "ZPre1_UpStairs***"
        if (regression_data['PCT_day_change_pre1'] < -2.5 and postlunchchange_low_pre1 < -1.5):
            intradaytech = "ZPre1_DownStairs***"
    if (intradaytech == '' and -1 < daychange < 1 and abs(regression_data['PCT_day_change']) < 1 and abs(daychange_pre1) > 3 * abs(daychange)):
        if ( daychange_pre1 > 5 and postlunchchange_high_pre1 > 1 and morningchange_high_pre1 > 1):
            intradaytech = "ZPre1_HighUpStairs1"
        if (daychange_pre1 < -5 and postlunchchange_low_pre1 < -1 and morningchange_low_pre1 < -1):
            intradaytech = "ZPre1_HighDownStairs1"

    if ('Up' in intradaytech and regression_data['year5HighChange'] < -20 and regression_data['year2HighChange'] > -5
        and regression_data['highTail_pre1'] < 1 and regression_data['highTail'] < 2
        ):
        intradaytech = intradaytech + ':AtYear2High:Year5HighChangeLT-30'
    elif ('Up' in intradaytech and regression_data['year5HighChange'] < -35 and regression_data['year2HighChange'] > -50
        and regression_data['highTail_pre1'] < 1 and regression_data['highTail'] < 2
        ):
        intradaytech = intradaytech + ':Year5HighChangeLT-30'
    if ('Down' in intradaytech and regression_data['year5LowChange'] > 20 and regression_data['year2LowChange'] < 5
        and regression_data['lowTail_pre1'] < 1 and regression_data['lowTail'] < 2
        ):
        intradaytech = intradaytech + ':AtYear2Low:Year5LowChangeGT30'
    elif ('Down' in intradaytech and regression_data['year5LowChange'] > 35 and regression_data['year2LowChange'] < 50
        and regression_data['lowTail_pre1'] < 1 and regression_data['lowTail'] < 2
        ):
        intradaytech = intradaytech + ':Year5LowChangeGT30'


    if (intradaytech != ''):
        return intradaytech

    open_pre2 = df.tail(75).loc[-forecast_out:, 'open'].values[0]
    open_cndl12_pre2 = df.tail(62).loc[-forecast_out:, 'open'].values[0]
    open_cndl25_pre2 = df.tail(75).loc[-forecast_out:, 'open'].values[0]
    high_pre2 = df.tail(51).loc[-forecast_out:, 'high'].values[0]
    high_cndl12_pre2 = df.tail(62).loc[-forecast_out:, 'high'].values[0]
    high_cndl25_pre2 = df.tail(75).loc[-forecast_out:, 'high'].values[0]
    low_pre2 = df.tail(51).loc[-forecast_out:, 'low'].values[0]
    low_cndl12_pre2 = df.tail(62).loc[-forecast_out:, 'low'].values[0]
    low_cndl25_pre2 = df.tail(75).loc[-forecast_out:, 'low'].values[0]
    close_pre2 = df.tail(51).loc[-forecast_out:, 'close'].values[0]
    close_cndl12_pre2 = df.tail(62).loc[-forecast_out:, 'close'].values[0]
    close_cndl25_pre2 = df.tail(75).loc[-forecast_out:, 'close'].values[0]
    daychange_pre2 = (close_pre2 - open_cndl25_pre2) * 100 / open_cndl25_pre2
    postlunchchange_high_pre2 = (high_pre2 - high_cndl12_pre2) * 100 / high_cndl12_pre2
    morningchange_high_pre2 = (high_cndl12_pre2 - high_cndl25_pre2) * 100 / high_cndl25_pre2
    postlunchchange_low_pre2 = (low_pre2 - low_cndl12_pre2) * 100 / low_cndl12_pre2
    morningchange_low_pre2 = (low_cndl12_pre2 - low_cndl25_pre2) * 100 / low_cndl25_pre2

    open_pre3 = df.tail(100).loc[-forecast_out:, 'open'].values[0]
    open_cndl12_pre3 = df.tail(87).loc[-forecast_out:, 'open'].values[0]
    open_cndl25_pre3 = df.tail(100).loc[-forecast_out:, 'open'].values[0]
    high_pre3 = df.tail(76).loc[-forecast_out:, 'high'].values[0]
    high_cndl12_pre3 = df.tail(87).loc[-forecast_out:, 'high'].values[0]
    high_cndl25_pre3 = df.tail(100).loc[-forecast_out:, 'high'].values[0]
    low_pre3 = df.tail(76).loc[-forecast_out:, 'low'].values[0]
    low_cndl12_pre3 = df.tail(87).loc[-forecast_out:, 'low'].values[0]
    low_cndl25_pre3 = df.tail(100).loc[-forecast_out:, 'low'].values[0]
    close_pre3 = df.tail(76).loc[-forecast_out:, 'close'].values[0]
    close_cndl12_pre3 = df.tail(87).loc[-forecast_out:, 'close'].values[0]
    close_cndl25_pre3 = df.tail(100).loc[-forecast_out:, 'close'].values[0]
    daychange_pre3 = (close_pre3 - open_cndl25_pre3) * 100 / open_cndl25_pre3
    postlunchchange_high_pre3 = (high_pre3 - high_cndl12_pre3) * 100 / high_cndl12_pre3
    morningchange_high_pre3 = (high_cndl12_pre3 - high_cndl25_pre3) * 100 / high_cndl25_pre3
    postlunchchange_low_pre3 = (low_pre3 - low_cndl12_pre3) * 100 / low_cndl12_pre3
    morningchange_low_pre3 = (low_cndl12_pre3 - low_cndl25_pre3) * 100 / low_cndl25_pre3

    open_pre4 = df.tail(125).loc[-forecast_out:, 'open'].values[0]
    open_cndl12_pre4 = df.tail(112).loc[-forecast_out:, 'open'].values[0]
    open_cndl25_pre4 = df.tail(125).loc[-forecast_out:, 'open'].values[0]
    high_pre4 = df.tail(101).loc[-forecast_out:, 'high'].values[0]
    high_cndl12_pre4 = df.tail(112).loc[-forecast_out:, 'high'].values[0]
    high_cndl25_pre4 = df.tail(125).loc[-forecast_out:, 'high'].values[0]
    low_pre4 = df.tail(101).loc[-forecast_out:, 'low'].values[0]
    low_cndl12_pre4 = df.tail(112).loc[-forecast_out:, 'low'].values[0]
    low_cndl25_pre4 = df.tail(125).loc[-forecast_out:, 'low'].values[0]
    close_pre4 = df.tail(101).loc[-forecast_out:, 'close'].values[0]
    close_cndl12_pre4 = df.tail(112).loc[-forecast_out:, 'close'].values[0]
    close_cndl25_pre4 = df.tail(125).loc[-forecast_out:, 'close'].values[0]
    daychange_pre4 = (close_pre4 - open_cndl25_pre4) * 100 / open_cndl25_pre4
    postlunchchange_high_pre4 = (high_pre4 - high_cndl12_pre4) * 100 / high_cndl12_pre4
    morningchange_high_pre4 = (high_cndl12_pre4 - high_cndl25_pre4) * 100 / high_cndl25_pre4
    postlunchchange_low_pre4 = (low_pre4 - low_cndl12_pre4) * 100 / low_cndl12_pre4
    morningchange_low_pre4 = (low_cndl12_pre4 - low_cndl25_pre4) * 100 / low_cndl25_pre4

    if (intradaytech == '' and abs(daychange) < 2
        and abs(daychange_pre1) < 1 and abs(daychange_pre2) < 1 and abs(daychange_pre3) < 1
        and abs(regression_data['PCT_change_pre5']) > 2 * abs(daychange_pre4) and (regression_data['PCT_change_pre5']) > 2 * abs(daychange_pre3) and abs(regression_data['PCT_change_pre5']) > 2 * abs(daychange_pre2) and abs(regression_data['PCT_change_pre5']) > 2 * abs(daychange_pre1)
        and abs(regression_data['PCT_change_pre5']) > 2 * abs(regression_data['PCT_change'])
        and abs(regression_data['PCT_day_change_pre5']) > 5
        ):
        if (regression_data['low'] > regression_data['low_pre5'] and regression_data['low'] > regression_data['open_pre5']
            and regression_data['highTail'] < 0.5 and regression_data['lowTail'] > 0.7
            ):
            intradaytech = "ZPre5_UpStairs"
        if (regression_data['high'] < regression_data['high_pre5'] and regression_data['high'] < regression_data['open_pre5']
            and regression_data['lowTail'] < 0.5 and regression_data['highTail'] > 0.7
            ):
            intradaytech = "ZPre5_DownStairs"

    if (intradaytech == '' and abs(daychange) < 1.5
        and abs(daychange_pre1) < 1 and abs(daychange_pre2) < 1
        and abs(daychange_pre4) > 2 * abs(daychange_pre3) and abs(daychange_pre4) > 2 * abs(daychange_pre2) and abs(daychange_pre4) > 2 * abs(daychange_pre1)
        and abs(regression_data['PCT_change_pre4']) > 2 * abs(regression_data['PCT_change'])
        and abs(daychange_pre4) > 5
        ):
        if (regression_data['low'] > regression_data['low_pre4'] and regression_data['low'] > regression_data['open_pre4']
            and regression_data['highTail'] < 0.5 and regression_data['lowTail'] > 0.7
            ):
            if (daychange_pre4 > 2.5 and (postlunchchange_high_pre4 > daychange / 3) and (morningchange_high_pre4 > daychange / 4)):
                intradaytech = "ZPre4_UpStairs"
            elif (daychange_pre4 > 2.5 and abs(postlunchchange_high_pre4) < abs(daychange) / 4):
                intradaytech = "Pre4_UpPostLunchConsolidation"
        if (regression_data['high'] < regression_data['high_pre4'] and regression_data['high'] < regression_data['open_pre4']
            and regression_data['lowTail'] < 0.5 and regression_data['highTail'] > 0.7
            ):
            if (daychange_pre4 < -2.5 and (postlunchchange_low_pre4 < daychange / 3) and (morningchange_low_pre4 < daychange / 4)):
                intradaytech = "ZPre4_DownStairs"
            elif (daychange_pre4 < -2.5 and abs(postlunchchange_low_pre4) > abs(daychange) / 4):
                intradaytech = "Pre4_DownPostLunchConsolidation"

    if (intradaytech == '' and abs(daychange) < 1.5
        and abs(daychange_pre1) < 1
        and abs(daychange_pre3) > 2 * abs(daychange_pre2) and abs(daychange_pre3) > 2 * abs(daychange_pre1)
        and abs(regression_data['PCT_change_pre3']) > 2 * abs(regression_data['PCT_change'])
        and abs(daychange_pre3) > 5):
        if (regression_data['low'] > regression_data['low_pre3'] and regression_data['low'] > regression_data['open_pre3']
            and regression_data['highTail'] < 0.5 and regression_data['lowTail'] > 0.7
            ):
            if (daychange_pre3 > 2 and (postlunchchange_high_pre3 > daychange / 3) and (morningchange_high_pre3 > daychange / 4)):
                intradaytech = "ZPre3_UpStairs"
            elif (daychange_pre3 > 2 and abs(postlunchchange_high_pre3) < abs(daychange) / 4):
                intradaytech = "Pre3_UpPostLunchConsolidation"
        if (regression_data['high'] < regression_data['high_pre3'] and regression_data['high'] < regression_data['open_pre3']
            and regression_data['lowTail'] < 0.5 and regression_data['highTail'] > 0.7
            ):
            if (daychange_pre3 < -2 and (postlunchchange_low_pre3 < daychange / 3) and (morningchange_low_pre3 < daychange / 4)):
                intradaytech = "ZPre3_DownStairs"
            elif (daychange_pre3 < -2 and abs(postlunchchange_low_pre3) > abs(daychange) / 4):
                intradaytech = "Pre3_DownPostLunchConsolidation"


    if (intradaytech == '' and abs(daychange) < 2
        and abs(daychange_pre1) < 1
        and abs(daychange_pre2) > 2 * abs(daychange_pre1)
        and abs(regression_data['PCT_change_pre2']) > 2 * abs(regression_data['PCT_change'])
        and abs(daychange_pre2) > 2
        ):
        if(regression_data['low'] > regression_data['low_pre2'] and regression_data['low'] > regression_data['open_pre2']):
            if (daychange_pre2 > 2 and (postlunchchange_high_pre2 > daychange / 3) and (
                    morningchange_high_pre2 > daychange / 4)):
                intradaytech = "ZPre2_UpStairs"
            elif (daychange_pre2 > 2 and abs(postlunchchange_high_pre2) < abs(daychange) / 4):
                intradaytech = "Pre2_UpPostLunchConsolidation"
        if (regression_data['high'] < regression_data['high_pre2'] and regression_data['high'] < regression_data['open_pre2']):
            if (daychange_pre2 < -2 and (postlunchchange_low_pre2 < daychange / 3) and (
                    morningchange_low_pre2 > daychange / 4)):
                intradaytech = "ZPre2_DownStairs"
            elif (daychange_pre2 < -2 and abs(postlunchchange_low_pre2) < abs(daychange) / 4):
                intradaytech = "Pre2_DownPostLunchConsolidation"

    if ('Up' in intradaytech and regression_data['year5HighChange'] < -30 and regression_data['year2HighChange'] > -5):
        intradaytech = intradaytech + ':AtYear2High:Year5HighChangeLT-30'
    elif ('Up' in intradaytech and regression_data['year5HighChange'] < -35 and regression_data['year2HighChange'] > -50):
        intradaytech = intradaytech + ':Year5HighChangeLT-30'
    if ('Down' in intradaytech and regression_data['year5LowChange'] > 35 and regression_data['year2LowChange'] < 5):
        intradaytech = intradaytech + ':AtYear2Low:Year5LowChangeGT30'
    elif ('Down' in intradaytech and regression_data['year5LowChange'] > 35 and regression_data['year2LowChange'] < 50):
        intradaytech = intradaytech + ':Year5LowChangeGT30'



    return intradaytech

def shortterm_tech_data_buy(regressionhigh, regressionlow):
    datahigh = ''

    if (regressionhigh['PCT_day_change'] > 1
            or (regressionhigh['PCT_day_change_pre1'] > 3 and regressionhigh['PCT_day_change'] > -0.5)
            or (regressionhigh['PCT_day_change_pre2'] > 3 and regressionhigh['PCT_day_change'] > -0.5)
            ):
            if (regressionhigh['week2HighChange'] < 0 and regressionhigh['weekHighChange'] > 0):
                datahigh = datahigh + '|' + "#week2high<LT0WeekHighGT0,"
            #elif (regressionhigh['week2HighChange'] > 0 and regressionhigh['weekHighChange'] > 0):
            #    datahigh = datahigh + '|' + "#BYYWEEK2HIGH>GT0,"

    if ('shortUpTrend' in regressionhigh['filter1'] and '(upTrend)' in regressionhigh['series_trend']
        and abs(regressionhigh['month3LowChange']) < abs(regressionhigh['month3HighChange'])
        ):
        datahigh = datahigh + '|' + 'shortUpTrend'

    if (regressionhigh['year5HighChange'] < -30 and regressionhigh['yearHighChange'] > -5):
        datahigh = 'AtYearHighLT-30' + '|' + datahigh

    return datahigh

def shortterm_tech_data_sell(regressionhigh, regressionlow):
    datalow = ''

    if (regressionlow['PCT_day_change'] < -1
            or (regressionlow['PCT_day_change_pre1'] < -3 and regressionlow['PCT_day_change'] < 0.5)
            or (regressionlow['PCT_day_change_pre2'] < -3 and regressionlow['PCT_day_change'] < 0.5)
            ):
            if (regressionlow['week2LowChange'] > 0 and regressionlow['weekLowChange'] < 0):
                datalow = datalow + '|' + "week2low>GT0WeekLowLT0,"
            #elif (regressionlow['week2LowChange'] < 0 and regressionlow['weekLowChange'] < 0):
            #    datalow = datalow + '|' + "##SLLWEEK2LOW<LT0,"

    if ('shorDownTrend' in regressionlow['filter1'] and '(downTrend)' in regressionlow['series_trend']
        and abs(regressionlow['month3LowChange']) > abs(regressionlow['month3HighChange'])
        ):
        datalow = datalow + '|' + 'shortDownTrend'

    if (regressionlow['year5LowChange'] > 30 and regressionlow['yearLowChange'] < 10):
        datalow = 'AtYearLow-GT30' + '|' + datalow

    return datalow


def get_index(scrip):
    data = db.scrip_futures.find_one({'scrip': scrip})
    if (data is not None):
        return "futures"
    else:
        return ""

def get_fundamental(scrip):
    data = db.fundamental.find_one({'scrip': scrip})
    if (data is not None):
        return data
    else:
        return ""


def result_data_summary(scrip):
    regression_high = db.highBuy.find_one({'scrip':scrip})
    regression_low = db.lowSell.find_one({'scrip':scrip})
    if(regression_high is not None):
        withoutml(regression_high, ws_highBuy)
    
    if(regression_low is not None):
        withoutml(regression_low, ws_lowSell)
                                                        
def result_data_reg(scrip):
    regression_high = db.regressionhigh.find_one({'scrip':scrip})
    regression_low = db.regressionlow.find_one({'scrip':scrip})
    if(regression_high is None or regression_low is None):
        return
    
    regression_data = regression_high
    if(regression_data is not None):
        regressionResultHigh = get_regressionResult(regression_data, scrip, db, 
                                            regression_low['mlpValue_reg'], 
                                            regression_low['kNeighboursValue_reg'],
                                            regression_low['mlpValue_cla'], 
                                            regression_low['kNeighboursValue_cla'],
                                            )
        buy_other_indicator(regression_data, regressionResultHigh, True, None)
        buy_filter_all_accuracy(regression_data, regressionResultHigh)
        
        if(is_filter_all_accuracy(regression_data, regression_high, regression_low, regressionResultHigh, 'High', None)):
            all_withoutml(regression_data, regressionResultHigh, None)
        buy_all_rule(regression_data, regressionResultHigh, True, None)
        #if (is_algo_buy(regression_data)):
            #all_withoutml(regression_data, regressionResultHigh, ws_highBuyReg)
        if (is_algo_buy(regression_high) and is_any_reg_algo_gt1_not_other(regression_data)):
            buy_all_common_High_Low(regression_data, regressionResultHigh, True, None)
            #all_withoutml(regression_data, regressionResultHigh, ws_highBuyStrong) 
        if (is_algo_buy(regression_high, True) and is_algo_buy(regression_low, True) and is_any_reg_algo_gt1(regression_data)):
            buy_all_common_High_Low(regression_data, regressionResultHigh, True, None)
            #all_withoutml(regression_data, regressionResultHigh, ws_lowBuyStrong)
        if (is_algo_buy(regression_high, True) and is_algo_buy(regression_low, True) 
            and is_any_reg_algo_gt1(regression_data)
            and is_any_reg_algo_gt1_not_other(regression_data)
            ):
            buy_all_common_High_Low(regression_data, regressionResultHigh, True, None)
            #all_withoutml(regression_data, regressionResultHigh, ws_highBuyStrongBoth)
         
        all_withoutml(regression_data, regressionResultHigh, ws_high)
        if (db['highBuy'].find_one({'scrip': regression_data['scrip']}) is None):
            regression_data['ml'] = ''
            db['highBuy'].insert_one(regression_data)
        
          
                
    regression_data = regression_low
    if(regression_data is not None):
        regressionResultLow = get_regressionResult(regression_data, scrip, db, 
                                            regression_high['mlpValue_reg'], 
                                            regression_high['kNeighboursValue_reg'],
                                            regression_high['mlpValue_cla'], 
                                            regression_high['kNeighboursValue_cla'],
                                            False
                                            )
        sell_other_indicator(regression_data, regressionResultLow, True, None)
        sell_filter_all_accuracy(regression_data, regressionResultLow)
        
        if(is_filter_all_accuracy(regression_data, regression_high, regression_low, regressionResultLow, 'Low', None)):
            all_withoutml(regression_data, regressionResultLow, None)
        sell_all_rule(regression_data, regressionResultLow, True, None)
        #if (is_algo_sell(regression_data)):
            #all_withoutml(regression_data, regressionResultLow, ws_lowSellReg)                               
        if (is_algo_sell(regression_high) and is_any_reg_algo_lt_minus1_not_other(regression_data)):
            sell_all_common_High_Low(regression_data, regressionResultLow, True, None)
            #all_withoutml(regression_data, regressionResultLow, ws_lowSellStrong)
        if (is_algo_sell(regression_high, True) and is_algo_sell(regression_low, True) and is_any_reg_algo_lt_minus1(regression_data)):
            sell_all_common_High_Low(regression_data, regressionResultLow, True, None)
            #all_withoutml(regression_data, regressionResultLow, ws_highSellStrong)
        if (is_algo_sell(regression_high, True) and is_algo_sell(regression_low, True) 
            and is_any_reg_algo_lt_minus1(regression_data)
            and is_any_reg_algo_lt_minus1_not_other(regression_data)
            ):
            sell_all_common_High_Low(regression_data, regressionResultLow, True, None)
            #all_withoutml(regression_data, regressionResultLow, ws_lowSellStrongBoth)
        
        all_withoutml(regression_data, regressionResultLow, ws_low)
        if (db['lowSell'].find_one({'scrip': regression_data['scrip']}) is None):
            regression_data['ml'] = ''
            db['lowSell'].insert_one(regression_data)
        
    
        regression_data = regression_high
        reg_data_filter = regression_data['filter']
        list = regression_data['filter'].partition(']:')
        lt_minus_2_count_any, lt_minus_2_cnt, lt_minus_2_max = filter_avg_lt_minus_2_count_any(regression_data)
        gt_2_count_any, gt_2_cnt, gt_2_max = filter_avg_gt_2_count_any(regression_data)
        if(list[1] == ']:'):
            reg_data_filter = list[2]
        # if ((is_algo_buy(regression_high) == True and 'Buy-AnyGT2' in regression_data['filter2'] and ('buy' in reg_data_filter or 'Buy' in reg_data_filter))
        #     or (is_algo_sell(regression_high) == True and 'Sell-AnyGT2' in regression_data['filter2'] and ('sell' in reg_data_filter or 'Sell' in reg_data_filter))
        #     or ((gt_2_count_any > 1 or ('Buy-SUPER' in regression_data['filter2'])) and gt_2_cnt > 1)
        #     or ((lt_minus_2_count_any > 1 or ('Sell-SUPER' in regression_data['filter2'])) and lt_minus_2_cnt > 1)
        #     ):
        #     all_withoutml(regression_data, regressionResultHigh, ws_highBuy)
        regression_data = regression_low
        reg_data_filter = regression_data['filter']
        list = regression_data['filter'].partition(']:')
        if(list[1] == ']:'):
            reg_data_filter = list[2]
        # if ((is_algo_buy(regression_low) == True and 'Buy-AnyGT2' in regression_data['filter2'] and  ('buy' in reg_data_filter or 'Buy' in reg_data_filter))
        #     or (is_algo_sell(regression_low) == True and 'Sell-AnyGT2' in regression_data['filter2'] and  ('sell' in reg_data_filter or 'Sell' in reg_data_filter))
        #     or ((gt_2_count_any > 1 or ('Buy-SUPER' in regression_data['filter2'])) and gt_2_cnt > 1)
        #     or ((lt_minus_2_count_any > 1 or ('Sell-SUPER' in regression_data['filter2'])) and lt_minus_2_cnt > 1)
        #     ):
        #     all_withoutml(regression_data, regressionResultLow, ws_lowSell)
            
        
        regression_high_copy = copy.deepcopy(regression_high)
        regression_high_copy1 = copy.deepcopy(regression_high)
        regression_high_copy2 = copy.deepcopy(regression_high)
        regression_low_copy = copy.deepcopy(regression_low)
        regression_low_copy1 = copy.deepcopy(regression_low)
        regression_low_copy2 = copy.deepcopy(regression_low)
        
        regression_data = regression_high_copy2
        if(buy_high_volatility(regression_high_copy2, regressionResultHigh)):
            all_withoutml(regression_high_copy2, regressionResultHigh, ws_highAnalysis)
            #filterbuy = is_buy_filterBuy(regression_data)
            #filter5 = is_buy_filter5(regression_data)
            # if(is_buy_filter2(regression_data)):
            #     if((db['highBuy'].find_one({'scrip':scrip}) is None)):
            #         record = {}
            #         record['scrip'] = scrip
            #         record['PCT_change'] =regression_data['PCT_change']
            #         record['PCT_day_change'] =regression_data['PCT_day_change']
            #         record['ml'] = ''
            #         record['filter'] = ''
            #         record['filter2'] = regression_data['filter2']
            #         record['filter3'] = ''
            #         json_data = json.loads(json.dumps(record, default=json_util.default))
            #         db['highBuy'].insert_one(json_data)
            #     else:
            #         db['highBuy'].update_one({'scrip':scrip}, { "$set": {'filter2':regression_data['filter2']}})
            # if(filterbuy != ""):
            #     if((db['highBuy'].find_one({'scrip':scrip}) is None)):
            #         record = {}
            #         record['scrip'] = scrip
            #         record['PCT_change'] =regression_data['PCT_change']
            #         record['PCT_day_change'] =regression_data['PCT_day_change']
            #         record['ml'] = ''
            #         record['filter'] = ''
            #         record['filter2'] = filterbuy
            #         record['filter3'] = ''
            #         json_data = json.loads(json.dumps(record, default=json_util.default))
            #         db['highBuy'].insert_one(json_data)
            #     else:
            #         temp = db['highBuy'].find_one({'scrip':scrip})
            #         db['highBuy'].update_one({'scrip':scrip}, { "$set": {'filter2':filterbuy + ',' + temp['filter2']}})
            # if(filter5 != ""):
            #     if((db['highBuy'].find_one({'scrip':scrip}) is None)):
            #         record = {}
            #         record['scrip'] = scrip
            #         record['PCT_change'] =regression_data['PCT_change']
            #         record['PCT_day_change'] =regression_data['PCT_day_change']
            #         record['ml'] = ''
            #         record['filter'] = ''
            #         record['filter2'] = filter5
            #         record['filter3'] = ''
            #         json_data = json.loads(json.dumps(record, default=json_util.default))
            #         db['highBuy'].insert_one(json_data)
            #     else:
            #         temp = db['highBuy'].find_one({'scrip':scrip})
            #         db['highBuy'].update_one({'scrip':scrip}, { "$set": {'filter2':filter5 + ',' + temp['filter2']}})
            
        if (is_algo_buy(regression_high_copy2)):
            all_withoutml(regression_data, regressionResultHigh, ws_highBuyReg)
        if ((is_algo_buy(regression_high_copy2) and is_any_reg_algo_gt1_not_other(regression_data) and trend_positive(regression_data))
            or (is_algo_buy(regression_high_copy2, True) and is_algo_buy(regression_low_copy2, True) and is_any_reg_algo_gt1(regression_data) and trend_positive(regression_data))
            ):
            buy_all_common_High_Low(regression_data, regressionResultHigh, True, None)
            all_withoutml(regression_data, regressionResultHigh, ws_highBuyStrong) 
            if((db['highBuy'].find_one({'scrip':scrip}) is None)):
                regression_data['ml'] = 'MLhighBuy'
                db['highBuy'].insert_one(regression_data)
            else:
                dataHighBuy = db['highBuy'].find_one({'scrip':scrip})
                if(dataHighBuy['ml'] == ''):
                    db['highBuy'].update_one({'scrip':scrip}, { "$set": {'ml':'MLhighBuy'}})
        if ((is_algo_sell(regression_high_copy2) and is_any_reg_algo_lt_minus1_not_other(regression_data))
            or (is_algo_sell(regression_high_copy2, True) and is_algo_sell(regression_low_copy2, True) and is_any_reg_algo_lt_minus1(regression_data))
            ):
            sell_all_common_High_Low(regression_data, regressionResultLow, True, None)
            all_withoutml(regression_data, regressionResultLow, ws_highSellStrong)
        
        if (is_algo_buy(regression_high_copy2, True) and is_algo_buy(regression_low_copy2, True) 
            and is_any_reg_algo_gt1(regression_data) and trend_positive(regression_data)
            #and is_any_reg_algo_gt1_not_other(regression_data)
            ):
            buy_all_common_High_Low(regression_data, regressionResultHigh, True, None)
            all_withoutml(regression_data, regressionResultHigh, ws_highBuyStrongBoth)
            if((db['highBuy'].find_one({'scrip':scrip}) is None)):
                regression_data['ml'] = 'MLhighBuyStrongBoth'
                db['highBuy'].insert_one(regression_data)
            else:
                db['highBuy'].update_one({'scrip':scrip}, { "$set": {'ml':'MLhighBuyStrongBoth'}})
                
        regression_data = regression_low_copy2
        if(sell_high_volatility(regression_low_copy2, regressionResultLow)):
            all_withoutml(regression_low_copy2, regressionResultLow, ws_lowAnalysis)
            #filtersell = is_sell_filterSell(regression_data)
            #filter5 = is_sell_filter5(regression_data)
            # if(is_sell_filter2(regression_data)):
            #     if((db['lowSell'].find_one({'scrip':scrip}) is None)):
            #         record = {}
            #         record['scrip'] = scrip
            #         record['PCT_change'] =regression_data['PCT_change']
            #         record['PCT_day_change'] =regression_data['PCT_day_change']
            #         record['ml'] = ''
            #         record['filter'] = ''
            #         record['filter2'] = regression_data['filter2']
            #         record['filter3'] = ''
            #         json_data = json.loads(json.dumps(record, default=json_util.default))
            #         db['lowSell'].insert_one(json_data)
            #     else:
            #         db['lowSell'].update_one({'scrip':scrip}, { "$set": {'filter2':regression_data['filter2']}})
            # if(filtersell != ""):
            #     if((db['lowSell'].find_one({'scrip':scrip}) is None)):
            #         record = {}
            #         record['scrip'] = scrip
            #         record['PCT_change'] =regression_data['PCT_change']
            #         record['PCT_day_change'] =regression_data['PCT_day_change']
            #         record['ml'] = ''
            #         record['filter'] = ''
            #         record['filter2'] = filtersell
            #         record['filter3'] = ''
            #         json_data = json.loads(json.dumps(record, default=json_util.default))
            #         db['lowSell'].insert_one(json_data)
            #     else:
            #         temp = db['lowSell'].find_one({'scrip':scrip})
            #         db['lowSell'].update_one({'scrip':scrip}, { "$set": {'filter2':(filtersell + ',' + temp['filter2'])}})
            # if(filter5 != ""):
            #     if((db['lowSell'].find_one({'scrip':scrip}) is None)):
            #         record = {}
            #         record['scrip'] = scrip
            #         record['PCT_change'] =regression_data['PCT_change']
            #         record['PCT_day_change'] =regression_data['PCT_day_change']
            #         record['ml'] = ''
            #         record['filter'] = ''
            #         record['filter2'] = filter5
            #         record['filter3'] = ''
            #         json_data = json.loads(json.dumps(record, default=json_util.default))
            #         db['lowSell'].insert_one(json_data)
            #     else:
            #         temp = db['lowSell'].find_one({'scrip':scrip})
            #         db['lowSell'].update_one({'scrip':scrip}, { "$set": {'filter2':(filter5 + ',' + temp['filter2'])}})
            
        if (is_algo_sell(regression_low_copy2)):
            all_withoutml(regression_data, regressionResultHigh, ws_lowSellReg)
        if ((is_algo_sell(regression_low_copy2) and is_any_reg_algo_lt_minus1_not_other(regression_data) and trend_negative(regression_data))
            or (is_algo_sell(regression_high_copy2, True) and is_algo_sell(regression_low_copy2, True) and is_any_reg_algo_lt_minus1(regression_data) and trend_negative(regression_data))
            ):
            sell_all_common_High_Low(regression_data, regressionResultLow, True, None)
            all_withoutml(regression_data, regressionResultLow, ws_lowSellStrong)
            if((db['lowSell'].find_one({'scrip':scrip}) is None)):
                regression_data['ml'] = 'MLlowSell'
                db['lowSell'].insert_one(regression_data)
            else:
                dataLowSell = db['lowSell'].find_one({'scrip':scrip})
                if(dataLowSell['ml'] == ''):
                    db['lowSell'].update_one({'scrip':scrip}, { "$set": {'ml':'MLlowSell'}})
        if ((is_algo_buy(regression_low_copy2) and is_any_reg_algo_gt1_not_other(regression_data))
            or (is_algo_buy(regression_high_copy2, True) and is_algo_buy(regression_low_copy2, True) and is_any_reg_algo_gt1(regression_data))
            ):
            buy_all_common_High_Low(regression_data, regressionResultHigh, True, None)
            all_withoutml(regression_data, regressionResultHigh, ws_lowBuyStrong)
        
        if (is_algo_sell(regression_high_copy2, True) and is_algo_sell(regression_low_copy2, True) 
            and is_any_reg_algo_lt_minus1(regression_data) and trend_negative(regression_data)
            #and is_any_reg_algo_lt_minus1_not_other(regression_data)
            ):
            sell_all_common_High_Low(regression_data, regressionResultLow, True, None)
            all_withoutml(regression_data, regressionResultLow, ws_lowSellStrongBoth)
            if((db['lowSell'].find_one({'scrip':scrip}) is None)):
                regression_data['ml'] = 'MLlowSellStrongBoth'
                db['lowSell'].insert_one(regression_data)
            else:
                db['lowSell'].update_one({'scrip':scrip}, { "$set": {'ml':'MLlowSellStrongBoth'}})
        
        regression_high_copy['filter2']=""
        if(is_filter_all_accuracy(regression_high_copy, regression_high, regression_low, regressionResultHigh, 'High', None)):
            all_withoutml(regression_high_copy, regressionResultHigh, ws_highBuyStrongFilterAcc)
            if(is_filter_hist_accuracy(regression_high_copy, regression_high, regression_low, regressionResultHigh, 'High', None)):
                if((db['highBuy'].find_one({'scrip':scrip}) is None)):
                    db['highBuy'].insert_one(regression_high_copy)
                else:
                    db['highBuy'].update_one({'scrip':scrip}, { "$set": {'filter2':regression_high_copy['filter2']}})
                            
            
        regression_low_copy['filter2']=""
        if(is_filter_all_accuracy(regression_low_copy, regression_high, regression_low, regressionResultLow, 'Low', None)):
            all_withoutml(regression_low_copy, regressionResultLow, ws_lowSellStrongFilterAcc)
            if(is_filter_hist_accuracy(regression_low_copy, regression_high, regression_low, regressionResultLow, 'Low', None)):
                if((db['lowSell'].find_one({'scrip':scrip}) is None)):
                    db['lowSell'].insert_one(regression_low_copy)
                else:
                    db['lowSell'].update_one({'scrip':scrip}, { "$set": {'filter2':regression_low_copy['filter2']}})
                    
        
        if(is_filter_all_accuracy(regression_high_copy1, regression_high, regression_low, regressionResultHigh, "None", None)
            and is_filter_all_accuracy(regression_low_copy1, regression_high, regression_low, regressionResultLow, "None", None)
            ):
            if(('Buy-AnyGT2'in regression_high_copy1['filter2'] or 'Buy-AnyGT2' in regression_low_copy1['filter2']
                and 'Buy-AnyGT2-1.0' not in regression_high_copy1['filter2']
                and 'Buy-AnyGT2-1.0' not in regression_low_copy1['filter2']
                and 'Buy-AnyGT2-2.0' not in regression_high_copy1['filter2']
                and 'Buy-AnyGT2-2.0' not in regression_low_copy1['filter2']
                and 'Sell-Any' not in regression_high_copy1['filter2']
                and 'Sell-Any' not in regression_low_copy1['filter2']
                )
                #and 'Buy-Any' in regression_high_copy1['filter2'] and 'Buy-Any' in regression_low_copy1['filter2']
                #and 'Sell-Any' not in regression_high_copy1['filter2'] and 'Sell-Any' not in regression_low_copy1['filter2']
                #and (abs(regression_high_copy1['PCT_change']) > 1.5 or abs(regression_high_copy1['PCT_change_pre1']) > 1.5)
                ):
                all_withoutml(regression_high_copy1, regressionResultHigh, ws_allFilterAcc)
                all_withoutml(regression_low_copy1, regressionResultLow, ws_allFilterAcc)  
            if(('Sell-AnyGT2'in regression_high_copy1['filter2'] or 'Sell-AnyGT2' in regression_low_copy1['filter2']
                and 'Sell-AnyGT2-1.0' not in regression_high_copy1['filter2']
                and 'Sell-AnyGT2-1.0' not in regression_low_copy1['filter2']
                and 'Sell-AnyGT2-2.0' not in regression_high_copy1['filter2']
                and 'Sell-AnyGT2-2.0' not in regression_low_copy1['filter2']
                and 'Buy-Any' not in regression_high_copy1['filter2']
                and 'Buy-Any' not in regression_low_copy1['filter2']
                )
                #and 'Sell-Any' in regression_high_copy1['filter2'] and 'Sell-Any' in regression_low_copy1['filter2']
                #and 'Buy-Any' not in regression_high_copy1['filter2'] and 'Buy-Any' not in regression_low_copy1['filter2']
                #and (abs(regression_high_copy1['PCT_change']) > 1.5 or abs(regression_high_copy1['PCT_change_pre1']) > 1.5)
                ):
                all_withoutml(regression_high_copy1, regressionResultHigh, ws_allFilterAcc)
                all_withoutml(regression_low_copy1, regressionResultLow, ws_allFilterAcc)
            '''
            if (('Buy-SUPER' in regression_high_copy1['filter2'] or 'Buy-SUPER' in regression_low_copy1['filter2'])
                and (abs(regression_high_copy1['PCT_change']) > 2 or abs(regression_high_copy1['PCT_change_pre1']) > 2)
                ):
                all_withoutml(regression_high_copy1, regressionResultHigh, ws_allFilterAcc)
                all_withoutml(regression_low_copy1, regressionResultLow, ws_allFilterAcc)
            if (('Sell-SUPER' in regression_high_copy1['filter2'] or 'Sell-SUPER' in regression_low_copy1['filter2'])
                and (abs(regression_high_copy1['PCT_change']) > 2 or abs(regression_high_copy1['PCT_change_pre1']) > 2)
                ):
                all_withoutml(regression_high_copy1, regressionResultHigh, ws_allFilterAcc)
                all_withoutml(regression_low_copy1, regressionResultLow, ws_allFilterAcc)
            '''

        intradaytech = ""
        intradaytech = intraday_tech_data(regression_data)
        shorttermtechbuy = shortterm_tech_data_buy(regression_high_copy2, regression_low_copy2)
        shorttermtechsell = shortterm_tech_data_sell(regression_high_copy2, regression_low_copy2)
        index = get_index(regression_data['scrip'])
        fundamenta_data = get_fundamental(regression_data['scrip'])
        marketcap = ''
        netprofit = ''
        peratio = ''
        publicholding = ''
        if (fundamenta_data != ''):
            marketcap = fundamenta_data['marketcap']
            netprofit = fundamenta_data['netprofit']
            peratio = fundamenta_data['peratio']
            publicholding = fundamenta_data['publicholding']

        db['highBuy'].update_one({'scrip': scrip}, {"$set": {'intradaytech': intradaytech}})
        db['lowSell'].update_one({'scrip': scrip}, {"$set": {'intradaytech': intradaytech}})
        db['highBuy'].update_one({'scrip': scrip}, {"$set": {'shorttermtech': shorttermtechbuy}})
        db['lowSell'].update_one({'scrip': scrip}, {"$set": {'shorttermtech': shorttermtechsell}})
        db['highBuy'].update_one({'scrip': scrip}, {"$set": {'index': index}})
        db['lowSell'].update_one({'scrip': scrip}, {"$set": {'index': index}})
        db['highBuy'].update_one({'scrip': scrip}, {"$set": {'marketcap': marketcap}})
        db['lowSell'].update_one({'scrip': scrip}, {"$set": {'marketcap': marketcap}})
        db['highBuy'].update_one({'scrip': scrip}, {"$set": {'netprofit': netprofit}})
        db['lowSell'].update_one({'scrip': scrip}, {"$set": {'netprofit': netprofit}})
        db['highBuy'].update_one({'scrip': scrip}, {"$set": {'peratio': peratio}})
        db['lowSell'].update_one({'scrip': scrip}, {"$set": {'peratio': peratio}})
        db['highBuy'].update_one({'scrip': scrip}, {"$set": {'publicholding': publicholding}})
        db['lowSell'].update_one({'scrip': scrip}, {"$set": {'publicholding': publicholding}})

        if (dbchartlink['highBuy'].find_one({'scrip': regression_data['scrip']}) is None):
            dbchartlink['highBuy'].insert_one(db['highBuy'].find_one({'scrip': regression_data['scrip']}))
        if (dbchartlink['lowSell'].find_one({'scrip': regression_data['scrip']}) is None):
            dbchartlink['lowSell'].insert_one(db['lowSell'].find_one({'scrip': regression_data['scrip']}))

        insert_year5LowBreakoutY2H(regression_data)
        insert_year5LowBreakoutYH(regression_data)
        insert_year2LowReversal(regression_data)
        insert_year5LowBreakoutMonthHigh(regression_data)
        insert_year5HighNearBreakout(regression_data)
        #insert_year2HighNearBreakout(regression_data)
        
def result_data_cla(scrip):
    regression_high = db.regressionhigh.find_one({'scrip':scrip})
    regression_low = db.regressionlow.find_one({'scrip':scrip})
    if(regression_high is None or regression_low is None):
        return
    
    regression_data = regression_high
    if(regression_data is not None):
        regressionResult = get_regressionResult(regression_data, scrip, db, 
                                            regression_low['mlpValue_reg'], 
                                            regression_low['kNeighboursValue_reg'],
                                            regression_low['mlpValue_cla'], 
                                            regression_low['kNeighboursValue_cla'],
                                            )
        if buy_all_rule_classifier(regression_data, regressionResult, buyIndiaAvg, None):
            buy_other_indicator(regression_data, regressionResult, True, None)
            buy_filter_all_accuracy(regression_data, regressionResult)
            buy_all_rule_classifier(regression_data, regressionResult, buyIndiaAvg, ws_highBuyStrongFilterAcc)
                
    regression_data = regression_low
    if(regression_data is not None):
        regressionResult = get_regressionResult(regression_data, scrip, db, 
                                            regression_high['mlpValue_reg'], 
                                            regression_high['kNeighboursValue_reg'],
                                            regression_high['mlpValue_cla'], 
                                            regression_high['kNeighboursValue_cla'],
                                            False
                                            )
        if sell_all_rule_classifier(regression_data, regressionResult, None, None):
            sell_other_indicator(regression_data, regressionResult, True, None)
            sell_filter_all_accuracy(regression_data, regressionResult)
            sell_all_rule_classifier(regression_data, regressionResult, None, ws_lowSellStrongFilterAcc)                              
                                             
def calculateParallel(threads=2, futures='Yes', processing_date=" "):
    pool = ThreadPool(threads)
    scrips = []
    
    if(processing_date == " "):
        processing_date = (datetime.date.today() - datetime.timedelta(days=0)).strftime('%Y-%m-%d')
    for data in db.scrip.find({'futures':futures}):
        #print('Scrip ', data)
        regdata = db.regressionlow.find_one({'scrip':data['scrip']})
        if(regdata is None):
            print('Missing or very less Data for ', data['scrip'])
        if(regdata is not None and regdata['date'] != processing_date):
            print('End Date ', regdata['date'], 'not recent for', data['scrip'])
        else: 
            scrips.append(data['scrip'])
    if (futures =='No' or futures =='NO' or futures =='no'):
        for data in db.scrip.find({'futures':'Yes'}):
            #print('Scrip ', data)
            regdata = db.regressionlow.find_one({'scrip':data['scrip']})
            if(regdata is None):
                print('Missing or very less Data for ', data['scrip'])
            if(regdata is not None and regdata['date'] != processing_date):
                print('End Date ', regdata['date'], 'not recent for', data['scrip'])
            else: 
                scrips.append(data['scrip'])
    scrips.sort()
    #pool.map(result_data, scrips)
    #pool.map(result_data_cla, scrips)
    pool.map(result_data_reg, scrips)
    pool.map(result_data_summary, scrips)

                      
if __name__ == "__main__":
    if not os.path.exists(directory):
        os.makedirs(directory)
    calculateParallel(1, sys.argv[1], sys.argv[2])
    connection.close()
    saveReports()