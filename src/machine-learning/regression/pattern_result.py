import os, logging, sys, json, csv
sys.path.insert(0, '../')

from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Color, PatternFill, Font, Border
from pymongo import MongoClient
from multiprocessing.dummy import Pool as ThreadPool

import quandl, math, time
import pandas as pd
import numpy as np
from talib.abstract import *

import datetime
import time
import gc

from util.util import getScore, all_day_pct_change_negative, all_day_pct_change_positive, no_doji_or_spinning_buy_india, no_doji_or_spinning_sell_india, scrip_patterns_to_dict

connection = MongoClient('localhost', 27017)
db = connection.Nsedata

buyPatternsDict=scrip_patterns_to_dict('../../data-import/nselist/all-buy-filter-by-PCT-Change.csv')
sellPatternsDict=scrip_patterns_to_dict('../../data-import/nselist/all-buy-filter-by-PCT-Change.csv')

directory = '../../output/final'
logname = '../../output/final' + '/pattern-result' + time.strftime("%d%m%y-%H%M%S")

newsDict = {}
wb = Workbook()
ws_buyPattern2 = wb.create_sheet("buyPattern2")
ws_buyPattern2.append(["BuyIndicators", "SellIndicators","Symbol", "VOL_change", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_Day_Change", "PCT_Change","Score", "MLP", "KNeighbors", "trend", "yHighChange","yLowChange", "ResultDate", "ResultDeclared", "ResultSentiment", "ResultComment", "Avg", "Count"])
ws_sellPattern2 = wb.create_sheet("sellPattern2")
ws_sellPattern2.append(["BuyIndicators", "SellIndicators","Symbol", "VOL_change", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_Day_Change", "PCT_Change","Score", "MLP", "KNeighbors", "trend", "yHighChange","yLowChange", "ResultDate", "ResultDeclared", "ResultSentiment", "ResultComment", "Avg", "Count"])
ws_sellYearHigh = wb.create_sheet("sellYearHigh")
ws_sellYearHigh.append(["BuyIndicators", "SellIndicators","Symbol", "VOL_change", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_Day_Change", "PCT_Change","Score", "MLP", "KNeighbors", "trend", "yHighChange","yLowChange", "ResultDate", "ResultDeclared", "ResultSentiment", "ResultComment", "Avg", "Count"])
ws_buyYearLow = wb.create_sheet("buyYearLow")
ws_buyYearLow.append(["BuyIndicators", "SellIndicators","Symbol", "VOL_change", "PCT", "PCT2", "PCT3", "PCT4", "PCT5", "PCT7", "PCT10", "PCT_Day_Change", "PCT_Change","Score", "MLP", "KNeighbors", "trend", "yHighChange","yLowChange", "ResultDate", "ResultDeclared", "ResultSentiment", "ResultComment", "Avg", "Count"])

def saveReports(run_type=None):
    # Add a default style with striped rows and banded columns
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
               showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    
    
    count = 0
    for row in ws_sellYearHigh.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:Y" + str(count))
    tab.tableStyleInfo = style
    ws_sellYearHigh.add_table(tab)
    
    count = 0
    for row in ws_buyYearLow.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:Y" + str(count))
    tab.tableStyleInfo = style
    ws_buyYearLow.add_table(tab)
    
    count = 0
    for row in ws_buyPattern2.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:Y" + str(count))
    tab.tableStyleInfo = style
    ws_buyPattern2.add_table(tab)
    
    count = 0
    for row in ws_sellPattern2.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:Y" + str(count))
    tab.tableStyleInfo = style
    ws_sellPattern2.add_table(tab)
      
    if(run_type == 'broker'):
        wb.save(logname + "broker_buy.xlsx")
    elif(run_type == 'result'):
        wb.save(logname + "result.xlsx")
    elif(run_type == 'result_declared'):
        wb.save(logname + "result_declared.xlsx")       
    else:
        wb.save(logname + ".xlsx")
   
def result_data(scrip):
    resultDeclared = ""
    resultDate = ""
    resultSentiment = ""
    resultComment = ""
    result_data = db.scrip_result.find_one({'scrip':scrip.replace('&','').replace('-','_')})
    if(result_data is not None):
        resultDate = result_data['result_date'].strip()
        resultSentiment = result_data['result_sentiment']
        resultComment = result_data['comment']
        start_date = (datetime.datetime.now() - datetime.timedelta(hours=0))
        start_date = datetime.datetime(start_date.year, start_date.month, start_date.day, start_date.hour)
        result_time = datetime.datetime.strptime(resultDate, "%Y-%m-%d")
        if result_time < start_date: 
            resultDeclared = resultDate 
            resultDate = ""
       
    regression_data = db.regressionhigh.find_one({'scrip':scrip.replace('&','').replace('-','_')})
    if(regression_data is not None):
        regressionResult = [ ]
        regressionResult.append(regression_data['buyIndia'])
        regressionResult.append(regression_data['sellIndia'])
        regressionResult.append(regression_data['scrip'])
        regressionResult.append(regression_data['forecast_day_VOL_change'])
        regressionResult.append(regression_data['forecast_day_PCT_change'])
        regressionResult.append(regression_data['forecast_day_PCT2_change'])
        regressionResult.append(regression_data['forecast_day_PCT3_change'])
        regressionResult.append(regression_data['forecast_day_PCT4_change'])
        regressionResult.append(regression_data['forecast_day_PCT5_change'])
        regressionResult.append(regression_data['forecast_day_PCT7_change'])
        regressionResult.append(regression_data['forecast_day_PCT10_change'])
        regressionResult.append(regression_data['PCT_day_change'])
        regressionResult.append(regression_data['PCT_change'])
        regressionResult.append(regression_data['score'])
        regressionResult.append(regression_data['mlpValue'])
        regressionResult.append(regression_data['kNeighboursValue'])
        regressionResult.append(regression_data['trend'])
        regressionResult.append(regression_data['yearHighChange'])
        regressionResult.append(regression_data['yearLowChange'])
        regressionResult.append(resultDate)
        regressionResult.append(resultDeclared)
        regressionResult.append(resultSentiment)
        regressionResult.append(resultComment)
        
        if regression_data['buyIndia'] != '' and regression_data['buyIndia'] in buyPatternsDict:
            if (abs(float(buyPatternsDict[regression_data['buyIndia']]['avg'])) >= .1):
                regressionResult.append(buyPatternsDict[regression_data['buyIndia']]['avg'])
                regressionResult.append(buyPatternsDict[regression_data['buyIndia']]['count'])
                if('P@[' not in str(regression_data['sellIndia'])):
                    if(float(buyPatternsDict[regression_data['buyIndia']]['avg']) > 0.8):
                       ws_buyPattern2.append(regressionResult) 
                    
    regression_data = db.regressionlow.find_one({'scrip':scrip.replace('&','').replace('-','_')})
    if(regression_data is not None):
        regressionResult = [ ]
        regressionResult.append(regression_data['buyIndia'])
        regressionResult.append(regression_data['sellIndia'])
        regressionResult.append(regression_data['scrip'])
        regressionResult.append(regression_data['forecast_day_VOL_change'])
        regressionResult.append(regression_data['forecast_day_PCT_change'])
        regressionResult.append(regression_data['forecast_day_PCT2_change'])
        regressionResult.append(regression_data['forecast_day_PCT3_change'])
        regressionResult.append(regression_data['forecast_day_PCT4_change'])
        regressionResult.append(regression_data['forecast_day_PCT5_change'])
        regressionResult.append(regression_data['forecast_day_PCT7_change'])
        regressionResult.append(regression_data['forecast_day_PCT10_change'])
        regressionResult.append(regression_data['PCT_day_change'])
        regressionResult.append(regression_data['PCT_change'])
        regressionResult.append(regression_data['score'])
        regressionResult.append(regression_data['mlpValue'])
        regressionResult.append(regression_data['kNeighboursValue'])
        regressionResult.append(regression_data['trend'])
        regressionResult.append(regression_data['yearHighChange'])
        regressionResult.append(regression_data['yearLowChange'])
        regressionResult.append(resultDate)
        regressionResult.append(resultDeclared)
        regressionResult.append(resultSentiment)
        regressionResult.append(resultComment)
        
        if regression_data['sellIndia'] != '' and regression_data['sellIndia'] in sellPatternsDict:
            if (abs(float(sellPatternsDict[regression_data['sellIndia']]['avg'])) >= .1):
                regressionResult.append(sellPatternsDict[regression_data['sellIndia']]['avg'])
                regressionResult.append(sellPatternsDict[regression_data['sellIndia']]['count'])
                if('P@[' not in str(regression_data['buyIndia'])):
                    if(float(sellPatternsDict[regression_data['sellIndia']]['avg']) < -0.8):
                        ws_sellPattern2.append(regressionResult)    
                                           
def calculateParallel(threads=2, futures=None):
    pool = ThreadPool(threads)
    scrips = []
    for data in db.scrip.find({'futures':futures}):
        scrips.append(data['scrip'].replace('&','').replace('-','_'))
    scrips.sort()
    pool.map(result_data, scrips)       
                     
if __name__ == "__main__":
    if not os.path.exists(directory):
        os.makedirs(directory)
    calculateParallel(1, sys.argv[1])
    connection.close()
    saveReports(sys.argv[1])