import os, logging, sys
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Color, PatternFill, Font, Border
from pymongo import MongoClient
from multiprocessing.dummy import Pool as ThreadPool

import quandl, math, time
import pandas as pd
import numpy as np
from sklearn import preprocessing, cross_validation, svm
from sklearn.linear_model import LinearRegression
from talib.abstract import *
from pip.req.req_file import preprocess
from Algorithms.regression_helpers import load_dataset, addFeatures, \
    mergeDataframes, count_missing, applyTimeLag, performRegression   
    
from technical import ta_lib_data    

connection = MongoClient('localhost', 27017)
db = connection.Nsedata

directory = '../../output' + '/' + time.strftime("%d%m%y-%H%M%S")
logname = '../../output' + '/mllog' + time.strftime("%d%m%y-%H%M%S")
logging.basicConfig(filename=logname, filemode='a', stream=sys.stdout, level=logging.INFO)
log = logging.getLogger(__name__)

wb = Workbook()
ws = wb.active
ws_filter = wb.create_sheet("Filter")
ws_gtltzero = wb.create_sheet("FilterAllgtlt0")
ws_RandomForest = wb.create_sheet("RandomForest")
ws_SVR = wb.create_sheet("SVR")
ws_Bagging = wb.create_sheet("Bagging")
ws_AdaBoost = wb.create_sheet("AdaBoost")
ws_KNeighbors = wb.create_sheet("KNeighbors")
ws_GradientBoosting = wb.create_sheet("GradientBoosting")
ws.append(["Symbol", "train set", "test set", "RandomForest", "accuracy", "SVR", "accuracy", "Bagging", "accuracy", "AdaBoost", "accuracy", "KNeighbors", "accuracy", "GradientBoosting", "accuracy", "BuyIndicators", "SellIndicators", "futures"])
ws_filter.append(["Symbol", "train set", "test set", "RandomForest", "accuracy", "SVR", "accuracy", "Bagging", "accuracy", "AdaBoost", "accuracy", "KNeighbors", "accuracy", "GradientBoosting", "accuracy", "BuyIndicators", "SellIndicators", "futures"])
ws_gtltzero.append(["Symbol", "train set", "test set", "RandomForest", "accuracy", "SVR", "accuracy", "Bagging", "accuracy", "AdaBoost", "accuracy", "KNeighbors", "accuracy", "GradientBoosting", "accuracy", "BuyIndicators", "SellIndicators", "futures"])
ws_RandomForest.append(["Symbol", "train set", "test set", "RandomForest", "accuracy", "SVR", "accuracy", "Bagging", "accuracy", "AdaBoost", "accuracy", "KNeighbors", "accuracy", "GradientBoosting", "accuracy", "BuyIndicators", "SellIndicators", "futures"])
ws_SVR.append(["Symbol", "train set", "test set", "RandomForest", "accuracy", "SVR", "accuracy", "Bagging", "accuracy", "AdaBoost", "accuracy", "KNeighbors", "accuracy", "GradientBoosting", "accuracy", "BuyIndicators", "SellIndicators", "futures"])
ws_Bagging.append(["Symbol", "train set", "test set", "RandomForest", "accuracy", "SVR", "accuracy", "Bagging", "accuracy", "AdaBoost", "accuracy", "KNeighbors", "accuracy", "GradientBoosting", "accuracy", "BuyIndicators", "SellIndicators", "futures"])
ws_AdaBoost.append(["Symbol", "train set", "test set", "RandomForest", "accuracy", "SVR", "accuracy", "Bagging", "accuracy", "AdaBoost", "accuracy", "KNeighbors", "accuracy", "GradientBoosting", "accuracy", "BuyIndicators", "SellIndicators", "futures"])
ws_KNeighbors.append(["Symbol", "train set", "test set", "RandomForest", "accuracy", "SVR", "accuracy", "Bagging", "accuracy", "AdaBoost", "accuracy", "KNeighbors", "accuracy", "GradientBoosting", "accuracy", "BuyIndicators", "SellIndicators", "futures"])
ws_GradientBoosting.append(["Symbol", "train set", "test set", "RandomForest", "accuracy", "SVR", "accuracy", "Bagging", "accuracy", "AdaBoost", "accuracy", "KNeighbors", "accuracy", "GradientBoosting", "accuracy", "BuyIndicators", "SellIndicators", "futures"])

def saveReports():
    # Add a default style with striped rows and banded columns
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
               showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    
    count = 0
    for row in ws.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:R" + str(count))
    tab.tableStyleInfo = style
    ws.add_table(tab)
    
    count = 0
    for row in ws_filter.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:R" + str(count))
    tab.tableStyleInfo = style
    ws_filter.add_table(tab)
    
    count = 0
    for row in ws_gtltzero.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:R" + str(count))
    tab.tableStyleInfo = style
    ws_gtltzero.add_table(tab)
    
    count = 0
    for row in ws_RandomForest.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:R" + str(count))
    tab.tableStyleInfo = style
    ws_RandomForest.add_table(tab)
    
    count = 0
    for row in ws_SVR.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:R" + str(count))
    tab.tableStyleInfo = style
    ws_SVR.add_table(tab)
    
    count = 0
    for row in ws_Bagging.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:R" + str(count))
    tab.tableStyleInfo = style
    ws_Bagging.add_table(tab)
    
    count = 0
    for row in ws_AdaBoost.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:R" + str(count))
    tab.tableStyleInfo = style
    ws_AdaBoost.add_table(tab)
    
    count = 0
    for row in ws_KNeighbors.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:R" + str(count))
    tab.tableStyleInfo = style
    ws_KNeighbors.add_table(tab)
    
    count = 0
    for row in ws_GradientBoosting.iter_rows(row_offset=1):
        count += 1
    tab = Table(displayName="Table1", ref="A1:R" + str(count))
    tab.tableStyleInfo = style
    ws_GradientBoosting.add_table(tab)
    
    wb.save(logname + ".xlsx")


def historical_data(data):
    ardate = np.array([x.encode('UTF8') for x in (np.array(data['data'])[:,0][::-1]).tolist()])
    aropen = np.array([float(x.encode('UTF8')) for x in (np.array(data['data'])[:,1][::-1]).tolist()])
    arhigh = np.array([float(x.encode('UTF8')) for x in (np.array(data['data'])[:,2][::-1]).tolist()])
    arlow  = np.array([float(x.encode('UTF8')) for x in (np.array(data['data'])[:,3][::-1]).tolist()])
    arlast = np.array([float(x.encode('UTF8')) for x in (np.array(data['data'])[:,4][::-1]).tolist()])
    arclose= np.array([float(x.encode('UTF8')) for x in (np.array(data['data'])[:,5][::-1]).tolist()])
    arquantity = np.array([float(x.encode('UTF8')) for x in (np.array(data['data'])[:,6][::-1]).tolist()])
    arturnover = np.array([float(x.encode('UTF8')) for x in (np.array(data['data'])[:,7][::-1]).tolist()])
    return ardate, aropen, arhigh, arlow, arlast, arclose, arquantity, arturnover

def regression_ta_data(scrip):
    data = db.history.find_one({'dataset_code':scrip.encode('UTF8').replace('&','').replace('-','_')})
    if(data is None or (np.array(data['data'])).size < 200):
        print('Missing or very less Data for ', scrip.encode('UTF8'))
        return
        
    hsdate, hsopen, hshigh, hslow, hslast, hsclose, hsquantity, hsturnover = historical_data(data)   
    df = pd.DataFrame({
        'date': hsdate,
        'open': hsopen,
        'high': hshigh,
        'low': hslow,
        'close': hsclose,
        'volume': hsquantity,
        'turnover':hsturnover
    })
    df = df[['date','open','high','low','close','volume','turnover']]
    #dfp = df[['open','high','low','close']]
        
    if (df is not None):
        df=df.rename(columns = {'total trade quantity':'volume'})
        df=df.rename(columns = {'turnover (lacs)': 'turnover'})
        df['PCT_day_change'] = (((df['close'] - df['open'])/df['open'])*100)
        df['HL_change'] = (((df['high'] - df['low'])/df['low'])*100).astype(int)
        df['volume_pre'] = df['volume'].shift(+1)
        df['close_pre'] = df['close'].shift(+1)
        #df.fillna(-99999, inplace=True)
        df.dropna(inplace=True)
        df['VOL_change'] = (((df['volume'] - df['volume_pre'])/df['volume_pre'])*100)
        df['PCT_change'] = (((df['close'] - df['close_pre'])/df['close_pre'])*100)
        
        dfp = df[['VOL_change']]
        maxdelta = 10
        delta = range(1, maxdelta)
        columns = df.columns
        close = columns[4]
        for dele in delta:
            addFeatures(df, dfp, close, dele)
            
        dfp['ADX'] = ADX(df).apply(lambda x: 1 if x > 20 else 0) #Average Directional Movement Index http://www.investopedia.com/terms/a/adx.asp
        dfp['ADXR'] = ADXR(df).apply(lambda x: 1 if x > 20 else 0) #Average Directional Movement Index Rating https://www.scottrade.com/knowledge-center/investment-education/research-analysis/technical-analysis/the-indicators/average-directional-movement-index-rating-adxr.html
        dfp['APO'] = APO(df).apply(lambda x: 1 if x > 0 else 0) #Absolute Price Oscillator https://www.fidelity.com/learning-center/trading-investing/technical-analysis/technical-indicator-guide/apo
#         aroon = AROON(df) #Aroon http://www.investopedia.com/terms/a/aroon.asp
#         dfp['AROONUP'], dfp['AROONDOWN'] = aroon['aroonup'], aroon['aroondown']
        dfp['AROONOSC'] = AROONOSC(df).apply(lambda x: 1 if x > 0 else 0)
        dfp['BOP'] = BOP(df).apply(lambda x: 1 if x > 0 else 0) #Balance Of Power https://www.marketvolume.com/technicalanalysis/balanceofpower.asp
#        dfp['CCI'] = CCI(df) #Commodity Channel Index http://www.investopedia.com/articles/trading/05/041805.asp
#        dfp['CMO'] = CMO(df) #Chande Momentum Oscillator https://www.fidelity.com/learning-center/trading-investing/technical-analysis/technical-indicator-guide/cmo
#        dfp['DX'] = DX(df) #Directional Movement Index http://www.investopedia.com/terms/d/dmi.asp
#         macd = MACD(df)
#         dfp['MACD'], dfp['MACDSIGNAL'], dfp['MACDHIST'] = macd['macd'], macd['macdsignal'], macd['macdhist']
#         #dfp['MACDEXT'] = MACDEXT(df)
#         #dfp['MACDFIX'] = MACDFIX(df)
#         dfp['MFI'] = MFI(df)
#         dfp['MINUS_DI'] = MINUS_DI(df)
#         dfp['MINUS_DM'] = MINUS_DM(df)
#         dfp['MOM'] = MOM(df)
#         dfp['PLUS_DI'] = PLUS_DI(df)
#         dfp['PLUS_DM'] = PLUS_DM(df)
#         dfp['PPO'] = PPO(df)
#         dfp['ROC'] = ROC(df)
#         dfp['ROCP'] = ROCP(df)
#         dfp['ROCR'] = ROCR(df)
#         dfp['ROCR100'] = ROCR100(df)
#        dfp['RSI'] = RSI(df)
        #dfp['STOCH'] = STOCH(df)
        #dfp['STOCHF'] = STOCHF(df)
        #dfp['STOCHRSI'] = STOCHRSI(df)
#         dfp['TRIX'] = TRIX(df)
#         dfp['ULTOSC'] = ULTOSC(df)
#         dfp['WILLR'] = WILLR(df)
#         
#         bbands = BBANDS(df)
#         dfp['BBANDSUPPER'], dfp['BBANDSMIDDLE'], dfp['BBANDSLOWER'] = bbands['upperband'], bbands['middleband'], bbands['lowerband']
#         
# #       dfp['DEMA'] = DEMA(df)
#         dfp['EMA9'] = EMA(df,9)
#         dfp['EMA21'] = EMA(df,21)
#         dfp['EMA25'] = EMA(df,25)
#         dfp['EMA50'] = EMA(df,50)
#         dfp['EMA100'] = EMA(df,100)
#         dfp['EMA200'] = EMA(df,200)
#         dfp['HT_TRENDLINE'] = HT_TRENDLINE(df)
#         dfp['KAMA'] = KAMA(df)
#         dfp['MA'] = MA(df)
#         #dfp['MAMA'] = MAMA(df)
#         #dfp['MAVP'] = MAVP(df)
#         dfp['MIDPOINT'] = MIDPOINT(df)
#         dfp['MIDPRICE'] = MIDPRICE(df)
#         dfp['SAR'] = SAR(df)
#         dfp['SAREXT'] = SAREXT(df)
#         dfp['SMA'] = SMA(df)
#         dfp['SMA9'] = SMA(df, 9)
#         dfp['T3'] = T3(df)
#         dfp['TEMA'] = TEMA(df)
#         dfp['TRIMA'] = TRIMA(df)
#         dfp['WMA'] = WMA(df)


        dfp['CDL2CROWS'] = CDL2CROWS(df)
        dfp['CDL3BLACKCROWS'] = CDL3BLACKCROWS(df)
        dfp['CDL3INSIDE'] = CDL3INSIDE(df)
        dfp['CDL3LINESTRIKE'] = CDL3LINESTRIKE(df)
        dfp['CDL3OUTSIDE'] = CDL3OUTSIDE(df)
        dfp['CDL3STARSINSOUTH'] = CDL3STARSINSOUTH(df)
        dfp['CDL3WHITESOLDIERS'] = CDL3WHITESOLDIERS(df)
        dfp['CDLABANDONEDBABY'] = CDLABANDONEDBABY(df)
        dfp['CDLADVANCEBLOCK'] = CDLADVANCEBLOCK(df) #Bearish reversal. prior trend Upward. Look for three white candles in an upward price trend. On each candle, price opens within the body of the previous candle. The height of the shadows grow taller on the last two candles.
        dfp['CDLBELTHOLD'] = CDLBELTHOLD(df) # Bearish reversal. prior trend upward. Price opens at the high for the day and closes near the low, forming a tall black candle, often with a small lower shadow.
        dfp['CDLBREAKAWAY'] = CDLBREAKAWAY(df) # Bearish reversal. prior trend upward. Look for 5 candle lines in an upward price trend with the first candle being a tall white one. The second day should be a white candle with a gap between the two bodies, but the shadows can overlap. Day three should have a higher close and the candle can be any color. Day 4 shows a white candle with a higher close. The last day is a tall black candle with a close within the gap between the bodies of the first two candles.
        dfp['CDLCLOSINGMARUBOZU'] = CDLCLOSINGMARUBOZU(df)
        dfp['CDLCONCEALBABYSWALL'] = CDLCONCEALBABYSWALL(df)
        dfp['CDLCOUNTERATTACK'] = CDLCOUNTERATTACK(df)
        dfp['CDLDARKCLOUDCOVER'] = CDLDARKCLOUDCOVER(df)
        dfp['CDLDOJI'] = CDLDOJI(df)
        dfp['CDLDOJISTAR'] = CDLDOJISTAR(df)
        dfp['CDLDRAGONFLYDOJI'] = CDLDRAGONFLYDOJI(df)
        dfp['CDLENGULFING'] = CDLENGULFING(df)
        dfp['CDLEVENINGDOJISTAR'] = CDLEVENINGDOJISTAR(df)
        dfp['CDLEVENINGSTAR'] = CDLEVENINGSTAR(df)
        dfp['CDLGAPSIDESIDEWHITE'] = CDLGAPSIDESIDEWHITE(df)
        dfp['CDLGRAVESTONEDOJI'] = CDLGRAVESTONEDOJI(df)
        dfp['CDLHAMMER'] = CDLHAMMER(df)
        dfp['CDLHANGINGMAN'] = CDLHANGINGMAN(df)
        dfp['CDLHARAMI'] = CDLHARAMI(df)
        dfp['CDLHARAMICROSS'] = CDLHARAMICROSS(df)
        dfp['CDLHIGHWAVE'] = CDLHIGHWAVE(df)
        dfp['CDLHIKKAKE'] = CDLHIKKAKE(df)
        dfp['CDLHIKKAKEMOD'] = CDLHIKKAKEMOD(df)
        dfp['CDLHOMINGPIGEON'] = CDLHOMINGPIGEON(df)
        dfp['CDLIDENTICAL3CROWS'] = CDLIDENTICAL3CROWS(df)
        dfp['CDLINNECK'] = CDLINNECK(df)
        dfp['CDLINVERTEDHAMMER'] = CDLINVERTEDHAMMER(df)
        dfp['CDLKICKING'] = CDLKICKING(df)
        dfp['CDLKICKINGBYLENGTH'] = CDLKICKINGBYLENGTH(df)
        dfp['CDLLADDERBOTTOM'] = CDLLADDERBOTTOM(df)
        dfp['CDLLONGLEGGEDDOJI'] = CDLLONGLEGGEDDOJI(df)
        dfp['CDLLONGLINE'] = CDLLONGLINE(df)
        dfp['CDLMARUBOZU'] = CDLMARUBOZU(df)
        dfp['CDLMATCHINGLOW'] = CDLMATCHINGLOW(df)
        dfp['CDLMATHOLD'] = CDLMATHOLD(df)
        dfp['CDLMORNINGDOJISTAR'] = CDLMORNINGDOJISTAR(df)
        dfp['CDLMORNINGSTAR'] = CDLMORNINGSTAR(df)
        dfp['CDLONNECK'] = CDLONNECK(df)
        dfp['CDLPIERCING'] = CDLPIERCING(df)
        dfp['CDLRICKSHAWMAN'] = CDLRICKSHAWMAN(df)
        dfp['CDLRISEFALL3METHODS'] = CDLRISEFALL3METHODS(df)
        dfp['CDLSEPARATINGLINES'] = CDLSEPARATINGLINES(df)
        dfp['CDLSHOOTINGSTAR'] = CDLSHOOTINGSTAR(df)
        dfp['CDLSHORTLINE'] = CDLSHORTLINE(df)
        dfp['CDLSPINNINGTOP'] = CDLSPINNINGTOP(df)
        dfp['CDLSTALLEDPATTERN'] = CDLSTALLEDPATTERN(df)
        dfp['CDLSTICKSANDWICH'] = CDLSTICKSANDWICH(df)
        dfp['CDLTAKURI'] = CDLTAKURI(df)
        dfp['CDLTASUKIGAP'] = CDLTASUKIGAP(df)
        dfp['CDLTHRUSTING'] = CDLTHRUSTING(df)
        dfp['CDLTRISTAR'] = CDLTRISTAR(df)
        dfp['CDLUNIQUE3RIVER'] = CDLUNIQUE3RIVER(df)
        dfp['CDLUPSIDEGAP2CROWS'] = CDLUPSIDEGAP2CROWS(df)
        dfp['CDLXSIDEGAP3METHODS'] = CDLXSIDEGAP3METHODS(df)

#         dfp['AVGPRICE'] = AVGPRICE(df)
#         dfp['MEDPRICE'] = MEDPRICE(df)
#         dfp['TYPPRICE'] = TYPPRICE(df)
#         dfp['WCLPRICE'] = WCLPRICE(df)
# 
# #         dfp['ATR'] = ATR(df)
# #         dfp['NATR'] = NATR(df)
# #         dfp['TRANGE'] = TRANGE(df)
#         
#        dfp['AD'] = AD(df)
#        dfp['ADOSC'] = ADOSC(df)
#        dfp['OBV'] = OBV(df)
        
        forecast_col = 'PCT_change1'
        dfp.fillna(-99999, inplace=True)
        forecast_out = 1
        dfp['label'] = dfp[forecast_col].shift(-forecast_out)
        
        dfp.to_csv(directory + '/' + scrip + '.csv', encoding='utf-8')    
        #print 'XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX'
        #print '--------', scrip, '----------'
        dfp = dfp.ix[50:]
#         X = np.array(dfp.drop(['label'], 1))
#         X = preprocessing.scale(X)
#         X = X[:-forecast_out]
#         X_lately = X[-forecast_out:]
#         
#         #dfp.dropna(inplace=True)
#         y = np.array(dfp['label'])
#         
#         X_train, X_test, y_train, y_test = cross_validation.train_test_split(X, y, test_size=0.1)
#         clf = LinearRegression()
#         #clf = svm.SVR(kernel='poly')
#         clf.fit(X_train, y_train)
#         accuracy = clf.score(X_test, y_test)
#         forecast_set = clf.predict(X_lately)
#         print(scrip, accuracy, forecast_set)
#         #print 'XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX'    
        print(scrip) 
        regressionResult = performRegression(dfp, 0.95, scrip, directory, forecast_out)
        buy, sell = ta_lib_data(scrip)
        regressionResult.append(str(buy))
        regressionResult.append(str(sell))
        regressionResult.append(data['futures'])
        #ws = wb.active
        ws.append(regressionResult)
        trainSize = int(regressionResult[1][1:-1].split(',')[0])
        #ws_filter = wb.create_sheet("Filter")
        if((trainSize> 1000) and (float(regressionResult[3]) > 1) and (float(regressionResult[7]) > 1) and (float(regressionResult[9]) > 0) and (float(regressionResult[11]) > 0) and (float(regressionResult[13]) > 0)):
            ws_filter.append(regressionResult)
            
        if((trainSize> 1000) and (float(regressionResult[3]) > 1) and (float(regressionResult[5]) > 0) and (float(regressionResult[7]) > 1) and (float(regressionResult[9]) > 0) and (float(regressionResult[13]) > 0)):
            ws_filter.append(regressionResult)    
        
        if((trainSize> 1000) and (float(regressionResult[3]) < -1) and (float(regressionResult[5]) < 0) and (float(regressionResult[7]) < 0) and (float(regressionResult[9]) < 0) and (float(regressionResult[13]) < 0)):
            ws_filter.append(regressionResult)
            
            
        #ws_gtltzero = wb.create_sheet("FilterAllgtlt0")
        if((trainSize> 1000) and (float(regressionResult[3]) > 0) and (float(regressionResult[7]) > 0) and (float(regressionResult[9]) > 0) and (float(regressionResult[11]) > 0) and (float(regressionResult[13]) > 0)):
            ws_gtltzero.append(regressionResult)
            
        if((trainSize> 1000) and (float(regressionResult[3]) < 0) and (float(regressionResult[7]) < 0) and (float(regressionResult[9]) < 0) and (float(regressionResult[11]) < 0) and (float(regressionResult[13]) < 0)):
            ws_gtltzero.append(regressionResult)  
            

        #ws_RandomForest = wb.create_sheet("RandomForest")
        if((trainSize> 1000) and (float(regressionResult[3]) > 1 or float(regressionResult[3]) < -1)):
            ws_RandomForest.append(regressionResult)
            
        #ws_SVR = wb.create_sheet("SVR")
        if((trainSize> 1000) and (float(regressionResult[5]) > 1.5 or float(regressionResult[5]) < -1.5)):
            ws_SVR.append(regressionResult)
       
        #ws_Bagging = wb.create_sheet("Bagging")
        if((trainSize> 1000) and (float(regressionResult[7]) > 1.5 or float(regressionResult[7]) < -1.5)):
            ws_Bagging.append(regressionResult)
            
        #ws_AdaBoost = wb.create_sheet("AdaBoost")
        if((trainSize> 1000) and (float(regressionResult[9]) > 1.5 or float(regressionResult[9]) < -1.5)):
            ws_AdaBoost.append(regressionResult)
            
        #ws_KNeighbors = wb.create_sheet("KNeighbors")
        if((trainSize> 1000) and (float(regressionResult[11]) > 1.5 or float(regressionResult[11]) < -1.5)):
            ws_KNeighbors.append(regressionResult)
            
        #ws_GradientBoosting = wb.create_sheet("GradientBoosting")
        if((trainSize> 1000) and (float(regressionResult[13]) > 1.5 or float(regressionResult[13]) < -1.5)):
            ws_GradientBoosting.append(regressionResult)
                                           
def calculateParallel(threads=2):
    pool = ThreadPool(threads)
    
    scrips = []
    for data in db.scrip.find():
        scrips.append((data['scrip']).encode('UTF8').replace('&','').replace('-','_'))
    scrips.sort()
    
    pool.map(regression_ta_data, scrips)
                     
if __name__ == "__main__":
    if not os.path.exists(directory):
        os.makedirs(directory)
    calculateParallel(1)
    connection.close()
    saveReports()